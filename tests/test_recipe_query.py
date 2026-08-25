from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "services" / "backend-api"
sys.path.insert(0, str(BACKEND_ROOT))


class RecipeQueryTests(unittest.TestCase):
    def setUp(self) -> None:
        self._old_sys_path = list(sys.path)
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        backend_root = str(BACKEND_ROOT)
        sys.path[:] = [item for item in sys.path if item != backend_root]
        sys.path.insert(0, str(BACKEND_ROOT))
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = self._old_sys_path

    def test_locate_recipe_source_uses_latest_export_not_active_dir(self) -> None:
        # formula 始终用 worker 最新全量导出，不再被人工激活的「活动 BOM」目录劫持。
        import os
        from app.recipes.bom_query import locate_recipe_source
        active_dir = self.tmp_path / "active"
        input_dir = self.tmp_path / "export"
        active_dir.mkdir()
        input_dir.mkdir()
        (active_dir / "bom_20260101_000000.xlsx").write_bytes(b"x")  # 旧的人工激活文件
        (input_dir / "bom_20260624_044650.xlsx").write_bytes(b"x")   # worker 最新导出
        saved = {k: os.environ.get(k) for k in
                 ("RECIPE_BOM_INPUT_PATH", "RECIPE_ACTIVE_BOM_DIR", "RECIPE_BOM_INPUT_DIR")}
        os.environ.pop("RECIPE_BOM_INPUT_PATH", None)
        os.environ["RECIPE_ACTIVE_BOM_DIR"] = str(active_dir)
        os.environ["RECIPE_BOM_INPUT_DIR"] = str(input_dir)
        try:
            result = locate_recipe_source()
        finally:
            for key, value in saved.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
        self.assertEqual("bom_20260624_044650.xlsx", result.name)

    def _write_source_workbook(self) -> Path:
        source = self.tmp_path / "bom_20260604.xlsx"
        material_rows = [
            {
                "父件编码": "30122027-3027",
                "父件名称": "BX3027-海尔洗衣机PP海灰色母",
                "规格型号": "PP",
                "版本号": "V1",
                "计量单位": "kg",
                "生产数量": 25,
                "默认BOM": 1,
                "停用": 0,
            },
            {
                "父件编码": "30122027-3027",
                "父件名称": "BX3027-停用旧版",
                "规格型号": "PP",
                "版本号": "V0",
                "计量单位": "kg",
                "生产数量": 25,
                "默认BOM": 0,
                "停用": 1,
            },
        ]
        child_rows = [
            {
                "版本号": "V1",
                "父件编码": "30122027-3027",
                "子件编码": "C001",
                "子件名称": "树脂",
                "规格型号": "A",
                "计量单位": "kg",
                "需用数量": 2,
            },
            {
                "版本号": "V1",
                "父件编码": "30122027-3027",
                "子件编码": "C002",
                "子件名称": "色粉",
                "规格型号": "B",
                "计量单位": "g",
                "需用数量": 500,
            },
            {
                "版本号": "V0",
                "父件编码": "30122027-3027",
                "子件编码": "C003",
                "子件名称": "旧版树脂",
                "规格型号": "C",
                "计量单位": "kg",
                "需用数量": 1,
            },
        ]
        with pd.ExcelWriter(source, engine="openpyxl") as writer:
            pd.DataFrame(material_rows).to_excel(writer, sheet_name="物料清单", index=False)
            pd.DataFrame(child_rows).to_excel(writer, sheet_name="子件明细", index=False)
        return source

    def test_detail_parse_is_cached_by_content_and_invalidated_on_change(self) -> None:
        from app.recipes.bom_query import file_content_signature, load_detail_from_workbook

        source = self._write_source_workbook()
        first = load_detail_from_workbook(source)
        second = load_detail_from_workbook(source)
        self.assertIs(first, second)  # 同内容命中缓存，不重复解析

        sig_before = file_content_signature(source)
        extra = self.tmp_path / "bom_changed.xlsx"
        with pd.ExcelWriter(extra, engine="openpyxl") as writer:
            pd.DataFrame(
                [{"父件编码": "X-1", "父件名称": "新件", "规格型号": "PP", "版本号": "V1", "计量单位": "kg", "生产数量": 1, "默认BOM": 1, "停用": 0}]
            ).to_excel(writer, sheet_name="物料清单", index=False)
            pd.DataFrame(
                [{"版本号": "V1", "父件编码": "X-1", "子件编码": "C9", "子件名称": "料", "规格型号": "A", "计量单位": "kg", "需用数量": 1}]
            ).to_excel(writer, sheet_name="子件明细", index=False)
        self.assertNotEqual(sig_before, file_content_signature(extra))  # 内容不同签名不同
        third = load_detail_from_workbook(extra)
        self.assertIsNot(third, first)  # 内容变化触发重新解析

    def test_query_merges_tplus_bom_workbook_and_keeps_disabled_versions(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        result = query_recipe_workbook(self._write_source_workbook(), query_text="3027", default_bom="all")

        self.assertEqual(3, result.match_count)
        self.assertEqual(2, result.recipe_count)
        self.assertEqual({"V0", "V1"}, set(result.detail["版本号_子件"].astype(str)))
        disabled = result.detail[result.detail["版本号_子件"] == "V0"].iloc[0]
        self.assertEqual("1", str(disabled["停用"]))

    def test_query_can_filter_to_default_bom_only(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        result = query_recipe_workbook(self._write_source_workbook(), query_text="3027", default_bom="1")

        self.assertEqual(2, result.match_count)
        self.assertEqual(1, result.recipe_count)
        self.assertEqual({"V1"}, set(result.detail["版本号_子件"].astype(str)))

    def test_query_preserves_leading_zero_codes_when_excel_infers_numeric_text(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_leading_zero.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["0003027", "BX3027-前导零配方", "PP", "V001", "kg", 25, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量"])
        component.append(["V001", "0003027", "03000012", "跳过项", "A", "kg", 2])
        component.append(["V001", "0003027", "0000008", "有效项", "B", "kg", 8])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="3027", default_bom="1")

        self.assertEqual(2, result.match_count)
        self.assertEqual({"0003027"}, set(result.detail["父件编码"]))
        self.assertEqual({"03000012", "0000008"}, set(result.detail["子件编码"]))
        skipped = result.detail[result.detail["子件编码"] == "03000012"].iloc[0]
        included = result.detail[result.detail["子件编码"] == "0000008"].iloc[0]
        self.assertTrue(pd.isna(skipped["比例"]))
        self.assertEqual(1.0, included["比例"])

    def test_ratio_excludes_packaging_by_unit_even_when_code_not_in_skip_list(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_unit_exclude.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["HYD-0601", "HYD-0601阻燃ABSPVC改性料", "ABS", "V1", "kg", 25, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量"])
        component.append(["V1", "HYD-0601", "NEWBAG01", "新包装袋(不在排除编码里)", "P", "条", 40])
        component.append(["V1", "HYD-0601", "NEWBOX01", "新包装件", "P", "个", 2])
        component.append(["V1", "HYD-0601", "C100", "树脂", "A", "kg", 20])
        component.append(["V1", "HYD-0601", "C200", "色粉", "B", "g", 5000])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="HYD-0601", default_bom="1")

        detail = result.detail.set_index("子件编码")
        self.assertTrue(pd.isna(detail.loc["NEWBAG01", "比例"]))
        self.assertTrue(pd.isna(detail.loc["NEWBOX01", "比例"]))
        # 分母只剩 20kg + 5kg(5000g) = 25kg
        self.assertAlmostEqual(0.8, float(detail.loc["C100", "比例"]))
        self.assertAlmostEqual(0.2, float(detail.loc["C200", "比例"]))

    def test_ratio_excludes_fen_unit_color_powder_pack(self) -> None:
        # 复刻生产上那本含色粉包的 BOM 的形状（编码/物料名用占位，本仓 PUBLIC 不落配方数据）：
        # 色粉包单位「份」不在旧黑名单里，被错误算进分母（101.5，应为 100.5）。
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_fen_unit.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["FEN-DEMO", "含色粉包成品", "X", "V1", "kg", 100, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量"])
        component.append(["V1", "FEN-DEMO", "MAT-A", "主料A", "A", "kg", 60])
        component.append(["V1", "FEN-DEMO", "MAT-B", "填料B", "B", "kg", 30])
        component.append(["V1", "FEN-DEMO", "MAT-C", "助剂C", "C", "kg", 10.5])
        component.append(["V1", "FEN-DEMO", "PACK-FEN", "色粉包", "D", "份", 1])
        component.append(["V1", "FEN-DEMO", "PACK-BAG", "包装袋", "E", "条", 4])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="FEN-DEMO", default_bom="1")

        detail = result.detail.set_index("子件编码")
        self.assertTrue(pd.isna(detail.loc["PACK-FEN", "比例"]))
        self.assertTrue(pd.isna(detail.loc["PACK-BAG", "比例"]))
        # 分母 = 60 + 30 + 10.5 = 100.5（不含色粉包的 1 份）
        self.assertAlmostEqual(60 / 100.5, float(detail.loc["MAT-A", "比例"]), places=9)
        self.assertAlmostEqual(30 / 100.5, float(detail.loc["MAT-B", "比例"]), places=9)
        self.assertAlmostEqual(1.0, float(detail.loc[["MAT-A", "MAT-B", "MAT-C"], "比例"].sum()), places=9)
        # 反证锚点：旧黑名单判据（{"条","个"}）会把色粉包算进分母得到 101.5；本行确保新测试真挡得住该 bug。
        self.assertNotAlmostEqual(60 / 101.5, float(detail.loc["MAT-A", "比例"]), places=6)

    def test_ratio_denominator_accepts_mass_unit_variants_only(self) -> None:
        # 白名单判据：质量单位的各种写法都进分母，其余（含「吨」——没有换算因子）一律排除。
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_unit_variants.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["UNITMIX", "单位变体", "X", "V1", "kg", 10, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量"])
        for code, unit, qty in [
            ("M_KG", "kg", 1), ("M_KGUP", "KG", 1), ("M_QIANKE", "千克", 1),
            ("M_GONGJIN", "公斤", 1), ("M_KE", "克", 1000), ("M_G", "g", 1000),
            ("X_DUN", "吨", 1), ("X_BAO", "包", 1), ("X_ZHI", "只", 1), ("X_PCS", "PCS", 1),
        ]:
            component.append(["V1", "UNITMIX", code, code, "S", unit, qty])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="UNITMIX", default_bom="1")

        detail = result.detail.set_index("子件编码")
        for code in ["M_KG", "M_KGUP", "M_QIANKE", "M_GONGJIN", "M_KE", "M_G"]:
            self.assertAlmostEqual(1 / 6, float(detail.loc[code, "比例"]), places=9, msg=code)
        for code in ["X_DUN", "X_BAO", "X_ZHI", "X_PCS"]:
            self.assertTrue(pd.isna(detail.loc[code, "比例"]), msg=code)

    def test_ratio_skips_unit_filter_when_unit_column_missing(self) -> None:
        # 判据反转会翻转降级方向：缺「计量单位_子件」列时必须不按单位过滤，否则整批比例会变空。
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_no_unit_column.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "生产数量", "默认BOM", "停用"])
        material.append(["NOUNIT", "无单位列", "X", "V1", 10, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "需用数量"])
        component.append(["V1", "NOUNIT", "A1", "料A", "S", 30])
        component.append(["V1", "NOUNIT", "A2", "料B", "S", 70])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="NOUNIT", default_bom="1")

        detail = result.detail.set_index("子件编码")
        self.assertAlmostEqual(0.3, float(detail.loc["A1", "比例"]), places=9)
        self.assertAlmostEqual(0.7, float(detail.loc["A2", "比例"]), places=9)

    def test_ratio_degrades_when_no_mass_unit_present_at_all(self) -> None:
        # 兜底：整列没有一个质量单位＝单位列大概率解析异常，此时按单位过滤会让全部比例变空，降级为不过滤。
        from app.recipes.bom_query import query_recipe_workbook

        source = self.tmp_path / "bom_all_non_mass.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["ALLBAD", "单位异常", "X", "V1", "kg", 10, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量"])
        component.append(["V1", "ALLBAD", "B1", "料A", "S", "KGS", 30])
        component.append(["V1", "ALLBAD", "B2", "料B", "S", "KGS", 70])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="ALLBAD", default_bom="1")

        detail = result.detail.set_index("子件编码")
        self.assertAlmostEqual(0.3, float(detail.loc["B1", "比例"]), places=9)
        self.assertAlmostEqual(0.7, float(detail.loc["B2", "比例"]), places=9)

    def test_simulated_quantities_recompute_ratio_without_packaging_units(self) -> None:
        from app.recipes.bom_query import calculate_recipe_costs, query_recipe_workbook

        source = self.tmp_path / "bom_simulated_quantity.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["HYD-0601", "HYD-0601阻燃ABSPVC改性料", "ABS", "V1", "kg", 25, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量", "系统单价"])
        component.append(["V1", "HYD-0601", "NEWBAG01", "新包装袋", "P", "条", 40, 1])
        component.append(["V1", "HYD-0601", "C100", "树脂", "A", "kg", 20, 3])
        component.append(["V1", "HYD-0601", "C200", "色粉", "B", "g", 5000, 4])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="HYD-0601", default_bom="1")
        recipes = calculate_recipe_costs(
            result,
            simulated_quantities={"NEWBAG01": 80, "C100": 10, "C200": 15000},
        )

        self.assertEqual(1, len(recipes))
        recipe = recipes[0]
        by_code = {line["child_code"]: line for line in recipe["lines"]}
        self.assertEqual(80.0, by_code["NEWBAG01"]["simulated_quantity"])
        self.assertEqual(0.0, by_code["NEWBAG01"]["simulated_ratio"])
        self.assertEqual(0.0, by_code["NEWBAG01"]["simulated_amount"])
        # 模拟分母只剩 10kg + 15kg(15000g) = 25kg。
        self.assertAlmostEqual(0.4, by_code["C100"]["simulated_ratio"])
        self.assertAlmostEqual(0.6, by_code["C200"]["simulated_ratio"])
        self.assertAlmostEqual(1.2, by_code["C100"]["simulated_amount"])
        self.assertAlmostEqual(2.4, by_code["C200"]["simulated_amount"])
        self.assertAlmostEqual(3.6, recipe["simulated_total"])

    def test_cost_uses_fallback_price_column_as_system_price(self) -> None:
        from app.recipes.bom_query import calculate_recipe_costs, query_recipe_workbook

        source = self.tmp_path / "bom_fallback_price.xlsx"
        wb = Workbook()
        material = wb.active
        material.title = "物料清单"
        material.append(["父件编码", "父件名称", "规格型号", "版本号", "计量单位", "生产数量", "默认BOM", "停用"])
        material.append(["HYD-0601", "HYD-0601阻燃ABSPVC改性料", "ABS", "V1", "kg", 25, 1, 0])
        component = wb.create_sheet("子件明细")
        component.append(["版本号", "父件编码", "子件编码", "子件名称", "规格型号", "计量单位", "需用数量", "材料单价"])
        component.append(["V1", "HYD-0601", "C100", "树脂", "A", "kg", 4, 12.5])
        wb.save(source)

        result = query_recipe_workbook(source, query_text="HYD-0601", default_bom="1")
        recipes = calculate_recipe_costs(result)

        self.assertEqual(1, len(recipes))
        line = recipes[0]["lines"][0]
        self.assertAlmostEqual(12.5, line["system_price"])
        self.assertAlmostEqual(12.5, line["current_price"])
        self.assertAlmostEqual(12.5, recipes[0]["system_total"])
        self.assertAlmostEqual(12.5, recipes[0]["current_total"])

    def test_concurrent_load_parses_only_once(self) -> None:
        # 并发 query+cost 命中同一文件时，加锁后只解析一次、共享同一份结果（防冷解析双倍 CPU/504）。
        import threading
        import time
        from app.recipes import bom_query as bq
        source = self._write_source_workbook()
        bq._DETAIL_CACHE.clear()
        calls = {"n": 0}
        real = bq._load_detail_from_workbook_uncached

        def counting(path):
            calls["n"] += 1
            time.sleep(0.3)  # 拖长解析，制造并发窗口
            return real(path)

        bq._load_detail_from_workbook_uncached = counting
        results: list = []
        try:
            threads = [
                threading.Thread(target=lambda: results.append(bq.load_detail_from_workbook(source)))
                for _ in range(4)
            ]
            for t in threads:
                t.start()
            for t in threads:
                t.join()
        finally:
            bq._load_detail_from_workbook_uncached = real
        self.assertEqual(1, calls["n"])  # 4 个并发只触发一次解析
        self.assertTrue(results and all(r is results[0] for r in results))

    def test_raw_export_filename_includes_codes_space_joined(self) -> None:
        from app.recipes.bom_query import recipe_raw_export_filename
        from datetime import datetime
        now = datetime(2026, 6, 28, 9, 30, 0)
        # 单编码
        self.assertEqual("配方查询_6800_20260628-093000.xlsx", recipe_raw_export_filename("6800", now=now))
        # 多编码：原分隔符(逗号/中文逗号/顿号)→空格
        self.assertEqual("配方查询_4791 4588_20260628-093000.xlsx", recipe_raw_export_filename("4791,4588", now=now))
        self.assertEqual("配方查询_4791 4588 4197_20260628-093000.xlsx", recipe_raw_export_filename("4791，4588、4197", now=now))
        # 空查询 → 全部
        self.assertEqual("配方查询_全部_20260628-093000.xlsx", recipe_raw_export_filename("", now=now))

    def test_float_or_zero_coercion_contract(self) -> None:
        # 锁定 _float_or_zero 的数值强转语义（重构掉 per-scalar pd.Series 后必须逐一致）。
        from app.recipes.bom_query import _float_or_zero
        cases = [
            (5, 5.0), (5.5, 5.5), ("5", 5.0), ("5.5", 5.5), ("  7 ", 7.0),
            (None, 0.0), ("", 0.0), ("abc", 0.0), ("1,234", 0.0), (True, 1.0),
        ]
        for value, expected in cases:
            self.assertEqual(expected, _float_or_zero(value), msg=repr(value))
        self.assertEqual(0.0, _float_or_zero(float("nan")))
        self.assertEqual(0.0, _float_or_zero(pd.NA))

    def test_cost_quantity_unit_and_coercion_contract(self) -> None:
        # 克类单位 /1000、其余原值；无法解析→0。
        from app.recipes.bom_query import _cost_quantity
        self.assertAlmostEqual(0.005, _cost_quantity(5, "克"))
        self.assertAlmostEqual(0.5, _cost_quantity(500, "g"))
        self.assertAlmostEqual(0.005, _cost_quantity(5, " G "))
        self.assertAlmostEqual(5.0, _cost_quantity(5, "kg"))
        self.assertAlmostEqual(0.0025, _cost_quantity("2.5", "g"))
        self.assertEqual(0.0, _cost_quantity(None, "克"))
        self.assertEqual(0.0, _cost_quantity("abc", "kg"))

    def test_save_recipe_workbook_creates_human_review_and_matrix_sheets(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook, save_recipe_workbook

        result = query_recipe_workbook(self._write_source_workbook(), query_text="3027", default_bom="1")
        output_path = self.tmp_path / "recipe.xlsx"

        save_recipe_workbook(output_path, result)

        wb = load_workbook(output_path, data_only=False)
        self.assertIn("配方表_人眼版", wb.sheetnames)
        self.assertIn("横向对比_矩阵", wb.sheetnames)
        self.assertIn("父件子件明细_提取", wb.sheetnames)
        self.assertIn("版本父件分组合计", wb.sheetnames)
        self.assertEqual("物料清单配方表", wb["配方表_人眼版"]["A1"].value)


class ChildReverseLookupTests(unittest.TestCase):
    """按子件反查配方：候选罗列 + 勾选后取整本配方。"""

    def setUp(self) -> None:
        self._old_sys_path = list(sys.path)
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        backend_root = str(BACKEND_ROOT)
        sys.path[:] = [item for item in sys.path if item != backend_root]
        sys.path.insert(0, str(BACKEND_ROOT))
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = self._old_sys_path

    def _write_source_workbook(self) -> Path:
        source = self.tmp_path / "bom_20260824.xlsx"
        # HYD-419 与 HYD-4197 互为前缀；C001 与 C0011 同理。两组都是用来钉死「精确匹配」的。
        material_rows = [
            {"父件编码": code, "父件名称": name, "规格型号": "PP", "版本号": version,
             "计量单位": "kg", "生产数量": 100, "默认BOM": default_bom, "停用": disabled}
            for code, name, version, default_bom, disabled in [
                ("HYD-419", "甲料", "V1", 1, 0),
                ("HYD-4197", "乙料", "V1", 1, 0),
                ("HYD-4200", "丙料", "V1", 1, 0),
                ("HYD-4300", "丁料", "V1", 1, 1),
            ]
        ]
        child_rows = [
            {"版本号": version, "父件编码": parent, "子件编码": code, "子件名称": name,
             "规格型号": "S", "计量单位": "kg", "需用数量": qty}
            for parent, version, code, name, qty in [
                ("HYD-419", "V1", "C001", "硬脂酸", 2),
                ("HYD-419", "V1", "C002", "树脂", 8),
                ("HYD-4197", "V1", "C001", "硬脂酸", 1),
                ("HYD-4197", "V1", "C003", "硬脂酸锌", 3),
                ("HYD-4200", "V1", "C0011", "硬脂酸钙", 5),
                ("HYD-4300", "V1", "C001", "硬脂酸", 4),
            ]
        ]
        with pd.ExcelWriter(source, engine="openpyxl") as writer:
            pd.DataFrame(material_rows).to_excel(writer, sheet_name="物料清单", index=False)
            pd.DataFrame(child_rows).to_excel(writer, sheet_name="子件明细", index=False)
        return source

    def test_search_lists_candidates_with_recipe_counts_sorted_by_usage(self) -> None:
        from app.recipes.bom_query import search_child_items

        items, truncated = search_child_items(self._write_source_workbook(), keyword="硬脂酸")

        self.assertFalse(truncated)
        by_code = {item["child_code"]: item for item in items}
        self.assertEqual({"C001", "C003", "C0011"}, set(by_code))
        # C001 被 HYD-419 / HYD-4197 / HYD-4300 三本用到
        self.assertEqual(3, by_code["C001"]["recipe_count"])
        self.assertEqual(1, by_code["C003"]["recipe_count"])
        self.assertEqual("硬脂酸钙", by_code["C0011"]["child_name"])
        # 用量最多的排最前，方便用户先看主料
        self.assertEqual("C001", items[0]["child_code"])

    def test_search_matches_child_code_and_respects_scope_filters(self) -> None:
        from app.recipes.bom_query import search_child_items

        source = self._write_source_workbook()
        by_code, _ = search_child_items(source, keyword="C003")
        self.assertEqual(["C003"], [item["child_code"] for item in by_code])

        # 停用的 HYD-4300 排除后，C001 只剩两本配方
        items, _ = search_child_items(source, keyword="硬脂酸", include_disabled=False)
        counts = {item["child_code"]: item["recipe_count"] for item in items}
        self.assertEqual(2, counts["C001"])

    def test_reverse_lookup_returns_whole_recipes_not_only_matched_lines(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        result = query_recipe_workbook(
            self._write_source_workbook(), query_text="硬脂酸", child_codes=["C001"],
        )

        self.assertEqual({"HYD-419", "HYD-4197", "HYD-4300"}, set(result.detail["父件编码"]))
        # 命中的是配方，不是行：同一本配方里没命中的子件也要在结果里，否则对比和成本核算都是残的
        self.assertIn("C002", set(result.detail["子件编码"]))
        self.assertEqual(["C001"], result.codes)

    def test_reverse_lookup_matches_child_code_exactly_not_as_prefix(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        result = query_recipe_workbook(
            self._write_source_workbook(), query_text="硬脂酸", child_codes=["C001"],
        )

        # C0011 只在 HYD-4200 里；子串匹配会把它一起捞进来，精确匹配不会
        self.assertNotIn("HYD-4200", set(result.detail["父件编码"]))

    def test_reverse_lookup_all_requires_every_selected_child(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        source = self._write_source_workbook()
        any_result = query_recipe_workbook(source, query_text="x", child_codes=["C001", "C003"], child_match="any")
        all_result = query_recipe_workbook(source, query_text="x", child_codes=["C001", "C003"], child_match="all")

        self.assertEqual({"HYD-419", "HYD-4197", "HYD-4300"}, set(any_result.detail["父件编码"]))
        # 同时含 C001 和 C003 的只有 HYD-4197
        self.assertEqual({"HYD-4197"}, set(all_result.detail["父件编码"]))

    def test_reverse_lookup_ignores_query_text_and_returns_empty_for_unknown_child(self) -> None:
        from app.recipes.bom_query import query_recipe_workbook

        source = self._write_source_workbook()
        # query_text 在反查模式下只作展示：即使它能匹配到父件，也不该混进结果
        result = query_recipe_workbook(source, query_text="HYD-419", child_codes=["C999"])

        self.assertTrue(result.detail.empty)
        self.assertEqual(0, result.recipe_count)


if __name__ == "__main__":
    unittest.main()
