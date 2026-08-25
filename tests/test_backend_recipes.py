from __future__ import annotations

from io import BytesIO
import os
import sys
import tempfile
import time
import unittest
import urllib.parse
from pathlib import Path

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "services" / "backend-api"


class BackendRecipeRouteTests(unittest.TestCase):
    def setUp(self) -> None:
        self._old_sys_path = list(sys.path)
        self._old_env = {
            "AUTH_TOKEN_SECRET": os.environ.get("AUTH_TOKEN_SECRET"),
            "RECIPE_BOM_INPUT_PATH": os.environ.get("RECIPE_BOM_INPUT_PATH"),
            "RECIPE_EXPORT_DIR": os.environ.get("RECIPE_EXPORT_DIR"),
            "TPLUS_BOM_SYNC_REQUEST_DIR": os.environ.get("TPLUS_BOM_SYNC_REQUEST_DIR"),
        }
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        backend_root = str(BACKEND_ROOT)
        sys.path[:] = [item for item in sys.path if item != backend_root]
        sys.path.insert(0, backend_root)

        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)
        self.source_path = self._write_source_workbook()
        os.environ["AUTH_TOKEN_SECRET"] = "test-recipe-secret"
        os.environ["RECIPE_BOM_INPUT_PATH"] = str(self.source_path)
        os.environ["RECIPE_EXPORT_DIR"] = str(self.tmp_path / "exports")
        os.environ["TPLUS_BOM_SYNC_REQUEST_DIR"] = str(self.tmp_path / "tplus-sync-requests")

        from app.core import _encode_token
        from app.main import app

        self._encode_token = _encode_token
        self.client = TestClient(app)

    def tearDown(self) -> None:
        self._tmp.cleanup()
        for key, value in self._old_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = self._old_sys_path

    def test_warm_recipe_caches_populates_detail_cache(self) -> None:
        # 后台预热：解析进缓存，使用户请求永不撞冷解析；价格表缺失时优雅跳过、不抛异常。
        from app.routers.recipes import warm_recipe_caches
        from app.recipes import bom_query as bq

        bq._DETAIL_CACHE.clear()
        warm_recipe_caches()
        self.assertTrue(bq._DETAIL_CACHE)

    def _write_source_workbook(self) -> Path:
        source = self.tmp_path / "bom_20260604.xlsx"
        material_rows = [
            {
                "父件编码": "30122027-3027",
                "父件名称": "BX3027-海尔洗衣机PP海灰色母",
                "规格型号": "PP",
                "版本号": "V1",
                "计量单位": "kg",
                "生产数量": 25,
                "默认BOM": 1,
                "停用": 0,
            },
            {
                "父件编码": "30122027-3027",
                "父件名称": "BX3027-停用旧版",
                "规格型号": "PP",
                "版本号": "V0",
                "计量单位": "kg",
                "生产数量": 25,
                "默认BOM": 0,
                "停用": 1,
            },
        ]
        child_rows = [
            {
                "版本号": "V1",
                "父件编码": "30122027-3027",
                "子件编码": "C001",
                "子件名称": "树脂",
                "规格型号": "A",
                "计量单位": "kg",
                "需用数量": 2,
                "系统单价": 10,
            },
            {
                "版本号": "V1",
                "父件编码": "30122027-3027",
                "子件编码": "C002",
                "子件名称": "色粉",
                "规格型号": "B",
                "计量单位": "g",
                "需用数量": 500,
                "系统单价": 20,
            },
            {
                "版本号": "V0",
                "父件编码": "30122027-3027",
                "子件编码": "C003",
                "子件名称": "旧版树脂",
                "规格型号": "C",
                "计量单位": "kg",
                "需用数量": 1,
                "系统单价": 30,
            },
        ]
        with pd.ExcelWriter(source, engine="openpyxl") as writer:
            pd.DataFrame(material_rows).to_excel(writer, sheet_name="物料清单", index=False)
            pd.DataFrame(child_rows).to_excel(writer, sheet_name="子件明细", index=False)
        return source

    def _token(self, *, roles: list[str] | None = None, permissions: list[str] | None = None) -> str:
        return self._encode_token(
            {
                "sub": "recipe-user",
                "roles": roles or [],
                "permissions": permissions or [],
                "exp": int(time.time()) + 3600,
            }
        )

    def test_formula_read_user_can_query_recipe_and_download_workbook(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/query",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027"},
        )

        self.assertEqual(200, response.status_code)
        data = response.json()
        self.assertEqual(3, data["match_count"])
        self.assertEqual(2, data["recipe_count"])
        self.assertIn("/v1/recipes/download/", data["download_url"])
        self.assertTrue(data["preview"])

        download = self.client.get(data["download_url"], headers={"Authorization": f"Bearer {token}"})

        self.assertEqual(200, download.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download.headers["content-type"],
        )
        self.assertGreater(len(download.content), 0)

    def test_children_search_lists_candidates_for_reverse_lookup(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/children/search",
            headers={"Authorization": f"Bearer {token}"},
            json={"keyword": "树脂"},
        )

        self.assertEqual(200, response.status_code)
        data = response.json()
        self.assertFalse(data["truncated"])
        codes = {item["child_code"]: item for item in data["items"]}
        self.assertEqual({"C001", "C003"}, set(codes))
        self.assertEqual(1, codes["C001"]["recipe_count"])

    def test_children_search_requires_formula_read_permission(self) -> None:
        response = self.client.post(
            "/v1/recipes/children/search",
            headers={"Authorization": f"Bearer {self._token()}"},
            json={"keyword": "树脂"},
        )

        self.assertEqual(403, response.status_code)

    def test_query_with_child_codes_returns_whole_recipes_and_survives_download(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/query",
            headers={"Authorization": f"Bearer {token}"},
            # query 在反查模式下只作展示：这里故意填一个匹配不到任何父件的关键字
            json={"query": "树脂", "child_codes": ["C001"], "child_match": "any"},
        )

        self.assertEqual(200, response.status_code)
        data = response.json()
        self.assertEqual(1, data["recipe_count"])
        # 整本配方都要在结果里，不能只剩命中行
        self.assertEqual({"C001", "C002"}, {row["子件编码"] for row in data["preview"]})

        # 延迟生成的导出文件必须沿用同一套反查参数，否则下载下来是空表
        download = self.client.get(data["download_url"], headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(200, download.status_code)
        self.assertGreater(len(download.content), 0)

    def test_download_with_sheet_human_returns_single_human_sheet(self) -> None:
        token = self._token(permissions=["formula.read"])
        data = self.client.post(
            "/v1/recipes/query",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027"},
        ).json()

        download = self.client.get(
            f"{data['download_url']}?sheet=human",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(200, download.status_code)
        workbook = load_workbook(BytesIO(download.content))
        self.assertEqual(["配方表_人眼版"], workbook.sheetnames)
        self.assertIn(urllib.parse.quote("人眼版"), download.headers.get("content-disposition", ""))

        # 全量下载不受 human 变体缓存影响，仍是多工作表
        full = self.client.get(data["download_url"], headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(200, full.status_code)
        full_workbook = load_workbook(BytesIO(full.content))
        self.assertIn("配方表_人眼版", full_workbook.sheetnames)
        self.assertGreater(len(full_workbook.sheetnames), 1)

    def test_download_rejects_unknown_sheet_param(self) -> None:
        token = self._token(permissions=["formula.read"])
        response = self.client.get(
            "/v1/recipes/download/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa?sheet=matrix",
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(400, response.status_code)

    def _compare_export_payload(self) -> dict:
        return {
            "query": "3027",
            "filter_label": "全部",
            "view": {"spec": True, "qty": True, "pct": True, "arrow": True, "delta": True, "newTag": True, "bar": True},
            "versions": [
                {"label": "A｜V1", "code": "A", "version": "V1", "is_base": True, "is_target": False},
                {"label": "B｜V2", "code": "B", "version": "V2", "is_base": False, "is_target": True},
            ],
            "rows": [
                {
                    "status": "change",
                    "item_code": "C001",
                    "item_name": "树脂",
                    "spec": "PP",
                    "unit": "kg",
                    "code_warn": False,
                    "cells": [
                        {"ratio": 0.994, "qty": 100, "delta": None, "is_new": False},
                        {"ratio": 0.913, "qty": 91.5, "delta": -0.081, "is_new": False},
                    ],
                },
                {
                    "status": "add",
                    "item_code": "C002",
                    "item_name": "色粉",
                    "spec": "",
                    "unit": "kg",
                    "code_warn": True,
                    "cells": [None, {"ratio": 0.087, "qty": 8.5, "delta": None, "is_new": True}],
                },
            ],
        }

    def test_compare_export_returns_styled_real_xlsx(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/compare/export",
            headers={"Authorization": f"Bearer {token}"},
            json=self._compare_export_payload(),
        )

        self.assertEqual(200, response.status_code)
        self.assertIn(
            urllib.parse.quote("配方比例对比表"),
            response.headers.get("content-disposition", ""),
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["比例对比"]
        self.assertEqual("配方比例对比表", sheet["A1"].value)
        # 表头深棕填充（伪xls时代 head CSS 被 Excel 忽略的问题，真 xlsx 彻底解决）
        self.assertEqual("FF4A3C28", sheet.cell(row=4, column=1).fill.start_color.rgb)
        # 状态列彩色填充（第5行=change 橙）
        self.assertEqual("FFFFEDD5", sheet.cell(row=5, column=1).fill.start_color.rgb)
        # 缺失格「—」灰底；冻结窗格生效
        self.assertEqual("—", sheet.cell(row=6, column=6).value)
        self.assertEqual("F5", sheet.freeze_panes)

    def test_compare_export_rejects_mismatched_cells(self) -> None:
        token = self._token(permissions=["formula.read"])
        payload = self._compare_export_payload()
        payload["rows"][0]["cells"] = [None]

        response = self.client.post(
            "/v1/recipes/compare/export",
            headers={"Authorization": f"Bearer {token}"},
            json=payload,
        )

        self.assertEqual(400, response.status_code)

    def test_compare_export_requires_login(self) -> None:
        response = self.client.post("/v1/recipes/compare/export", json=self._compare_export_payload())
        self.assertIn(response.status_code, {401, 403})

    def test_admin_user_can_query_recipe(self) -> None:
        token = self._token(roles=["admin"], permissions=["admin.access"])

        response = self.client.post(
            "/v1/recipes/query",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027", "default_bom": "1"},
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(2, response.json()["match_count"])

    def test_query_requires_bearer_token(self) -> None:
        response = self.client.post("/v1/recipes/query", json={"query": "3027"})

        self.assertIn(response.status_code, {401, 403})

    def test_query_missing_source_returns_404_without_local_path(self) -> None:
        missing_path = self.tmp_path / "missing" / "bom.xlsx"
        os.environ["RECIPE_BOM_INPUT_PATH"] = str(missing_path)
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/query",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027"},
        )

        self.assertEqual(404, response.status_code)
        detail = response.json()["detail"]
        self.assertEqual("BOM 输入文件未找到", detail)
        self.assertNotIn(str(missing_path), detail)

    def test_download_requires_bearer_token(self) -> None:
        response = self.client.get("/v1/recipes/download/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")

        self.assertIn(response.status_code, {401, 403})

    def test_download_requires_formula_read_permission(self) -> None:
        token = self._token(permissions=["production.schedule.read"])

        response = self.client.get(
            "/v1/recipes/download/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(403, response.status_code)

    def test_download_rejects_invalid_file_id(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.get(
            "/v1/recipes/download/not-a-valid-file-id",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(400, response.status_code)

    def test_download_returns_404_for_missing_file(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.get(
            "/v1/recipes/download/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(404, response.status_code)

    def test_formula_read_user_can_request_manual_bom_sync(self) -> None:
        import json

        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/sync-bom",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(200, response.status_code)
        data = response.json()
        self.assertEqual("pending", data["status"])
        self.assertEqual("bom", data["module"])
        self.assertTrue(data["include_disabled"])
        request_files = list((self.tmp_path / "tplus-sync-requests").glob("*.json"))
        self.assertEqual(1, len(request_files))
        payload = json.loads(request_files[0].read_text(encoding="utf-8"))
        self.assertEqual("bom", payload["module"])
        self.assertEqual("manual_bom_full_include_disabled", payload["mode"])
        self.assertTrue(payload["include_disabled"])

    def test_manual_bom_sync_requires_formula_read_permission(self) -> None:
        token = self._token(permissions=["production.schedule.read"])

        response = self.client.post(
            "/v1/recipes/sync-bom",
            headers={"Authorization": f"Bearer {token}"},
        )

        self.assertEqual(403, response.status_code)

    def test_formula_cost_user_can_calculate_recipe_cost_with_manual_prices_and_simulation(self) -> None:
        token = self._token(permissions=["formula.cost.calculate"])

        response = self.client.post(
            "/v1/recipes/cost",
            headers={"Authorization": f"Bearer {token}"},
            json={
                "query": "3027",
                "default_bom": "1",
                "manual_prices": {"C002": 25},
                "simulated_quantities": {"C001": 1, "C002": 1000},
            },
        )

        self.assertEqual(200, response.status_code)
        data = response.json()
        self.assertEqual(1, data["recipe_count"])
        recipe = data["recipes"][0]
        self.assertEqual("V1", recipe["version"])
        # 分价 = 比例 × 单价：C001 比例 0.8、C002 比例 0.2
        self.assertAlmostEqual(12.0, recipe["system_total"])   # 0.8*10 + 0.2*20
        self.assertAlmostEqual(13.0, recipe["current_total"])  # 0.8*10 + 0.2*25
        by_code = {line["child_code"]: line for line in recipe["lines"]}
        self.assertAlmostEqual(20.0, by_code["C002"]["system_price"])
        self.assertAlmostEqual(25.0, by_code["C002"]["current_price"])
        self.assertAlmostEqual(5.0, by_code["C002"]["current_amount"])  # 0.2 * 25
        self.assertAlmostEqual(0.5, by_code["C001"]["simulated_ratio"])
        self.assertAlmostEqual(0.5, by_code["C002"]["simulated_ratio"])
        self.assertAlmostEqual(12.5, by_code["C002"]["simulated_amount"])  # 0.5 * 25
        self.assertAlmostEqual(17.5, recipe["simulated_total"])

    def test_recipe_cost_requires_formula_cost_permission(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/cost",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027"},
        )

        self.assertEqual(403, response.status_code)

    def test_formula_cost_user_can_export_cost_workbook_with_one_sheet_per_recipe(self) -> None:
        token = self._token(permissions=["formula.cost.calculate"])

        response = self.client.post(
            "/v1/recipes/cost/export",
            headers={"Authorization": f"Bearer {token}"},
            json={
                "query": "3027",
                "manual_prices": {"C002": 25},
                "simulated_quantities": {"C001": 1, "C002": 1000, "C003": 2},
            },
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        disposition = urllib.parse.unquote(response.headers["content-disposition"])
        self.assertIn("配方核算", disposition)
        self.assertIn("3027", disposition)
        self.assertIn("2个配方", disposition)
        wb = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(2, len(wb.sheetnames))
        self.assertTrue(any("V1" in name for name in wb.sheetnames))
        sheet = wb[wb.sheetnames[0]]
        merged = {str(item) for item in sheet.merged_cells.ranges}
        self.assertTrue({"A1:H1", "I1:J1", "K1:M1"}.issubset(merged))
        headers = [sheet.cell(2, col).value for col in range(1, 14)]
        self.assertEqual(
            ["子件编码", "子件名称", "规格型号", "单位", "数量", "比例", "系统单价", "系统分价", "当下价格", "当下分价", "模拟数量", "模拟比例", "模拟分价"],
            headers,
        )
        self.assertEqual("left", sheet["B3"].alignment.horizontal)
        self.assertEqual("left", sheet["C3"].alignment.horizontal)
        self.assertEqual("right", sheet["E3"].alignment.horizontal)
        self.assertLessEqual(sheet.column_dimensions["J"].width, 12)

    def test_recipe_cost_export_requires_formula_cost_permission(self) -> None:
        token = self._token(permissions=["formula.read"])

        response = self.client.post(
            "/v1/recipes/cost/export",
            headers={"Authorization": f"Bearer {token}"},
            json={"query": "3027"},
        )

        self.assertEqual(403, response.status_code)


if __name__ == "__main__":
    unittest.main()
