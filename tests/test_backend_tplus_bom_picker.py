from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "services" / "backend-api"


def _write_inventory(directory: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Code", "Name", "Specification", "BaseUnitCode", "BaseUnitName", "Disabled"])
    ws.append(["515", "测试原料515", "25kg", "1", "kg", "False"])
    ws.append(["FG-1", "测试成品", "", "2", "个", "False"])
    ws.append(["06000088", "停用旧父件", "", "1", "kg", "True"])
    wb.save(directory / "inventory_20260711_010830.xlsx")


def _write_stock(directory: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["WarehouseCode", "InventoryCode", "ExistingQuantity", "AvailableQuantity"])
    ws.append(["001", "515", "10", "8"])
    ws.append(["012", "515", "3", "2"])
    ws.append(["002", "FG-1", "9", "9"])
    wb.save(directory / "current_stock_20260711_010830.xlsx")


class TPlusBomPickerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path.insert(0, str(BACKEND_ROOT))
        from app.routers import tplus_bom as module
        cls.main = module

    @classmethod
    def tearDownClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = [item for item in sys.path if item != str(BACKEND_ROOT)]

    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.old_dir = os.environ.get("TPLUS_EXPORT_DIR")
        os.environ["TPLUS_EXPORT_DIR"] = self.tmp.name
        directory = Path(self.tmp.name)
        _write_inventory(directory)
        _write_stock(directory)
        self.old_scope = self.main._inventory_scope_config
        self.main._inventory_scope_config = lambda: ({"001", "012"}, {"001"})

    def tearDown(self) -> None:
        self.main._inventory_scope_config = self.old_scope
        if self.old_dir is None:
            os.environ.pop("TPLUS_EXPORT_DIR", None)
        else:
            os.environ["TPLUS_EXPORT_DIR"] = self.old_dir
        self.tmp.cleanup()

    def _user(self) -> dict:
        return {"sub": "tester", "roles": [], "permissions": ["tplus.bom.write"]}

    def test_reads_live_base_unit_columns(self):
        result = self.main.tplus_inventory_choices(q="515", limit=20, scope="all", user=self._user())
        self.assertEqual(1, result["total"])
        self.assertEqual("kg", result["items"][0]["unit_name"])
        self.assertEqual("1", result["items"][0]["unit_code"])

    def test_material_scope_joins_raw_warehouses_and_quantity(self):
        result = self.main.tplus_inventory_choices(q="测试", limit=20, scope="material", user=self._user())
        self.assertEqual(["515"], [item["code"] for item in result["items"]])
        self.assertEqual(10.0, result["items"][0]["available_quantity"])
        self.assertEqual(13.0, result["items"][0]["existing_quantity"])

    def test_disabled_rows_hidden_by_default(self):
        result = self.main.tplus_inventory_choices(q="06000088", limit=20, scope="all", user=self._user())
        self.assertEqual(0, result["total"])

    def test_include_disabled_reveals_disabled_rows(self):
        result = self.main.tplus_inventory_choices(
            q="06000088", limit=20, scope="all", include_disabled=True, user=self._user()
        )
        self.assertEqual(["06000088"], [item["code"] for item in result["items"]])

    def test_create_options_use_tplus_classes_and_synced_units(self):
        old_post = self.main._chanjet_read_post
        self.main._chanjet_read_post = lambda endpoint, payload: [
            {"Code": "01", "Name": "原材料", "IsEndNode": True},
            {"Code": "12", "Name": "代加工", "IsEndNode": False},
        ]
        try:
            result = self.main.tplus_inventory_create_options(user=self._user())
        finally:
            self.main._chanjet_read_post = old_post
        self.assertEqual([{"code": "01", "name": "原材料"}], result["classes"])
        self.assertEqual(
            [{"code": "1", "name": "kg"}, {"code": "2", "name": "个"}],
            result["units"],
        )


def _write_inventory_for_suggestion(directory: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Code", "Name", "Specification", "BaseUnitCode", "BaseUnitName", "Disabled"])
    ws.append(["06000009", "旧父件九", "", "1", "kg", "False"])
    ws.append(["06000012", "旧父件十二", "", "1", "kg", "False"])
    ws.append(["0316-CO712", "历史杂码", "", "1", "kg", "False"])
    ws.append(["069999", "位数不足不算", "", "1", "kg", "False"])
    ws.append(["01000030", "原料", "", "1", "kg", "False"])
    wb.save(directory / "inventory_20260713_010000.xlsx")


class TPlusCodeSuggestionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path.insert(0, str(BACKEND_ROOT))
        from app.routers import tplus_bom as module
        cls.main = module

    @classmethod
    def tearDownClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = [item for item in sys.path if item != str(BACKEND_ROOT)]

    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.old_dir = os.environ.get("TPLUS_EXPORT_DIR")
        os.environ["TPLUS_EXPORT_DIR"] = self.tmp.name
        _write_inventory_for_suggestion(Path(self.tmp.name))
        self.old_read_post = self.main._chanjet_read_post
        self.main._chanjet_read_post = lambda endpoint, payload: []
        self.old_memory = self.main._suggested_code_memory
        self.main._suggested_code_memory = lambda: set()
        self.old_remember = self.main._remember_inventory_code
        self.remembered: list[tuple[str, str, str]] = []
        self.main._remember_inventory_code = (
            lambda code, source, actor: self.remembered.append((code, source, actor))
        )

    def tearDown(self) -> None:
        self.main._chanjet_read_post = self.old_read_post
        self.main._suggested_code_memory = self.old_memory
        self.main._remember_inventory_code = self.old_remember
        if self.old_dir is None:
            os.environ.pop("TPLUS_EXPORT_DIR", None)
        else:
            os.environ["TPLUS_EXPORT_DIR"] = self.old_dir
        self.tmp.cleanup()

    def _user(self) -> dict:
        return {"sub": "tester", "roles": [], "permissions": ["tplus.bom.write"]}

    def test_suggests_max_serial_plus_one_for_prefix(self):
        result = self.main.tplus_inventory_code_suggestion(class_code="06", user=self._user())
        self.assertEqual("06000013", result["suggested"])
        self.assertEqual("06", result["prefix"])
        self.assertTrue(result["live_checked"])

    def test_first_code_when_prefix_unused(self):
        result = self.main.tplus_inventory_code_suggestion(class_code="09", user=self._user())
        self.assertEqual("09000001", result["suggested"])

    def test_subclass_code_uses_first_two_digits(self):
        # 末级分类 1201 → 前缀 12
        result = self.main.tplus_inventory_code_suggestion(class_code="1201", user=self._user())
        self.assertEqual("12000001", result["suggested"])

    def test_rejects_non_numeric_prefix(self):
        from fastapi import HTTPException
        with self.assertRaises(HTTPException) as ctx:
            self.main.tplus_inventory_code_suggestion(class_code="AB", user=self._user())
        self.assertEqual(400, ctx.exception.status_code)

    def test_live_duplicate_skips_to_next_free_and_remembers(self):
        taken = {"06000013"}
        def fake_query(endpoint, payload):
            code = payload["param"]["Code"]
            return [{"Code": code}] if code in taken else []
        self.main._chanjet_read_post = fake_query
        result = self.main.tplus_inventory_code_suggestion(class_code="06", user=self._user())
        self.assertEqual("06000014", result["suggested"])
        self.assertTrue(result["live_checked"])
        self.assertEqual(("06000013", "live_duplicate"), self.remembered[0][:2])

    def test_memory_codes_are_skipped(self):
        self.main._suggested_code_memory = lambda: {"06000013"}
        result = self.main.tplus_inventory_code_suggestion(class_code="06", user=self._user())
        self.assertEqual("06000014", result["suggested"])

    def test_falls_back_to_export_when_live_check_unavailable(self):
        from fastapi import HTTPException
        def fake_query(endpoint, payload):
            raise HTTPException(status_code=503, detail="T+ 查询凭据未配置")
        self.main._chanjet_read_post = fake_query
        result = self.main.tplus_inventory_code_suggestion(class_code="06", user=self._user())
        self.assertEqual("06000013", result["suggested"])
        self.assertFalse(result["live_checked"])


class TPlusBomPendingTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path.insert(0, str(BACKEND_ROOT))
        from app.routers import tplus_bom as module
        cls.main = module

    @classmethod
    def tearDownClass(cls) -> None:
        for name in list(sys.modules):
            if name == "app" or name.startswith("app."):
                del sys.modules[name]
        sys.path[:] = [item for item in sys.path if item != str(BACKEND_ROOT)]

    def _user(self) -> dict:
        return {"sub": "tester", "roles": [], "permissions": ["tplus.bom.audit"]}

    def test_mine_scope_returns_only_pending(self):
        self.main._load_success_submissions = lambda: [
            {"bom": {"dto": {"Inventory": {"Code": "06000001"}, "Version": "V1"}}},
            {"bom": {"dto": {"Inventory": {"Code": "06000002"}, "Version": "V1"}}},
        ]
        def fake_query(endpoint, payload):
            code = payload["dto"]["Code"]
            state = "00" if code == "06000001" else "01"
            return [{"Code": code, "Version": "V1", "ID": 1, "VoucherState": {"Code": state, "Name": "x"}, "BOMChildDTOs": []}]
        self.main._chanjet_read_post = fake_query
        result = self.main.tplus_bom_pending(user=self._user())
        self.assertEqual(["06000001"], [i["code"] for i in result["items"]])

    def test_requires_audit_permission(self):
        from fastapi import HTTPException
        with self.assertRaises(HTTPException) as ctx:
            self.main.tplus_bom_pending(user={"sub": "x", "roles": [], "permissions": []})
        self.assertEqual(403, ctx.exception.status_code)

    def test_audit_success_when_recheck_flips_to_audited(self):
        calls = []
        def fake_audit_post(endpoint, payload):
            calls.append(endpoint)
            return {"result": "ok"}, ""
        def fake_query(endpoint, payload):
            return [{"Code": "06000001", "Version": "V1", "ID": 1, "VoucherState": {"Code": "01", "Name": "已审"}}]
        self.main._chanjet_business_post = fake_audit_post
        self.main._chanjet_read_post = fake_query
        body = self.main.BomAuditBody(code="06000001", version="V1", bom_id="1")
        result = self.main.tplus_bom_audit(body=body, user=self._user())
        self.assertTrue(result["audited"])
        self.assertEqual("01", result["voucher_state"]["code"])
        self.assertIn(self.main.BOM_AUDIT_ENDPOINT, calls)

    def test_audit_not_audited_when_recheck_still_pending(self):
        self.main._chanjet_business_post = lambda endpoint, payload: ({}, "存货记账中，暂不能审核")
        self.main._chanjet_read_post = lambda endpoint, payload: [
            {"Code": "06000001", "Version": "V1", "ID": 1, "VoucherState": {"Code": "00", "Name": "未审"}}]
        body = self.main.BomAuditBody(code="06000001", version="V1", bom_id="1")
        result = self.main.tplus_bom_audit(body=body, user=self._user())
        self.assertFalse(result["audited"])
        self.assertIn("记账中", result["message"])

    def test_audit_requires_permission(self):
        from fastapi import HTTPException
        body = self.main.BomAuditBody(code="06000001", version="V1", bom_id="1")
        with self.assertRaises(HTTPException) as ctx:
            self.main.tplus_bom_audit(body=body, user={"sub": "x", "roles": [], "permissions": []})
        self.assertEqual(403, ctx.exception.status_code)


if __name__ == "__main__":
    unittest.main()
