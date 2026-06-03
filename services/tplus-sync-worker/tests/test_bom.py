import json
import tempfile
import unittest
from pathlib import Path

from config.settings import Settings
from tplus_datahub.modules.bom.export_bom import export_bom
from tplus_datahub.modules.bom.sync_bom import sync_bom
from tplus_datahub.modules.bom.transform_bom import transform_bom_rows, transform_bom_workbook_rows


class FakeClient:
    def __init__(self):
        self.payloads = []

    def post(self, endpoint, payload):
        self.payloads.append((endpoint, payload.copy()))
        page_index = payload["param"]["PageIndex"]
        if page_index == 1:
            disabled = payload["param"].get("Disabled")
            if disabled == "0":
                return {"Result": {"Rows": [{"Code": "BOM_ACTIVE", "Disabled": "False"}]}}
            if disabled == "1":
                return {"Result": {"Rows": [{"Code": "BOM_DISABLED", "Disabled": "True"}]}}
            return {"Result": {"Rows": [{"id": 1, "material": {"code": "M001"}}]}}
        return {"Result": {"Rows": []}}


class BomSyncTests(unittest.TestCase):
    def test_sync_bom_fetches_enabled_and_disabled_pages_and_saves_raw_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = Settings(
                base_url="https://openapi.example.com",
                app_key="app-key",
                app_secret="app-secret",
                open_token="open-token",
                default_page_size=1,
                timeout_connect=5,
                timeout_read=30,
                output_dir=str(Path(tmp) / "output"),
                data_dir=str(Path(tmp) / "data"),
            )
            client = FakeClient()

            rows = sync_bom(settings=settings, client=client, timestamp="20260601_153000")

            self.assertEqual(
                rows,
                [
                    {"Code": "BOM_ACTIVE", "Disabled": "False"},
                    {"Code": "BOM_DISABLED", "Disabled": "True"},
                ],
            )
            self.assertEqual(len(client.payloads), 4)
            self.assertEqual(client.payloads[0][1]["param"]["Disabled"], "0")
            self.assertEqual(client.payloads[0][1]["param"]["PageIndex"], 1)
            self.assertEqual(client.payloads[0][1]["param"]["PageSize"], 1)
            self.assertEqual(client.payloads[1][1]["param"]["PageIndex"], 2)
            self.assertEqual(client.payloads[2][1]["param"]["Disabled"], "1")
            self.assertEqual(client.payloads[2][1]["param"]["PageIndex"], 1)
            self.assertEqual(client.payloads[3][1]["param"]["PageIndex"], 2)
            enabled_raw = Path(tmp) / "data" / "raw" / "bom" / "20260601_153000_enabled_page_1.json"
            disabled_raw = Path(tmp) / "data" / "raw" / "bom" / "20260601_153000_disabled_page_1.json"
            self.assertTrue(enabled_raw.exists())
            self.assertTrue(disabled_raw.exists())
            self.assertEqual(json.loads(enabled_raw.read_text(encoding="utf-8"))["Result"]["Rows"][0]["Code"], "BOM_ACTIVE")
            self.assertEqual(json.loads(disabled_raw.read_text(encoding="utf-8"))["Result"]["Rows"][0]["Code"], "BOM_DISABLED")

    def test_transform_bom_rows_flattens_nested_dicts(self):
        rows = [{"id": 1, "material": {"code": "M001", "name": "Steel"}}]

        result = transform_bom_rows(rows)

        self.assertEqual(result, [{"id": 1, "material.code": "M001", "material.name": "Steel"}])

    def test_transform_bom_workbook_rows_splits_parent_and_child_details(self):
        rows = [
            {
                "Code": "40000008",
                "Name": "0463-ABS奶茶色母",
                "Specification": "4%",
                "Version": "20250110",
                "Unit": {"Name": "kg"},
                "ProduceQuantity": "150.00000000000000",
                "Manufactureplant": {"Code": "005", "Name": "生产部"},
                "Warehouse": {"Code": "001", "Name": "原料仓"},
                "IsDefaultBom": "True",
                "YieldRate": "1.00000000000000",
                "Disabled": "False",
                "CreateDate": "2026-01-01 10:00:00",
                "BOMChilds": [
                    {
                        "Code": "10001001",
                        "Name": "ABS树脂0215H",
                        "Specification": "35",
                        "Unit": {"Name": "kg"},
                        "RequiredQuantity": "16.50000000000000",
                        "ProduceQuantity": "150.00000000000000",
                        "WasteRate": "0.00000000000000",
                        "BackflushMaterial": "False",
                        "Warehouse": {"Code": "002", "Name": "材料仓"},
                    }
                ],
            }
        ]

        parent_rows, child_rows = transform_bom_workbook_rows(rows)

        self.assertEqual(parent_rows[0]["父件编码"], "40000008")
        self.assertEqual(parent_rows[0]["父件名称"], "0463-ABS奶茶色母")
        self.assertNotIn("BOMChilds", parent_rows[0])
        self.assertEqual(child_rows[0]["版本号"], "20250110")
        self.assertEqual(child_rows[0]["父件编码"], "40000008")
        self.assertEqual(child_rows[0]["子件编码"], "10001001")
        self.assertEqual(child_rows[0]["需用数量"], "16.50000000000000")
        self.assertEqual(child_rows[0]["标准用量"], "16.50000000000000")

    def test_export_bom_writes_parent_and_child_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = Settings(
                base_url="https://openapi.example.com",
                app_key="app-key",
                app_secret="app-secret",
                open_token="open-token",
                default_page_size=1,
                timeout_connect=5,
                timeout_read=30,
                output_dir=str(Path(tmp) / "output"),
                data_dir=str(Path(tmp) / "data"),
            )
            rows = [
                {
                    "Code": "40000008",
                    "Name": "0463-ABS奶茶色母",
                    "Version": "20250110",
                    "Unit": {"Name": "kg"},
                    "ProduceQuantity": "150.00000000000000",
                    "BOMChilds": [{"Code": "10001001", "Name": "ABS树脂0215H", "RequiredQuantity": "16.5"}],
                }
            ]

            target = export_bom(rows, settings=settings, timestamp="20260602_210000")

            from openpyxl import load_workbook

            workbook = load_workbook(target)
            self.assertEqual(workbook.sheetnames, ["物料清单", "子件明细"])
            parent_headers = [cell.value for cell in next(workbook["物料清单"].iter_rows(max_row=1))]
            child_headers = [cell.value for cell in next(workbook["子件明细"].iter_rows(max_row=1))]
            self.assertIn("父件编码", parent_headers)
            self.assertIn("子件编码", child_headers)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
