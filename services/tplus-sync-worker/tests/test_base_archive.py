import json
import tempfile
import unittest
from pathlib import Path

from config.settings import Settings
from tplus_datahub.modules.base_archive.export_base_archive import export_base_archive
from tplus_datahub.modules.base_archive.sync_base_archive import sync_base_archive


class FakeQueryClient:
    def __init__(self):
        self.calls = []

    def post(self, endpoint, payload):
        self.calls.append((endpoint, payload.copy()))
        # 第一页返回一条（满页 PageSize=1），第二页返回空让分页正常终止，
        # 否则「满页就翻下一页」的逻辑会无限翻页。
        if len(self.calls) == 1:
            return {"result": [{"Code": "01", "Name": "Main"}]}
        return {"result": []}


class BaseArchiveSyncTests(unittest.TestCase):
    def _settings(self, tmp):
        return Settings(
            base_url="https://openapi.example.com",
            app_key="app-key",
            app_secret="app-secret",
            open_token="open-token",
            default_page_size=1,
            timeout_connect=5,
            timeout_read=30,
            output_dir=str(Path(tmp) / "output"),
            data_dir=str(Path(tmp) / "data"),
        )

    def test_sync_base_archive_uses_query_body_and_saves_raw_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = self._settings(tmp)
            client = FakeQueryClient()

            rows = sync_base_archive(
                module_name="warehouse",
                endpoint="/tplus/api/v2/warehouse/Query",
                settings=settings,
                client=client,
                timestamp="20260604_011000",
            )

            self.assertEqual(rows, [{"Code": "01", "Name": "Main"}])
            self.assertEqual(
                client.calls,
                [
                    ("/tplus/api/v2/warehouse/Query", {"param": {"PageIndex": 1, "PageSize": 1}}),
                    ("/tplus/api/v2/warehouse/Query", {"param": {"PageIndex": 2, "PageSize": 1}}),
                ],
            )
            raw_file = Path(tmp) / "data" / "raw" / "warehouse" / "20260604_011000_page_1.json"
            self.assertTrue(raw_file.exists())
            self.assertEqual(json.loads(raw_file.read_text(encoding="utf-8"))["result"][0]["Code"], "01")

    def test_export_base_archive_writes_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = self._settings(tmp)

            target = export_base_archive(
                "warehouse",
                [{"Code": "01", "Name": "Main"}],
                settings=settings,
                timestamp="20260604_011100",
            )

            from openpyxl import load_workbook

            workbook = load_workbook(target)
            self.assertEqual(workbook.sheetnames, ["Sheet1"])
            headers = [cell.value for cell in next(workbook["Sheet1"].iter_rows(max_row=1))]
            self.assertEqual(headers, ["Code", "Name"])
            workbook.close()


if __name__ == "__main__":
    unittest.main()
