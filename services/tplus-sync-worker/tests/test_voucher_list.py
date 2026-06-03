import json
import tempfile
import unittest
from pathlib import Path

from config.settings import Settings
from tplus_datahub.modules.voucher.export_voucher_list import export_voucher_list
from tplus_datahub.modules.voucher.sync_voucher_list import sync_voucher_list


class FakeVoucherListClient:
    def __init__(self):
        self.calls = []

    def post(self, endpoint, payload):
        self.calls.append((endpoint, payload.copy()))
        page_index = payload["pageIndex"]
        if page_index == 0:
            return {
                "code": "0",
                "message": "ok",
                "data": {
                    "TotalCount": "2",
                    "TotalPageNum": "2",
                    "Columns": ["id", "voucherdate", "code"],
                    "Rows": [["1", "2026-06-01", "SO-001"]],
                },
            }
        return {
            "code": "0",
            "message": "ok",
            "data": {
                "TotalCount": "2",
                "TotalPageNum": "2",
                "Columns": ["id", "voucherdate", "code"],
                "Rows": [["2", "2026-06-02", "SO-002"]],
            },
        }


class VoucherListSyncTests(unittest.TestCase):
    def _settings(self, tmp):
        return Settings(
            base_url="https://openapi.example.com",
            app_key="app-key",
            app_secret="app-secret",
            open_token="open-token",
            default_page_size=1,
            timeout_connect=5,
            timeout_read=30,
            output_dir=str(Path(tmp) / "output"),
            data_dir=str(Path(tmp) / "data"),
        )

    def test_sync_voucher_list_pages_from_zero_and_maps_columns_to_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = self._settings(tmp)
            client = FakeVoucherListClient()

            rows = sync_voucher_list(
                module_name="sale_order_list",
                endpoint="/tplus/api/v2/SaleOrderOpenApi/FindVoucherList",
                select_fields=["SaleOrder.ID", "SaleOrder.VoucherDate", "SaleOrder.Code"],
                settings=settings,
                client=client,
                timestamp="20260604_012000",
            )

            self.assertEqual(
                rows,
                [
                    {"id": "1", "voucherdate": "2026-06-01", "code": "SO-001"},
                    {"id": "2", "voucherdate": "2026-06-02", "code": "SO-002"},
                ],
            )
            self.assertEqual(client.calls[0][1]["pageIndex"], 0)
            self.assertEqual(client.calls[0][1]["pageSize"], 1)
            self.assertEqual(client.calls[1][1]["pageIndex"], 1)
            self.assertEqual(client.calls[0][1]["paramDic"], {})
            raw_file = Path(tmp) / "data" / "raw" / "sale_order_list" / "20260604_012000_page_1.json"
            self.assertTrue(raw_file.exists())
            self.assertEqual(json.loads(raw_file.read_text(encoding="utf-8"))["data"]["TotalCount"], "2")

    def test_export_voucher_list_writes_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings = self._settings(tmp)

            target = export_voucher_list(
                "sale_order_list",
                [{"id": "1", "code": "SO-001"}],
                settings=settings,
                timestamp="20260604_012100",
            )

            from openpyxl import load_workbook

            workbook = load_workbook(target)
            headers = [cell.value for cell in next(workbook["Sheet1"].iter_rows(max_row=1))]
            self.assertEqual(headers, ["id", "code"])
            workbook.close()


if __name__ == "__main__":
    unittest.main()
