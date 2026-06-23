from __future__ import annotations

import base64
import calendar
import hashlib
import hmac
import json
import os
import secrets
import shutil
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from contextlib import closing
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import psycopg
from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from passlib.context import CryptContext
from psycopg.types.json import Jsonb
from pydantic import BaseModel, Field

from app.integrations.events import build_ops_attention_items
from app.logging_utils import configure_logging, log_event
from app.recipes.active_bom import copy_latest_bom_source, export_active_bom_rows
from app.recipes.bom_query import (
    calculate_recipe_costs,
    export_path_for_id,
    locate_recipe_source,
    new_export_path,
    query_recipe_workbook,
    recipe_cost_export_filename,
    save_recipe_cost_workbook,
    save_recipe_workbook,
)
from app.recipes.price_lookup import latest_purchase_prices, latest_sales_prices
from app.routers.webhooks import router as webhooks_router


app = FastAPI(title="AliECS Backend API", version="0.4.0")
app.include_router(webhooks_router)

def _cors_origins() -> list[str]:
    defaults = [
        "http://localhost:8080",
        "http://localhost:8081",
        "http://127.0.0.1:8080",
        "http://127.0.0.1:8081",
        "https://localhost:8080",
        "https://localhost:8081",
        "https://127.0.0.1:8080",
        "https://127.0.0.1:8081",
    ]
    origins = {x.strip() for x in defaults if x.strip()}

    app_base = os.getenv("APP_BASE_URL", "").strip()
    if app_base:
        origins.add(app_base.rstrip("/"))

    extra = os.getenv("CORS_ALLOWED_ORIGINS", "").strip()
    if extra:
        origins.update({x.strip().rstrip("/") for x in extra.split(",") if x.strip()})

    return sorted(origins)


app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins(),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

_request_logger = configure_logging("aliecs.request")


@app.middleware("http")
async def _log_requests(request, call_next):
    started = time.monotonic()
    response = await call_next(request)
    duration_ms = round((time.monotonic() - started) * 1000, 2)
    log_event(
        _request_logger,
        "request completed",
        request_id=request.headers.get("x-request-id", uuid.uuid4().hex),
        method=request.method,
        path=request.url.path,
        status_code=response.status_code,
        duration_ms=duration_ms,
    )
    return response

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")

DEFAULT_FEATURES: list[dict[str, Any]] = [
    {"id": 9, "code": "raw_inventory", "title": "原材料库存", "description": "原材料库存查询", "url": "/inventory/raw-materials/", "category": "业务查询", "required_permission": "inventory.raw.read", "status": "active", "sort_order": 10},
    {"id": 10, "code": "finished_inventory", "title": "成品库存", "description": "成品库存查询", "url": "/inventory/finished-goods/", "category": "业务查询", "required_permission": "inventory.finished.read", "status": "active", "sort_order": 20},
    {"id": 7, "code": "formula_query", "title": "系统配方", "description": "配方检索、BOM 同步与成本核算", "url": "/formula/", "category": "业务查询", "required_permission": "formula.read", "status": "active", "sort_order": 30},
    {"id": 1, "code": "new_model_form", "title": "新品型号录入表", "description": "新品型号登记", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_4b7094", "category": "业务录入", "required_permission": "production.schedule.write", "status": "active", "sort_order": 40},
    {"id": 2, "code": "schedule_form", "title": "排产登记表", "description": "排产信息登记", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_e3792e", "category": "业务录入", "required_permission": "production.schedule.write", "status": "active", "sort_order": 50},
    {"id": 3, "code": "pending_return_alert", "title": "待处理+退货提醒", "description": "待处理与退货提醒", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_4501d0", "category": "业务录入", "required_permission": "production.schedule.read", "status": "active", "sort_order": 60},
    {"id": 4, "code": "naming_form", "title": "产品命名登记", "description": "产品命名录入", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_a577fc", "category": "业务录入", "required_permission": "formula.write", "status": "active", "sort_order": 70},
    {"id": 5, "code": "qc_form", "title": "检测数据登记表", "description": "检测数据登记", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_b669cf", "category": "质检", "required_permission": "formula.read", "status": "active", "sort_order": 80},
    {"id": 6, "code": "density_calculator", "title": "配方密度计算器", "description": "配方密度工具", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_bac993", "category": "质检", "required_permission": "formula.read", "status": "active", "sort_order": 90},
    {"id": 8, "code": "midea_requirement", "title": "美的需求", "description": "需求查询入口", "url": None, "category": "业务查询", "required_permission": "midea.requirement.read", "status": "reserved", "sort_order": 100},
    {"id": 11, "code": "personal_section", "title": "个人板块", "description": "个人工具入口", "url": "https://doc.weixin.qq.com/smartsheet/form/1_wp7hSPEQAAT1c_JcnLpU1STlUJOXWRPA_0c521a", "category": "个人", "required_permission": "personal.access", "status": "active", "sort_order": 110},
    {"id": 12, "code": "admin_ui", "title": "Admin UI", "description": "管理后台入口", "url": "/admin/", "category": "系统", "required_permission": "admin.access", "status": "active", "sort_order": 120},
]


def _database_url() -> str:
    return os.getenv("DATABASE_URL", "")


def _conn() -> psycopg.Connection:
    database_url = _database_url()
    if not database_url:
        raise HTTPException(status_code=500, detail="DATABASE_URL is empty")
    return psycopg.connect(database_url, connect_timeout=3)


def _db_ping() -> tuple[bool, str]:
    if not _database_url():
        return False, "DATABASE_URL is empty"

    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
        return True, "db ok"
    except Exception as exc:
        return False, f"db error: {exc}"


def _token_secret() -> str:
    secret = os.getenv("AUTH_TOKEN_SECRET", "change-this-in-production")
    env_name = os.getenv("ENV", "dev")
    if env_name == "prod" and secret == "change-this-in-production":
        raise HTTPException(status_code=500, detail="AUTH_TOKEN_SECRET must be changed in production")
    if env_name == "prod" and len(secret) < 32:
        raise HTTPException(status_code=500, detail="AUTH_TOKEN_SECRET must be at least 32 characters in production")
    return secret


def _token_ttl_seconds() -> int:
    raw = os.getenv("AUTH_TOKEN_TTL_SECONDS", "28800")
    try:
        return int(raw)
    except ValueError as exc:
        raise HTTPException(status_code=500, detail="AUTH_TOKEN_TTL_SECONDS must be integer") from exc


def _sign(payload: str) -> str:
    return hmac.new(_token_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def _encode_token(payload: dict[str, Any]) -> str:
    payload = dict(payload)
    payload.setdefault("jti", uuid.uuid4().hex)
    body = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    b64 = base64.urlsafe_b64encode(body.encode("utf-8")).decode("utf-8").rstrip("=")
    return f"{b64}.{_sign(b64)}"


def _decode_token(token: str) -> dict[str, Any]:
    try:
        b64, sig = token.split(".", 1)
    except ValueError as exc:
        raise HTTPException(status_code=401, detail="invalid token") from exc

    if not hmac.compare_digest(sig, _sign(b64)):
        raise HTTPException(status_code=401, detail="invalid token")

    try:
        body = base64.urlsafe_b64decode(b64 + "=" * (-len(b64) % 4)).decode("utf-8")
        payload = json.loads(body)
    except Exception as exc:
        raise HTTPException(status_code=401, detail="invalid token") from exc

    if int(payload.get("exp", 0)) <= int(time.time()):
        raise HTTPException(status_code=401, detail="token expired")

    return payload


def _extract_bearer(authorization: str | None) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="missing authorization")
    return authorization[7:].strip()


def _audit(
    actor: str | None,
    action: str,
    target_type: str | None = None,
    target_id: str | None = None,
    detail: dict[str, Any] | None = None,
) -> None:
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO audit_logs(actor_username, action, target_type, target_id, detail)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (actor, action, target_type, target_id, Jsonb(detail or {})),
                )
            conn.commit()
    except Exception:
        # 审计失败不应阻断主流程。
        pass


def _user_roles_permissions(user_id: int, is_admin: bool = False) -> tuple[list[str], list[str]]:
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT r.code
                    FROM roles r
                    JOIN user_roles ur ON ur.role_id = r.id
                    WHERE ur.user_id = %s
                    ORDER BY r.id
                    """,
                    (user_id,),
                )
                roles = [row[0] for row in cur.fetchall()]

                cur.execute(
                    """
                    SELECT DISTINCT p.code
                    FROM permissions p
                    JOIN role_permissions rp ON rp.permission_id = p.id
                    JOIN user_roles ur ON ur.role_id = rp.role_id
                    WHERE ur.user_id = %s
                    ORDER BY p.code
                    """,
                    (user_id,),
                )
                permissions = [row[0] for row in cur.fetchall()]
    except Exception:
        # 兼容旧环境（RBAC 关联表缺失或尚未初始化）：
        # 允许登录继续，按 is_admin 做最小权限回退。
        roles = []
        permissions = []

    if is_admin and "admin" not in roles:
        roles.append("admin")
    if is_admin and "admin.access" not in permissions:
        permissions.append("admin.access")

    return roles, permissions


def _current_token_version(user_id: int) -> int | None:
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT token_version FROM users WHERE id = %s", (user_id,))
                row = cur.fetchone()
    except Exception:
        return None
    if not row:
        return None
    return int(row[0])


def _bootstrap_admin_if_needed() -> None:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM users")
            user_count = cur.fetchone()[0]
            if user_count > 0:
                return

            username = os.getenv("ADMIN_BOOTSTRAP_USERNAME", "admin")
            password = os.getenv("ADMIN_BOOTSTRAP_PASSWORD", "admin123")
            display_name = os.getenv("ADMIN_BOOTSTRAP_DISPLAY_NAME", "系统管理员")

            cur.execute(
                """
                INSERT INTO users(username, display_name, password_hash, is_admin, status)
                VALUES (%s, %s, %s, true, 'active')
                RETURNING id
                """,
                (username, display_name, pwd_ctx.hash(password)),
            )
            user_id = cur.fetchone()[0]

            try:
                cur.execute("SELECT id FROM roles WHERE code = 'admin'")
                role_row = cur.fetchone()
                if role_row:
                    cur.execute(
                        """
                        INSERT INTO user_roles(user_id, role_id)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        (user_id, role_row[0]),
                    )
            except Exception:
                # 兼容旧环境（roles / user_roles 尚未迁移完成）：
                # 不阻断管理员账户初始化，登录后由 is_admin 回退授予最小后台权限。
                pass

        conn.commit()

    _audit(username, "auth.bootstrap_admin", "users", str(user_id))


def get_current_user(authorization: str | None = Header(default=None)) -> dict[str, Any]:
    token = _extract_bearer(authorization)
    payload = _decode_token(token)
    uid = payload.get("uid")
    if uid is not None and "tv" in payload:
        current = _current_token_version(int(uid))
        if current is not None and int(payload["tv"]) != current:
            raise HTTPException(status_code=401, detail="token revoked")
    return payload


def require_login(user: dict[str, Any] = Depends(get_current_user)) -> dict[str, Any]:
    return user


def require_admin(user: dict[str, Any] = Depends(get_current_user)) -> dict[str, Any]:
    roles = user.get("roles", [])
    permissions = user.get("permissions", [])
    if "admin" in roles or "admin.access" in permissions:
        return user
    raise HTTPException(status_code=403, detail="permission denied")


def require_permission(permission: str, user: dict[str, Any]) -> dict[str, Any]:
    roles = user.get("roles", [])
    permissions = user.get("permissions", [])
    if "admin" in roles or "admin.access" in permissions or permission in permissions:
        return user
    raise HTTPException(status_code=403, detail="permission denied")


def _couple_feature_enabled() -> bool:
    return os.getenv("COUPLE_FEATURE_ENABLED", "true").lower() in {"1", "true", "yes", "on"}


def _couple_route() -> str:
    route = os.getenv("COUPLE_ROUTE", "/couple/").strip()
    if not route.startswith("/"):
        return "/couple"
    return route


def _has_couple_access(user: dict[str, Any]) -> bool:
    if not _couple_feature_enabled():
        return False

    permissions = user.get("permissions", [])
    if "couple_memory_access" in permissions:
        return True

    username = str(user.get("username") or user.get("sub") or "").strip().lower()
    email = str(user.get("email") or "").strip().lower()
    roles = [str(r).lower() for r in user.get("roles", [])]
    if "admin" in roles:
        return True

    allowed_users = [i.strip().lower() for i in os.getenv("COUPLE_ALLOWED_USERS", "").split(",") if i.strip()]
    if username and username in allowed_users:
        return True

    allowed_emails = [i.strip().lower() for i in os.getenv("COUPLE_ALLOWED_EMAILS", "").split(",") if i.strip()]
    if email and email in allowed_emails:
        return True

    return False


class LoginRequest(BaseModel):
    username: str
    password: str


class RegisterRequest(BaseModel):
    username: str
    password: str
    note: str | None = None


class CreateUserRequest(BaseModel):
    username: str
    display_name: str
    password: str
    is_admin: bool = False


class PatchUserRequest(BaseModel):
    display_name: str | None = None
    status: str | None = None
    is_admin: bool | None = None


class ResetPasswordRequest(BaseModel):
    new_password: str = Field(min_length=6)


class CreateRoleRequest(BaseModel):
    code: str
    name: str
    description: str | None = None


class PatchRoleRequest(BaseModel):
    name: str | None = None
    description: str | None = None


class PutRoleIdsRequest(BaseModel):
    role_ids: list[int]


class PutPermissionIdsRequest(BaseModel):
    permission_ids: list[int]


class ManagedContactUpsertRequest(BaseModel):
    channel: str
    peer_id: str
    display_name: str | None = None
    remark: str | None = None
    enabled: bool = True
    project_url: str | None = None
    project_name: str | None = None
    tags: str | None = None
    daily_quota: int | None = None
    notes: str | None = None
    source_sheet: str | None = None


class MemoryUpsertRequest(BaseModel):
    couple_space_id: int | None = None
    title: str
    content: str | None = None
    memory_date: str | None = None
    place_name: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    cover_photo_url: str | None = None
    visibility: str = "private"
    tags: list[str] = Field(default_factory=list)


class ImmichAssetBindRequest(BaseModel):
    asset_ids: list[str] = Field(default_factory=list)
    immich_asset_id: str | None = None
    immich_album_id: str | None = None
    original_filename: str | None = None
    taken_at: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    thumbnail_cache_key: str | None = None
    sort_order: int = 0


class PatchCoupleSpaceRequest(BaseModel):
    name: str | None = None
    start_date: str | None = None
    theme: str | None = None
    cover_image_url: str | None = None


class AnniversaryUpsertRequest(BaseModel):
    couple_space_id: int | None = None
    title: str
    date: str
    repeat_type: str = "yearly"
    description: str | None = None


class BucketItemUpsertRequest(BaseModel):
    couple_space_id: int | None = None
    title: str
    description: str | None = None
    status: str = "want"
    target_date: str | None = None
    completed_memory_id: int | None = None


class CreateShareLinkRequest(BaseModel):
    expires_in_days: int | None = Field(default=None, ge=1, le=365)


class RecipeQueryRequest(BaseModel):
    query: str = Field(min_length=1, max_length=100)
    default_bom: str = "all"
    include_disabled: bool = True


class RecipeCostRequest(RecipeQueryRequest):
    manual_prices: dict[str, float] = Field(default_factory=dict)
    simulated_quantities: dict[str, float] = Field(default_factory=dict)


class ReconciliationActionRequest(BaseModel):
    action: str = Field(pattern="^(use_current|use_previous|use_full|use_incremental|ignore)$")
    note: str | None = Field(default=None, max_length=500)


def _tplus_bom_sync_request_dir() -> Path:
    return Path(os.getenv("TPLUS_BOM_SYNC_REQUEST_DIR", "/tmp/aliecs-tplus-sync-requests"))


def _create_tplus_bom_sync_request(requested_by: str | None) -> dict[str, Any]:
    request_dir = _tplus_bom_sync_request_dir()
    request_dir.mkdir(parents=True, exist_ok=True)
    request_id = uuid.uuid4().hex
    payload = {
        "id": request_id,
        "provider": "chanjet",
        "module": "bom",
        "mode": "manual_bom_full_include_disabled",
        "include_disabled": True,
        "requested_by": requested_by or "",
        "requested_at": int(time.time()),
    }
    (request_dir / f"{request_id}.json").write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return {"id": request_id, "status": "pending", "mode": payload["mode"]}


ALLOWED_UPLOAD_MIMES = {
    "image/jpeg": {".jpg", ".jpeg"},
    "image/png": {".png"},
    "image/webp": {".webp"},
    "image/gif": {".gif"},
}


def _max_upload_bytes() -> int:
    raw = os.getenv("MAX_UPLOAD_MB", os.getenv("PHOTO_MAX_UPLOAD_MB", "15"))
    try:
        mb = max(1, int(raw))
    except ValueError:
        mb = 15
    return mb * 1024 * 1024


def _detect_image_mime(data: bytes) -> str | None:
    if data.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data.startswith(b"RIFF") and len(data) >= 12 and data[8:12] == b"WEBP":
        return "image/webp"
    return None


def _validate_photo_upload(filename: str | None, declared_mime: str | None, content: bytes) -> tuple[str, str]:
    ext = Path(filename or "").suffix.lower()
    detected_mime = _detect_image_mime(content)
    declared = (declared_mime or "").split(";", 1)[0].strip().lower()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")
    if len(content) > _max_upload_bytes():
        raise HTTPException(status_code=400, detail="file too large")
    if not detected_mime or detected_mime not in ALLOWED_UPLOAD_MIMES:
        raise HTTPException(status_code=400, detail="unsupported file type")
    if ext not in ALLOWED_UPLOAD_MIMES[detected_mime]:
        raise HTTPException(status_code=400, detail="file extension does not match image type")
    if declared and declared not in ALLOWED_UPLOAD_MIMES:
        raise HTTPException(status_code=400, detail="unsupported file type")
    if declared and declared != detected_mime:
        raise HTTPException(status_code=400, detail="mime type does not match file content")
    return ext, detected_mime


def _public_upload_url(filename: str) -> str:
    public_base = os.getenv("APP_BASE_URL", "").rstrip("/")
    relative_url = f"/uploads/{filename}"
    return f"{public_base}{relative_url}" if public_base else relative_url


def _public_photo_content_url(key: str) -> str:
    public_base = os.getenv("APP_BASE_URL", "").rstrip("/")
    relative_url = f"/api/v1/photos/content/{urllib.parse.quote(key, safe='')}"
    return f"{public_base}{relative_url}" if public_base else relative_url


def _webdock_photo_base_url() -> str:
    return os.getenv("WEBDOCK_PHOTO_BASE_URL", "http://host.docker.internal:11800").rstrip("/")


def _webdock_photo_token() -> str:
    token = os.getenv("WEBDOCK_PHOTO_API_TOKEN") or os.getenv("WEB_DOCK_API_TOKEN") or ""
    if not token:
        raise HTTPException(status_code=500, detail="WEBDOCK_PHOTO_API_TOKEN is required")
    return token


def _webdock_photo_timeout() -> int:
    raw = os.getenv("WEBDOCK_PHOTO_TIMEOUT_SECONDS", "30")
    try:
        return max(1, int(raw))
    except ValueError:
        return 30


def _webdock_photo_url(key: str | None = None) -> str:
    if key is None:
        return f"{_webdock_photo_base_url()}/storage/photos"
    return f"{_webdock_photo_base_url()}/storage/photos/{urllib.parse.quote(key, safe='')}"


def _multipart_photo_body(filename: str | None, content_type: str, content: bytes) -> tuple[bytes, str]:
    boundary = f"----aliecs-{uuid.uuid4().hex}"
    safe_name = (filename or "photo").replace("\\", "_").replace('"', "_")
    parts = [
        f"--{boundary}\r\n".encode("utf-8"),
        (
            f'Content-Disposition: form-data; name="file"; filename="{safe_name}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8"),
        content,
        f"\r\n--{boundary}--\r\n".encode("utf-8"),
    ]
    return b"".join(parts), f"multipart/form-data; boundary={boundary}"


def _webdock_photo_request(
    method: str,
    key: str | None = None,
    *,
    body: bytes | None = None,
    headers: dict[str, str] | None = None,
) -> bytes:
    request_headers = {"Authorization": f"Bearer {_webdock_photo_token()}"}
    if headers:
        request_headers.update(headers)
    request = urllib.request.Request(
        _webdock_photo_url(key),
        data=body,
        headers=request_headers,
        method=method,
    )
    try:
        with urllib.request.urlopen(request, timeout=_webdock_photo_timeout()) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            raise HTTPException(status_code=404, detail="photo not found") from exc
        raise HTTPException(status_code=502, detail=f"webdock photo storage returned {exc.code}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail="webdock photo storage unavailable") from exc


def _webdock_photo_key(original_storage_url: str | None) -> str | None:
    if not original_storage_url or not original_storage_url.startswith("webdock:"):
        return None
    key = original_storage_url.split(":", 1)[1].strip()
    return key or None


class PhotoStorage:
    driver = "local"

    async def save(self, file: UploadFile) -> dict[str, str]:
        raise NotImplementedError

    def delete(self, original_storage_url: str | None) -> None:
        return None


def _upload_disk_usage(path: Path | str) -> dict[str, Any]:
    try:
        usage = shutil.disk_usage(path)
    except OSError as exc:
        # The upload dir is created lazily by LocalPhotoStorage, so it may not
        # exist yet at healthcheck time. A missing/unstat-able path must not make
        # /healthz fail — report it as unavailable instead of raising.
        return {"path": str(path), "available": False, "percent": None, "error": str(exc)}
    percent = round(usage.used * 100 / usage.total, 1) if usage.total else 0.0
    return {
        "path": str(path),
        "available": True,
        "total": usage.total,
        "used": usage.used,
        "free": usage.free,
        "percent": percent,
    }


class LocalPhotoStorage(PhotoStorage):
    driver = "local"

    def __init__(self) -> None:
        self.base_dir = Path(os.getenv("LOCAL_UPLOAD_DIR", "/app/uploads"))
        self.base_dir.mkdir(parents=True, exist_ok=True)

    async def save(self, file: UploadFile) -> dict[str, str]:
        content = await file.read()
        return self.save_content(file.filename, file.content_type, content)

    def save_content(self, filename: str | None, content_type: str | None, content: bytes) -> dict[str, str]:
        ext, _mime = _validate_photo_upload(filename, content_type, content)
        filename = f"{uuid.uuid4().hex}{ext}"
        full_path = self.base_dir / filename
        full_path.write_bytes(content)
        self._warn_if_disk_high()
        public_url = _public_upload_url(filename)
        return {
            "original_storage_url": str(full_path),
            "display_url": public_url,
            "thumbnail_url": public_url,
            "storage_driver": self.driver,
        }

    def delete(self, original_storage_url: str | None) -> None:
        if original_storage_url:
            Path(original_storage_url).unlink(missing_ok=True)

    def _warn_if_disk_high(self) -> None:
        info = _upload_disk_usage(self.base_dir)
        if info.get("percent") is None:
            return
        raw = os.getenv("UPLOAD_DISK_WARN_PCT", "").strip()
        if not raw:
            return
        try:
            threshold = float(raw)
        except ValueError:
            return
        if info["percent"] >= threshold:
            log_event(
                _request_logger,
                "upload disk usage high",
                path=info["path"],
                percent=info["percent"],
                threshold=threshold,
            )


class OssPhotoStorage(PhotoStorage):
    driver = "oss"

    def __init__(self) -> None:
        from app.oss_client import OssClient, config_from_env

        config = config_from_env()
        if not config.enabled:
            self._client = None
        else:
            self._client = OssClient(config)

    async def save(self, file: UploadFile) -> dict[str, str]:
        content = await file.read()
        if self._client is None:
            raise HTTPException(status_code=501, detail="OSS storage is not configured in this build")

        ext, mime = _validate_photo_upload(file.filename, file.content_type, content)
        key = f"couple/{uuid.uuid4().hex}{ext}"
        from app.oss_client import OssError

        try:
            self._client.put_object(key, content, mime)
        except OssError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc

        public_url = self._client.object_url(key)
        return {
            "original_storage_url": f"oss:{key}",
            "display_url": public_url,
            "thumbnail_url": public_url,
            "storage_driver": self.driver,
        }

    def delete(self, original_storage_url: str | None) -> None:
        if not original_storage_url or not original_storage_url.startswith("oss:"):
            return
        if self._client is None:
            return
        key = original_storage_url.split(":", 1)[1]
        from app.oss_client import OssError

        try:
            self._client.delete_object(key)
        except OssError:
            pass


class WebDockPhotoStorage(PhotoStorage):
    driver = "webdock"

    async def save(self, file: UploadFile) -> dict[str, str]:
        content = await file.read()
        _ext, mime = _validate_photo_upload(file.filename, file.content_type, content)
        body, content_type = _multipart_photo_body(file.filename, mime, content)
        try:
            raw = _webdock_photo_request("POST", body=body, headers={"Content-Type": content_type})
        except HTTPException as exc:
            if exc.status_code == 502:
                return LocalPhotoStorage().save_content(file.filename, mime, content)
            raise
        try:
            payload = json.loads(raw.decode("utf-8"))
        except Exception as exc:
            raise HTTPException(status_code=502, detail="invalid webdock photo storage response") from exc
        key = str(payload.get("key") or "").strip()
        if not key:
            raise HTTPException(status_code=502, detail="webdock photo storage did not return key")
        public_url = _public_photo_content_url(key)
        return {
            "original_storage_url": f"webdock:{key}",
            "display_url": public_url,
            "thumbnail_url": public_url,
            "storage_driver": self.driver,
        }

    def delete(self, original_storage_url: str | None) -> None:
        key = _webdock_photo_key(original_storage_url)
        if key:
            _webdock_photo_request("DELETE", key)


def photo_storage() -> PhotoStorage:
    driver = os.getenv("STORAGE_DRIVER", "local").strip().lower() or "local"
    if driver == "local":
        return LocalPhotoStorage()
    if driver == "oss":
        return OssPhotoStorage()
    if driver == "webdock":
        return WebDockPhotoStorage()
    raise HTTPException(status_code=500, detail="invalid STORAGE_DRIVER")


@app.get("/uploads/{name}")
def serve_upload(name: str) -> FileResponse:
    base = Path(os.getenv("LOCAL_UPLOAD_DIR", "/app/uploads")).resolve()
    target = (base / name).resolve()
    if base not in target.parents or not target.is_file():
        raise HTTPException(status_code=404, detail="not found")
    return FileResponse(target)


class CreateFeatureRequest(BaseModel):
    code: str
    title: str
    description: str | None = None
    url: str | None = None
    category: str | None = None
    required_permission: str | None = None
    status: str = "active"
    sort_order: int = 100


class PatchFeatureRequest(BaseModel):
    title: str | None = None
    description: str | None = None
    url: str | None = None
    category: str | None = None
    required_permission: str | None = None
    status: str | None = None
    sort_order: int | None = None


class CreateCoupleSpaceRequest(BaseModel):
    name: str
    start_date: str | None = None
    theme: str | None = None


class AddCoupleMemberRequest(BaseModel):
    username: str
    role: str = "member"


@app.get("/healthz")
def healthz() -> dict[str, object]:
    db_ok, db_message = _db_ping()
    upload_dir = os.getenv("LOCAL_UPLOAD_DIR", "/tmp/aliecs-uploads")
    return {
        "status": "ok" if db_ok else "degraded",
        "service": "backend-api",
        "database": {"ok": db_ok, "message": db_message},
        "upload_disk": _upload_disk_usage(upload_dir),
    }


@app.get("/v1/ops/tplus/runs")
def ops_tplus_runs(
    limit: int = Query(default=20, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    """全部 T+ 同步执行记录（含每小时 scheduled_full 与手动 bom），分页。
    数据源 integration_sync_runs，比 ops_status 里只取 10 条的 recent_requests 完整。"""
    items: list[dict[str, Any]] = []
    total = 0
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM integration_sync_runs WHERE provider = 'chanjet'")
                total = int(cur.fetchone()[0])
                cur.execute(
                    """
                    SELECT sr.id, sr.module, sr.mode, sr.status, sr.finished_at, sr.exit_code, sr.row_count,
                           req.id, req.reason_event_id
                    FROM integration_sync_runs sr
                    LEFT JOIN LATERAL (
                        SELECT id, reason_event_id
                        FROM integration_sync_requests
                        WHERE provider = 'chanjet' AND sync_run_id = sr.id
                        ORDER BY requested_at DESC NULLS LAST, id DESC
                        LIMIT 1
                    ) req ON TRUE
                    WHERE sr.provider = 'chanjet'
                    ORDER BY sr.finished_at DESC NULLS LAST, sr.id DESC
                    LIMIT %s OFFSET %s
                    """,
                    (limit, offset),
                )
                items = [
                    {
                        "id": row[0],
                        "module": row[1],
                        "mode": row[2],
                        "status": row[3],
                        "finished_at": str(row[4]) if row[4] else None,
                        "exit_code": row[5],
                        "row_count": row[6],
                        "request_id": row[7],
                        "reason_event_id": row[8],
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取 T+ 同步记录失败：{type(exc).__name__}") from exc
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@app.get("/v1/ops/tplus/requests")
def ops_tplus_requests(
    limit: int = Query(default=20, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    """全部 T+ 同步请求（畅捷通回调事件触发的 bom 同步），分页。reason_event_id=回调事件ID。"""
    items: list[dict[str, Any]] = []
    total = 0
    counts = {"pending": 0, "running": 0, "failed": 0}
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM integration_sync_requests WHERE provider = 'chanjet'")
                total = int(cur.fetchone()[0])
                cur.execute(
                    "SELECT status, COUNT(*) FROM integration_sync_requests WHERE provider = 'chanjet' GROUP BY status"
                )
                status_map = {row[0]: int(row[1]) for row in cur.fetchall()}
                counts = {key: status_map.get(key, 0) for key in ("pending", "running", "failed")}
                cur.execute(
                    """
                    SELECT r.id, r.module, r.mode, r.status, r.requested_at, r.started_at,
                           r.finished_at, r.reason_event_id, r.target_json, r.sync_run_id,
                           r.error_json, sr.detail_json, sr.error_json, sr.row_count, sr.exit_code
                    FROM integration_sync_requests r
                    LEFT JOIN integration_sync_runs sr ON sr.id = r.sync_run_id
                    WHERE r.provider = 'chanjet'
                    ORDER BY r.requested_at DESC, r.id DESC
                    LIMIT %s OFFSET %s
                    """,
                    (limit, offset),
                )
                items = [
                    {
                        "id": row[0],
                        "module": row[1],
                        "mode": row[2],
                        "status": row[3],
                        "requested_at": str(row[4]) if row[4] else None,
                        "started_at": str(row[5]) if row[5] else None,
                        "finished_at": str(row[6]) if row[6] else None,
                        "reason_event_id": row[7],
                        "target_json": _json_value(row[8]),
                        "sync_run_id": row[9],
                        "request_error_json": _json_value(row[10]),
                        "detail_json": _json_value(row[11]),
                        "error_json": _json_value(row[12]),
                        "row_count": row[13],
                        "exit_code": row[14],
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取 T+ 同步请求失败：{type(exc).__name__}") from exc
    return {"items": items, "total": total, "limit": limit, "offset": offset, **counts}


@app.get("/v1/ops/tplus/timeline")
def ops_tplus_timeline(
    limit: int = Query(default=20, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    """统一时间线：执行(run) + 无执行的孤儿请求，按时间倒序分页；附产出 Excel 与变化摘要。"""
    items: list[dict[str, Any]] = []
    total = 0
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT (SELECT COUNT(*) FROM integration_sync_runs WHERE provider='chanjet')
                         + (SELECT COUNT(*) FROM integration_sync_requests
                            WHERE provider='chanjet' AND sync_run_id IS NULL)
                    """
                )
                total = int(cur.fetchone()[0])
                cur.execute(
                    """
                    SELECT kind, id, module, mode, status, event_time, row_count, exit_code,
                           reason_event_id, request_id, detail_json, reconciliation_id
                    FROM (
                        SELECT 'run' AS kind, sr.id AS id, sr.module, sr.mode, sr.status,
                               sr.finished_at AS event_time, sr.row_count, sr.exit_code,
                               req.reason_event_id, req.id AS request_id, sr.detail_json,
                               rec.id AS reconciliation_id
                        FROM integration_sync_runs sr
                        LEFT JOIN LATERAL (
                            SELECT id, reason_event_id FROM integration_sync_requests
                            WHERE provider='chanjet' AND sync_run_id = sr.id
                            ORDER BY requested_at DESC NULLS LAST, id DESC LIMIT 1
                        ) req ON TRUE
                        LEFT JOIN LATERAL (
                            SELECT d.id FROM integration_reconciliation_diffs d
                            JOIN integration_sync_snapshots s ON s.id = d.full_snapshot_id
                            WHERE d.provider='chanjet' AND d.status='needs_review'
                              AND s.created_at <= sr.finished_at
                            ORDER BY s.created_at DESC LIMIT 1
                        ) rec ON TRUE
                        WHERE sr.provider='chanjet'
                        UNION ALL
                        SELECT 'request' AS kind, r.id, r.module, r.mode, r.status,
                               r.requested_at AS event_time, NULL::int, NULL::int,
                               r.reason_event_id, r.id, r.error_json, NULL::bigint
                        FROM integration_sync_requests r
                        WHERE r.provider='chanjet' AND r.sync_run_id IS NULL
                    ) merged
                    ORDER BY event_time DESC NULLS LAST, kind, id DESC
                    LIMIT %s OFFSET %s
                    """,
                    (limit, offset),
                )
                rows = cur.fetchall()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取 T+ 时间线失败：{type(exc).__name__}") from exc

    run_rows = [(r[1], r[5]) for r in rows if r[0] == "run"]
    export_dir = _tplus_export_dir()
    disk_files = [p.name for p in export_dir.glob("*.xlsx")] if export_dir.is_dir() else []
    fallback = _match_export_files_to_runs(run_rows, disk_files)
    existing = set(disk_files)

    for (kind, rid, module, mode, status, event_time, row_count, exit_code,
         reason_event_id, request_id, detail, reconciliation_id) in rows:
        detail = _json_value(detail) or {}
        diff_summary = detail.get("diff_summary")
        row: dict[str, Any] = {
            "kind": kind,
            "number": f"#{rid}" if kind == "run" else f"请求·R{rid}",
            "id": rid,
            "module": module,
            "mode": mode,
            "status": status,
            "event_time": str(event_time) if event_time else None,
            "row_count": row_count,
            "exit_code": exit_code,
            "reason_event_id": reason_event_id,
            "request_id": request_id,
            "diff_summary": diff_summary,
            "needs_review": bool((diff_summary or {}).get("needs_review")),
            "reconciliation_id": reconciliation_id,
            "export_files": [],
        }
        if kind == "run":
            names = list(detail.get("export_files") or []) or fallback.get(rid, [])
            row["export_files"] = [
                {"name": name,
                 "download_url": f"/v1/exports/tplus/{name}" if name in existing else None,
                 "pruned": name not in existing}
                for name in names
            ]
        items.append(row)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@app.get("/v1/ops/status")
def ops_status(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    db_ok, db_message = _db_ping()
    status: dict[str, Any] = {
        "status": "ok" if db_ok else "degraded",
        "service": "backend-api",
        "database": {"ok": db_ok, "message": db_message},
        "system": _system_status(),
        "tplus": _tplus_status_from_db() if db_ok else _empty_tplus_status(),
        "reconciliation": _reconciliation_status_from_db() if db_ok else {"needs_review": 0, "recent": []},
        "hosts": _configured_host_statuses(),
    }
    status["attention_items"] = build_ops_attention_items(status)
    if status["attention_items"]:
        status["status"] = "degraded"
    return status


@app.get("/v1/ops/hosts")
def ops_hosts() -> dict[str, Any]:
    return {"items": _configured_host_statuses()}


@app.get("/v1/ops/hosts/{host_name}/refresh")
def ops_host_refresh(host_name: str) -> dict[str, Any]:
    for target in _ops_http_targets():
        if str(target.get("name") or target.get("url") or "target") == host_name:
            return _probe_http_target(target)
    raise HTTPException(status_code=404, detail="host target not found")


def _empty_tplus_status() -> dict[str, Any]:
    return {
        "pending_requests": 0,
        "running_requests": 0,
        "failed_requests": 0,
        "last_success_at": None,
        "last_run": None,
        "recent_requests": [],
    }


def _system_status() -> dict[str, Any]:
    disk = shutil.disk_usage("/")
    memory = _memory_status()
    result: dict[str, Any] = {
        "disk_total": disk.total,
        "disk_used": disk.used,
        "disk_free": disk.free,
        "disk_percent": round(disk.used / disk.total * 100, 1) if disk.total else 0.0,
        **memory,
    }
    try:
        load1, load5, load15 = os.getloadavg()
        result["loadavg"] = [round(load1, 2), round(load5, 2), round(load15, 2)]
    except (AttributeError, OSError):
        result["loadavg"] = []
    return result


def _memory_status() -> dict[str, Any]:
    meminfo = Path("/proc/meminfo")
    if not meminfo.is_file():
        return {"memory_total": 0, "memory_available": 0, "memory_percent": 0.0}
    values: dict[str, int] = {}
    for line in meminfo.read_text(encoding="utf-8", errors="ignore").splitlines():
        if ":" not in line:
            continue
        key, raw = line.split(":", 1)
        parts = raw.strip().split()
        if parts and parts[0].isdigit():
            values[key] = int(parts[0]) * 1024
    total = values.get("MemTotal", 0)
    available = values.get("MemAvailable", 0)
    used = max(total - available, 0)
    return {
        "memory_total": total,
        "memory_available": available,
        "memory_percent": round(used / total * 100, 1) if total else 0.0,
    }


def _tplus_status_from_db() -> dict[str, Any]:
    status = _empty_tplus_status()
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT status, COUNT(*)
                    FROM integration_sync_requests
                    WHERE provider = 'chanjet' AND module = 'bom'
                    GROUP BY status
                    """
                )
                for row_status, count in cur.fetchall():
                    if row_status == "pending":
                        status["pending_requests"] = int(count)
                    elif row_status == "running":
                        status["running_requests"] = int(count)
                    elif row_status == "failed":
                        status["failed_requests"] = int(count)
                cur.execute(
                    """
                    SELECT id, module, mode, status, started_at, finished_at, row_count, exit_code, detail_json
                    FROM integration_sync_runs
                    WHERE provider = 'chanjet'
                    ORDER BY started_at DESC, id DESC
                    LIMIT 1
                    """
                )
                last_run = cur.fetchone()
                if last_run:
                    status["last_run"] = _sync_run_to_dict(last_run)
                cur.execute(
                    """
                    SELECT finished_at
                    FROM integration_sync_runs
                    WHERE provider = 'chanjet' AND status = 'success'
                    ORDER BY finished_at DESC NULLS LAST, id DESC
                    LIMIT 1
                    """
                )
                last_success = cur.fetchone()
                if last_success and last_success[0]:
                    status["last_success_at"] = str(last_success[0])
                cur.execute(
                    """
                    SELECT r.id, r.module, r.mode, r.status, r.requested_at, r.started_at,
                           r.finished_at, r.reason_event_id, r.target_json, r.sync_run_id,
                           r.error_json, sr.detail_json, sr.error_json, sr.row_count, sr.exit_code
                    FROM integration_sync_requests r
                    LEFT JOIN integration_sync_runs sr ON sr.id = r.sync_run_id
                    WHERE r.provider = 'chanjet'
                    ORDER BY r.requested_at DESC, r.id DESC
                    LIMIT 10
                    """
                )
                status["recent_requests"] = [
                    {
                        "id": row[0],
                        "module": row[1],
                        "mode": row[2],
                        "status": row[3],
                        "requested_at": str(row[4]) if row[4] else None,
                        "started_at": str(row[5]) if row[5] else None,
                        "finished_at": str(row[6]) if row[6] else None,
                        "reason_event_id": row[7],
                        "target_json": _json_value(row[8]),
                        "sync_run_id": row[9],
                        "request_error_json": _json_value(row[10]),
                        "detail_json": _json_value(row[11]),
                        "error_json": _json_value(row[12]),
                        "row_count": row[13],
                        "exit_code": row[14],
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        status["error"] = str(exc)
    return status


def _reconciliation_status_from_db() -> dict[str, Any]:
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM integration_reconciliation_diffs
                    WHERE status = 'needs_review'
                    """
                )
                needs_review = int(cur.fetchone()[0])
                cur.execute(
                    """
                    SELECT id, provider, module, severity, summary, created_at
                    FROM integration_reconciliation_diffs
                    WHERE status = 'needs_review'
                    ORDER BY created_at DESC, id DESC
                    LIMIT 10
                    """
                )
                recent = [
                    {
                        "id": row[0],
                        "provider": row[1],
                        "module": row[2],
                        "severity": row[3],
                        "summary": row[4],
                        "created_at": str(row[5]) if row[5] else None,
                    }
                    for row in cur.fetchall()
                ]
        return {"needs_review": needs_review, "recent": recent}
    except Exception as exc:
        return {"needs_review": 0, "recent": [], "error": str(exc)}


def _json_value(value: Any) -> Any:
    if value is None:
        return {}
    if isinstance(value, (dict, list)):
        return value
    if hasattr(value, "obj"):
        return getattr(value, "obj")
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return {"raw": value}
    return value


def _reconciliation_diff_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "id": row[0],
        "provider": row[1],
        "module": row[2],
        "status": row[3],
        "severity": row[4],
        "summary": row[5],
        "diff_json": _json_value(row[6]),
        "full_snapshot_id": row[7],
        "incremental_snapshot_id": row[8],
        "created_at": str(row[9]) if row[9] else None,
        "reviewed_at": str(row[10]) if row[10] else None,
        "reviewed_by": row[11],
        "resolution": _json_value(row[12]),
    }


@app.get("/v1/ops/reconciliation/{diff_id}")
def ops_reconciliation_detail(diff_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, provider, module, status, severity, summary, diff_json,
                       full_snapshot_id, incremental_snapshot_id, created_at,
                       reviewed_at, reviewed_by, resolution_json
                FROM integration_reconciliation_diffs
                WHERE id = %s
                """,
                (diff_id,),
            )
            row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="reconciliation diff not found")
    return _reconciliation_diff_to_dict(row)


@app.post("/v1/ops/reconciliation/{diff_id}/actions")
def ops_reconciliation_action(
    diff_id: int,
    body: ReconciliationActionRequest,
    user: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    next_status = "ignored" if body.action == "ignore" else "resolved"
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, provider, module, status, severity, summary, diff_json,
                       full_snapshot_id, incremental_snapshot_id, created_at,
                       reviewed_at, reviewed_by, resolution_json
                FROM integration_reconciliation_diffs
                WHERE id = %s
                FOR UPDATE
                """,
                (diff_id,),
            )
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="reconciliation diff not found")

            existing_diff = _reconciliation_diff_to_dict(existing)
            selected_snapshot_id = _selected_reconciliation_snapshot(existing_diff, body.action)
            resolution = {
                "action": body.action,
                "note": body.note or "",
                "resolved_at": datetime.now(timezone.utc).isoformat(),
            }
            if next_status == "resolved":
                if selected_snapshot_id is None:
                    raise HTTPException(status_code=400, detail="selected snapshot not found")
                activation = _activate_bom_snapshot(conn, selected_snapshot_id, allow_latest_fallback=body.action != "use_previous")
                resolution.update(activation)
                resolution["selected_snapshot_id"] = selected_snapshot_id

            cur.execute(
                """
                UPDATE integration_reconciliation_diffs
                SET status = %s,
                    resolution_json = %s,
                    reviewed_at = NOW(),
                    reviewed_by = %s
                WHERE id = %s
                RETURNING id, provider, module, status, severity, summary, diff_json,
                          full_snapshot_id, incremental_snapshot_id, created_at,
                          reviewed_at, reviewed_by, resolution_json
                """,
                (next_status, Jsonb(resolution), user.get("sub", ""), diff_id),
            )
            row = cur.fetchone()
            if next_status == "resolved":
                cur.execute(
                    """
                    UPDATE integration_reconciliation_diffs
                    SET status = 'superseded',
                        resolution_json = %s,
                        reviewed_at = NOW(),
                        reviewed_by = %s
                    WHERE provider = %s
                      AND module = %s
                      AND status = 'needs_review'
                      AND id < %s
                    """,
                    (
                        Jsonb(
                            {
                                "action": "superseded_by_newer_resolution",
                                "superseded_by_diff_id": diff_id,
                                "selected_snapshot_id": selected_snapshot_id,
                                "active_export_name": resolution.get("active_export_name"),
                                "resolved_at": resolution["resolved_at"],
                            }
                        ),
                        user.get("sub", ""),
                        existing_diff["provider"],
                        existing_diff["module"],
                        diff_id,
                    ),
                )
        conn.commit()
    _audit(user.get("sub"), "ops.reconciliation.resolve", "integration_reconciliation_diffs", str(diff_id), resolution)
    return _reconciliation_diff_to_dict(row)


def _selected_reconciliation_snapshot(diff: dict[str, Any], action: str) -> int | None:
    diff_json = diff.get("diff_json") if isinstance(diff.get("diff_json"), dict) else {}
    if action == "ignore":
        return None
    if action in {"use_current", "use_incremental"}:
        return _int_or_none(diff.get("incremental_snapshot_id")) or _int_or_none(diff_json.get("current_snapshot_id")) or _int_or_none(diff.get("full_snapshot_id"))
    if action == "use_previous":
        return _int_or_none(diff_json.get("previous_snapshot_id"))
    if action == "use_full":
        return _int_or_none(diff.get("full_snapshot_id")) or _int_or_none(diff_json.get("current_snapshot_id"))
    return None


def _activate_bom_snapshot(conn: psycopg.Connection, snapshot_id: int, *, allow_latest_fallback: bool = True) -> dict[str, Any]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT source_json
            FROM integration_sync_snapshots
            WHERE id = %s
            """,
            (snapshot_id,),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="selected snapshot not found")
    source = _json_value(row[0])
    records = source.get("records") if isinstance(source.get("records"), list) else []
    if records:
        return export_active_bom_rows(records)
    if not allow_latest_fallback:
        raise HTTPException(status_code=409, detail="selected snapshot has no stored BOM records")
    return copy_latest_bom_source()


def _int_or_none(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _sync_run_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "id": row[0],
        "module": row[1],
        "mode": row[2],
        "status": row[3],
        "started_at": str(row[4]) if row[4] else None,
        "finished_at": str(row[5]) if row[5] else None,
        "row_count": row[6],
        "exit_code": row[7],
        "detail_json": _json_value(row[8]),
    }


def _configured_host_statuses() -> list[dict[str, Any]]:
    return [_probe_http_target(item) for item in _ops_http_targets()]


def _ops_http_targets() -> list[dict[str, Any]]:
    raw = os.getenv("OPS_HEALTH_HTTP_TARGETS_JSON", "").strip()
    if not raw:
        targets = [
            {
                "name": "AliECS Backend API",
                "url": os.getenv("OPS_HEALTH_BACKEND_URL", "").strip(),
                "description": "公网反代后的 AliECS 后端健康检查。",
                "timeout": 3,
            },
            {
                "name": "AliECS Public Web",
                "url": os.getenv("OPS_HEALTH_PUBLIC_WEB_URL", "").strip(),
                "description": "公网首页入口。",
                "timeout": 3,
            },
            {
                "name": "WebDock API",
                "url": os.getenv("OPS_HEALTH_WEBDOCK_API_URL", "http://host.docker.internal:11800/healthz"),
                "description": "旧电脑 WebDock API，经服务器 SSH 隧道 11800 端口探测。",
                "timeout": 3,
            },
        ]
        novnc_url = os.getenv("OPS_HEALTH_WEBDOCK_NOVNC_URL", "").strip()
        if novnc_url:
            targets.append(
                {
                    "name": "WebDock noVNC",
                    "url": novnc_url,
                    "description": "旧电脑 noVNC 页面，需显式配置可由 backend-api 访问的地址。",
                    "timeout": 3,
                }
            )
        # Drop targets without an explicit URL (e.g. backend/public-web before
        # OPS_HEALTH_*_URL is configured) so the health page doesn't render
        # blank "url is empty" rows.
        return [t for t in targets if str(t.get("url") or "").strip()]
    try:
        targets = json.loads(raw)
    except Exception as exc:
        return [{"name": "OPS_HEALTH_HTTP_TARGETS_JSON", "url": "", "description": f"invalid json: {exc}"}]
    if not isinstance(targets, list):
        return [{"name": "OPS_HEALTH_HTTP_TARGETS_JSON", "url": "", "description": "must be a list"}]
    return [item for item in targets if isinstance(item, dict)]


def _probe_http_target(item: dict[str, Any]) -> dict[str, Any]:
    name = str(item.get("name") or item.get("url") or "target")
    url = str(item.get("url") or "")
    description = str(item.get("description") or "")
    checked_at = datetime.now(timezone.utc).isoformat()
    timeout = float(item.get("timeout") or 2)
    if not url:
        return {"name": name, "url": url, "description": description, "ok": False, "message": "url is empty", "last_checked_at": checked_at}
    started = time.perf_counter()
    try:
        request = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(request, timeout=timeout) as response:
            status_code = int(getattr(response, "status", 0) or 0)
        elapsed_ms = round((time.perf_counter() - started) * 1000)
        return {
            "name": name,
            "url": url,
            "description": description,
            "ok": 200 <= status_code < 500,
            "status_code": status_code,
            "latency_ms": elapsed_ms,
            "last_checked_at": checked_at,
        }
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000)
        return {
            "name": name,
            "url": url,
            "description": description,
            "ok": False,
            "message": str(exc),
            "latency_ms": elapsed_ms,
            "last_checked_at": checked_at,
        }


def _wechat_login_qr_from_gateway() -> dict[str, Any] | None:
    url = os.getenv("OPENCLAW_WECHAT_LOGIN_QR_URL", "").strip()
    if not url:
        return None
    request = urllib.request.Request(url, method="GET")
    try:
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"openclaw qr gateway failed: {exc}") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=502, detail="openclaw qr gateway returned non-object json")
    if not (payload.get("qr_image_base64") or payload.get("qr_url")):
        raise HTTPException(status_code=502, detail="openclaw qr gateway response missing qr_image_base64/qr_url")
    payload = dict(payload)
    payload.setdefault("source", "gateway")
    return payload


def _wechat_login_qr_from_file() -> dict[str, Any] | None:
    raw_path = os.getenv("OPENCLAW_WECHAT_LOGIN_QR_FILE", "").strip()
    if not raw_path:
        return None
    path = Path(raw_path)
    if not path.is_file():
        raise HTTPException(status_code=404, detail="wechat login qr file not found")
    data = path.read_bytes()
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return {"qr_image_base64": "data:image/png;base64," + base64.b64encode(data).decode("ascii"), "source": "file"}
    text = data.decode("utf-8", errors="replace").strip()
    if text.startswith(("http://", "https://")):
        return {"qr_url": text, "source": "file"}
    if text.startswith("data:image/"):
        return {"qr_image_base64": text, "source": "file"}
    return {"qr_image_base64": "data:image/png;base64," + base64.b64encode(data).decode("ascii"), "source": "file"}


@app.get("/v1/ops/wechat/login-qr")
def ops_wechat_login_qr(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    payload = _wechat_login_qr_from_gateway() or _wechat_login_qr_from_file()
    if payload:
        payload.setdefault("expires_at", None)
        payload.setdefault("message", "等待新用户扫码加入微信 clawbot。")
        return payload
    raise HTTPException(
        status_code=503,
        detail="未配置稳定二维码来源。请在 OpenClaw 主机运行 openclaw channels login --channel openclaw-weixin，或配置 OPENCLAW_WECHAT_LOGIN_QR_URL/FILE。",
    )


def _extract_wecom_b_message(payload: dict[str, Any]) -> dict[str, Any]:
    body = payload.get("body") if isinstance(payload.get("body"), dict) else payload
    sender = body.get("from") if isinstance(body.get("from"), dict) else {}
    msg_type = str(body.get("msgtype") or body.get("msg_type") or "")
    content = ""
    if msg_type == "text" and isinstance(body.get("text"), dict):
        content = str(body["text"].get("content") or "")
    elif msg_type and isinstance(body.get(msg_type), dict):
        content = json.dumps(body[msg_type], ensure_ascii=False, sort_keys=True)
    msg_id = str(body.get("msgid") or body.get("msg_id") or payload.get("msgid") or "").strip()
    if not msg_id:
        raise HTTPException(status_code=400, detail="missing msgid")
    return {
        "msg_id": msg_id,
        "bot_id": str(body.get("aibotid") or body.get("bot_id") or ""),
        "chat_id": str(body.get("chatid") or body.get("chat_id") or ""),
        "chat_type": str(body.get("chattype") or body.get("chat_type") or ""),
        "sender_id": str(sender.get("userid") or sender.get("user_id") or body.get("from_user_id") or ""),
        "msg_type": msg_type,
        "content": content,
    }


@app.post("/v1/webhooks/wecom-b/messages")
def wecom_b_capture_message(
    payload: dict[str, Any],
    x_wecom_capture_token: str | None = Header(default=None),
) -> dict[str, str]:
    expected = os.getenv("WECOM_B_CAPTURE_TOKEN", "").strip()
    if expected and x_wecom_capture_token != expected:
        raise HTTPException(status_code=403, detail="invalid capture token")
    message = _extract_wecom_b_message(payload)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO wecom_b_messages(
                    msg_id, bot_id, chat_id, chat_type, sender_id, msg_type, content, raw_json, received_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT(msg_id) DO NOTHING
                """,
                (
                    message["msg_id"],
                    message["bot_id"],
                    message["chat_id"],
                    message["chat_type"],
                    message["sender_id"],
                    message["msg_type"],
                    message["content"],
                    Jsonb(payload),
                ),
            )
        conn.commit()
    return {"status": "received", "msg_id": message["msg_id"]}


@app.get("/readyz")
def readyz() -> dict[str, object]:
    db_ok, db_message = _db_ping()
    if db_ok:
        return {"status": "ready"}
    return {"status": "not-ready", "reason": db_message}


@app.get("/v1/ping")
def ping() -> dict[str, str]:
    return {"message": "pong"}


@app.post("/v1/auth/login")
def auth_login(body: LoginRequest) -> dict[str, Any]:
    _bootstrap_admin_if_needed()

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, username, display_name, password_hash, status, is_admin, token_version
                FROM users
                WHERE username = %s
                """,
                (body.username,),
            )
            row = cur.fetchone()

            try:
                verified = bool(row) and row[4] == "active" and pwd_ctx.verify(body.password, row[3])
            except Exception:
                verified = False
            if not verified:
                raise HTTPException(status_code=401, detail="invalid credentials")

            user_id = row[0]
            roles, permissions = _user_roles_permissions(user_id, bool(row[5]))

            now = int(time.time())
            payload = {
                "sub": row[1],
                "uid": user_id,
                "display_name": row[2],
                "roles": roles,
                "permissions": permissions,
                "tv": int(row[6]),
                "jti": uuid.uuid4().hex,
                "iat": now,
                "exp": now + _token_ttl_seconds(),
            }

            cur.execute(
                "UPDATE users SET last_login_at = NOW(), updated_at = NOW() WHERE id = %s",
                (user_id,),
            )

        conn.commit()

    _audit(row[1], "auth.login")
    return {
        "token": _encode_token(payload),
        "expires_in": _token_ttl_seconds(),
        "user": {
            "username": row[1],
            "display_name": row[2],
            "roles": roles,
            "permissions": permissions,
        },
    }


@app.post("/v1/auth/register")
def auth_register(body: RegisterRequest) -> dict[str, Any]:
    _audit(body.username, "auth.register.request", detail={"note": body.note})
    return {
        "status": "pending_review",
        "message": "注册请求已提交，请联系管理员在后台分配账号权限。",
        "requested_user": body.username,
    }


@app.post("/v1/auth/logout")
def auth_logout(_: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    return {"status": "ok"}


@app.get("/v1/auth/me")
def auth_me(user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    return user


@app.get("/v1/features")
def features(
    authorization: str | None = Header(default=None),
    include_all: bool = Query(default=False),
) -> dict[str, Any]:
    guest = True
    user: dict[str, Any] = {"sub": "guest", "roles": ["guest"], "permissions": []}

    if authorization:
        try:
            user = get_current_user(authorization)
            guest = False
        except HTTPException:
            guest = True

    is_admin = "admin" in user.get("roles", []) or "admin.access" in user.get("permissions", [])

    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, code, title, description, url, category, required_permission, status, sort_order
                    FROM features
                    WHERE status != 'disabled'
                    ORDER BY sort_order ASC, id ASC
                    """
                )
                rows = cur.fetchall()
    except Exception:
        # 兼容迁移尚未完成场景（features 表缺失）：
        # 首页仍返回 200 与用户态，并回退到默认功能列表，避免首页空白。
        rows = [
            (
                feature["id"],
                feature["code"],
                feature["title"],
                feature["description"],
                feature["url"],
                feature["category"],
                feature["required_permission"],
                feature["status"],
                feature["sort_order"],
            )
            for feature in DEFAULT_FEATURES
        ]

    items = []
    for row in rows:
        required_permission = row[6]
        allowed = False

        if is_admin:
            allowed = True
        elif required_permission is None:
            allowed = True
        elif not guest and required_permission in user.get("permissions", []):
            allowed = True

        if include_all or allowed:
            items.append(
                {
                    "id": row[0],
                    "code": row[1],
                    "title": row[2],
                    "description": row[3],
                    "url": row[4],
                    "category": row[5],
                    "required_permission": required_permission,
                    "status": row[7],
                    "sort_order": row[8],
                    "allowed": allowed,
                }
            )

    return {
        "user": {
            "username": user.get("sub", "guest"),
            "roles": user.get("roles", []),
            "permissions": user.get("permissions", []),
        },
        "features": items,
    }


# 查询结果导出文件延迟生成：查询时只记录上下文，用户真正点「下载原始明细」时才写 xlsx
# （save_recipe_workbook 约 2.3s，绝大多数查询并不会下载，没必要每次都写）。
_RECIPE_QUERY_CONTEXT: dict[str, dict[str, object]] = {}
_RECIPE_QUERY_CONTEXT_MAX = 256


def _remember_recipe_query(file_id: str, query: str, default_bom: str | None, include_disabled: bool) -> None:
    _RECIPE_QUERY_CONTEXT[file_id] = {
        "query": query,
        "default_bom": default_bom,
        "include_disabled": include_disabled,
    }
    while len(_RECIPE_QUERY_CONTEXT) > _RECIPE_QUERY_CONTEXT_MAX:
        _RECIPE_QUERY_CONTEXT.pop(next(iter(_RECIPE_QUERY_CONTEXT)))


def _latest_bom_sync_run() -> dict[str, Any] | None:
    """产出当前 BOM 文件的那次同步：最近一次成功且会导出 bom 的 run（scheduled_full 或手动 bom）。
    locate_recipe_source 取 mtime 最新文件 ⇔ 最近一次成功 bom 同步，故二者对应。任何异常降级为 None。"""
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, module, mode, status, finished_at
                    FROM integration_sync_runs
                    WHERE provider = 'chanjet' AND status = 'success' AND module IN ('all', 'bom')
                    ORDER BY finished_at DESC NULLS LAST, id DESC
                    LIMIT 1
                    """
                )
                row = cur.fetchone()
                if not row:
                    return None
                return {
                    "id": row[0],
                    "module": row[1],
                    "mode": row[2],
                    "status": row[3],
                    "finished_at": str(row[4]) if row[4] else None,
                }
    except Exception:
        return None


@app.post("/v1/recipes/query")
def recipe_query(body: RecipeQueryRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    require_permission("formula.read", user)
    try:
        source_path = locate_recipe_source()
        result = query_recipe_workbook(
            source_path,
            query_text=body.query,
            default_bom=body.default_bom,
            include_disabled=body.include_disabled,
        )
        file_id, _output_path = new_export_path()
        _remember_recipe_query(file_id, body.query, body.default_bom, body.include_disabled)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="BOM 输入文件未找到") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"配方查询失败：{type(exc).__name__}") from exc

    return {
        "query": body.query,
        "source_file": source_path.name,
        "source_sync": _latest_bom_sync_run(),
        "match_count": result.match_count,
        "recipe_count": result.recipe_count,
        "default_bom": result.default_bom,
        "include_disabled": result.include_disabled,
        "file_id": file_id,
        "download_url": f"/v1/recipes/download/{file_id}",
        "preview": result.preview_rows(limit=5000),
    }


def _recipe_price_maps() -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]]]:
    """最新采购/销售价格映射；任何读取异常都降级为空，不影响成本核算。"""
    try:
        return latest_purchase_prices(), latest_sales_prices()
    except Exception:
        return {}, {}


@app.post("/v1/recipes/cost")
def recipe_cost(body: RecipeCostRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    require_permission("formula.cost.calculate", user)
    try:
        source_path = locate_recipe_source()
        result = query_recipe_workbook(
            source_path,
            query_text=body.query,
            default_bom=body.default_bom,
            include_disabled=body.include_disabled,
        )
        purchase_prices, sales_prices = _recipe_price_maps()
        recipes = calculate_recipe_costs(
            result,
            manual_prices=body.manual_prices,
            simulated_quantities=body.simulated_quantities,
            purchase_prices=purchase_prices,
            sales_prices=sales_prices,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="BOM 输入文件未找到") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"配方成本核算失败：{type(exc).__name__}") from exc

    return {
        "query": body.query,
        "source_file": source_path.name,
        "recipe_count": len(recipes),
        "default_bom": result.default_bom,
        "include_disabled": result.include_disabled,
        "manual_price_count": len(body.manual_prices),
        "simulated_quantity_count": len(body.simulated_quantities),
        "recipes": recipes,
    }


@app.post("/v1/recipes/cost/export")
def recipe_cost_export(body: RecipeCostRequest, user: dict[str, Any] = Depends(require_login)) -> FileResponse:
    require_permission("formula.cost.calculate", user)
    try:
        source_path = locate_recipe_source()
        result = query_recipe_workbook(
            source_path,
            query_text=body.query,
            default_bom=body.default_bom,
            include_disabled=body.include_disabled,
        )
        purchase_prices, sales_prices = _recipe_price_maps()
        recipes = calculate_recipe_costs(
            result,
            manual_prices=body.manual_prices,
            simulated_quantities=body.simulated_quantities,
            purchase_prices=purchase_prices,
            sales_prices=sales_prices,
        )
        _file_id, output_path = new_export_path()
        save_recipe_cost_workbook(output_path, recipes)
        filename = recipe_cost_export_filename(recipes, body.query)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="BOM 输入文件未找到") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"配方成本核算导出失败：{type(exc).__name__}") from exc

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )



@app.post("/v1/recipes/sync-bom")
def recipe_sync_bom(user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    require_permission("formula.read", user)
    try:
        request = _create_tplus_bom_sync_request(user.get("sub"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"BOM 同步请求创建失败：{type(exc).__name__}") from exc
    return {
        **request,
        "module": "bom",
        "include_disabled": True,
        "message": "已请求同步 T+ 物料清单 BOM；默认全量包含停用配方。",
    }


@app.get("/v1/recipes/download/{file_id}")
def recipe_download(file_id: str, user: dict[str, Any] = Depends(require_login)) -> FileResponse:
    require_permission("formula.read", user)
    try:
        path = export_path_for_id(file_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not path.is_file():
        # 延迟生成：用查询时记录的上下文按需写出导出文件
        context = _RECIPE_QUERY_CONTEXT.get(file_id)
        if context is None:
            raise HTTPException(status_code=404, detail="下载文件已过期，请重新查询后再下载。")
        try:
            result = query_recipe_workbook(
                locate_recipe_source(),
                query_text=str(context["query"]),
                default_bom=context["default_bom"],
                include_disabled=bool(context["include_disabled"]),
            )
            save_recipe_workbook(path, result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"导出文件生成失败：{type(exc).__name__}") from exc
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"配方查询_{file_id}.xlsx",
    )


_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


_TPLUS_EXPORT_DESCRIPTIONS = {
    "bom": "BOM 父件和子件用量；看成品由哪些原材料组成，不含价格。",
    "inventory": "存货基础档案；含分类、采购/销售/材料标记、税率、RetailPriceNew/InvUnitPriceDTOs。销售价格先看这里；采购价未接入。",
    "current_stock": "仓库×存货现存量/可用量；看原材库库存数量，不含价格。",
    "partner": "往来单位档案；客户/供应商及价格等级，不是商品价格。",
    "warehouse": "仓库档案；用于识别原材库、成品库等仓库编码。",
    "unit_group": "计量单位组档案；辅助理解单位换算关系。",
    "unit": "计量单位档案；含主单位、换算率等单位信息。",
    "project": "项目档案；当前可能为空，主要用于项目辅助核算。",
    "project_class": "项目分类档案；用于识别项目类别。",
    "brand": "品牌档案；当前可能为空。",
    "district": "地区档案；当前可能为空。",
    "sale_order_list": "销售订单列表，仅单据ID/日期/单号；不含明细售价金额。",
    "sale_delivery_list": "销货单列表，仅单据ID/日期/单号；不含明细售价金额。",
    "purchase_order_list": "采购订单列表，仅单据ID/日期/单号；不含明细单价金额，不能核对原材料采购价。",
    "purchase_arrival_list": "采购到货单列表，仅单据ID/日期/单号；不含明细单价金额。",
    "purchase_receive_list": "采购入库单列表，仅单据ID/日期/单号；不含明细单价金额。",
    "material_dispatch_list": "材料出库/领料单列表，仅单据ID/日期/单号；不含价格。",
    "purchase_price": "采购价格表（采购到货明细，来自 T+ 报表）；原材料采购单价/含税单价优先看这里，成本核算系统单价取此最新价。",
    "sales_price": "销售价格表（销货单明细，来自 T+ 报表）；商品销售单价/含税单价优先看这里。",
}


def _tplus_export_description(module: str) -> str:
    return _TPLUS_EXPORT_DESCRIPTIONS.get(module, "暂未配置说明；请按表头人工判断内容。")


def _tplus_export_dir() -> Path:
    return Path(os.getenv("TPLUS_EXPORT_DIR", "/app/tplus-output/excel"))


def _tplus_module_of(file_name: str) -> str:
    stem = file_name[:-5] if file_name.endswith(".xlsx") else file_name
    parts = stem.rsplit("_", 2)
    if len(parts) == 3 and parts[1].isdigit() and parts[2].isdigit():
        return parts[0]
    return stem


def _parse_export_timestamp(file_name: str) -> datetime | None:
    stem = file_name[:-5] if file_name.endswith(".xlsx") else file_name
    parts = stem.rsplit("_", 2)
    if len(parts) == 3 and len(parts[1]) == 8 and len(parts[2]) == 6 and parts[1].isdigit() and parts[2].isdigit():
        try:
            return datetime.strptime(parts[1] + parts[2], "%Y%m%d%H%M%S")
        except ValueError:
            return None
    return None


def _match_export_files_to_runs(runs: list[tuple[Any, Any]], files: list[str]) -> dict[Any, list[str]]:
    """把每个文件归给 finished_at >= 文件时间戳 的最早一次 run。
    runs: [(run_id, finished_at_iso_or_dt)]，可乱序。返回 {run_id: [file,...]}。"""
    parsed_runs = []
    for run_id, finished in runs:
        if finished is None:
            continue
        dt = finished if isinstance(finished, datetime) else datetime.fromisoformat(str(finished).replace("Z", "")[:19])
        parsed_runs.append((dt, run_id))
    parsed_runs.sort()
    mapping: dict[Any, list[str]] = {}
    for name in files:
        t = _parse_export_timestamp(name)
        if t is None:
            continue
        chosen = next((rid for dt, rid in parsed_runs if dt >= t), None)
        if chosen is not None:
            mapping.setdefault(chosen, []).append(name)
    return mapping


def _latest_tplus_exports() -> list[dict[str, Any]]:
    directory = _tplus_export_dir()
    latest: dict[str, Path] = {}
    if directory.is_dir():
        for item in directory.glob("*.xlsx"):
            module = _tplus_module_of(item.name)
            current = latest.get(module)
            if current is None or item.name > current.name:
                latest[module] = item
    items: list[dict[str, Any]] = []
    for module in sorted(latest):
        path = latest[module]
        stat = path.stat()
        items.append(
            {
                "name": module,
                "file_name": path.name,
                "description": _tplus_export_description(module),
                "size_bytes": stat.st_size,
                "updated_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "download_url": f"/v1/exports/tplus/{path.name}",
            }
        )
    return items


def _external_source_tab_key(provider: str, env_profile: str) -> str:
    if provider == "feishu":
        return "feishu"
    return f"{provider}_{env_profile.lower()}"


@app.get("/v1/exports/catalog")
def exports_catalog(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    tabs: dict[str, dict[str, Any]] = {
        "tplus": {"key": "tplus", "title": "T+ ERP", "items": _latest_tplus_exports()},
        "wecom_company_a": {"key": "wecom_company_a", "title": "企微A", "items": []},
        "wecom_company_b": {"key": "wecom_company_b", "title": "企微B", "items": []},
        "feishu": {"key": "feishu", "title": "飞书", "items": []},
    }
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            # 按工作簿（文档）聚合：每个智能表格文档一条，下载时整簿多 sheet 导出。
            cur.execute(
                """
                SELECT s.provider, s.env_profile, s.external_doc_id,
                       COALESCE(MAX(NULLIF(s.document_name, '')), MAX(NULLIF(s.source_name, ''))) AS document_name,
                       COALESCE(MIN(s.id) FILTER (WHERE s.external_sheet_id = ''), MIN(s.id)) AS first_source_id,
                       COUNT(DISTINCT s.id) FILTER (WHERE s.external_sheet_id <> '') AS sheet_count,
                       COUNT(r.id) FILTER (WHERE s.external_sheet_id <> '') AS row_count,
                       MAX(s.last_sync_at) FILTER (WHERE s.external_sheet_id <> '') AS last_sync_at
                FROM external_sources s
                LEFT JOIN external_records r ON r.source_id = s.id AND s.external_sheet_id <> ''
                WHERE s.status = 'active' AND s.external_doc_id <> ''
                GROUP BY s.provider, s.env_profile, s.external_doc_id
                ORDER BY MIN(s.id)
                """
            )
            rows = cur.fetchall()
    for provider, env_profile, _doc_id, document_name, first_source_id, sheet_count, row_count, last_sync_at in rows:
        key = _external_source_tab_key(str(provider or ""), str(env_profile or ""))
        tab = tabs.setdefault(key, {"key": key, "title": key, "items": []})
        tab["items"].append(
            {
                "name": document_name or f"{provider} 文档",
                "source_id": first_source_id,
                "sheets": int(sheet_count or 0),
                "rows": int(row_count or 0),
                "updated_at": str(last_sync_at) if last_sync_at else None,
                "download_url": f"/v1/exports/external-doc/{first_source_id}" if int(sheet_count or 0) > 0 else None,
            }
        )
    return {"tabs": list(tabs.values())}


_STOCK_COLUMNS = {
    "WarehouseCode": "仓库编码",
    "WarehouseName": "仓库",
    "InventoryCode": "存货编码",
    "InventoryName": "存货名称",
    "InventoryClassName": "存货分类",
    "Specification": "规格型号",
    "UnitName": "单位",
    "ExistingQuantity": "现存量",
    "AvailableQuantity": "可用量",
}


def _latest_tplus_export_file(module: str) -> Path | None:
    directory = _tplus_export_dir()
    if not directory.is_dir():
        return None
    candidates = [item for item in directory.glob(f"{module}_*.xlsx") if _tplus_module_of(item.name) == module]
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.name)


# 库存页仓库范围：原材料=原材库(001)+L-代加工库(012)；成品=除原材库外全部。
_RAW_STOCK_WAREHOUSES = {"001", "012"}
_FINISHED_EXCLUDED_WAREHOUSES = {"001"}


@app.get("/v1/inventory/current-stock")
def inventory_current_stock(
    q: str = Query(default=""),
    warehouse: str = Query(default=""),
    scope: str = Query(default="raw"),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    if scope not in ("raw", "finished"):
        raise HTTPException(status_code=400, detail="scope must be raw or finished")
    roles = user.get("roles", [])
    permissions = user.get("permissions", [])
    scope_permission = "inventory.raw.read" if scope == "raw" else "inventory.finished.read"
    allowed = "admin" in roles or "admin.access" in permissions or scope_permission in permissions
    if not allowed:
        raise HTTPException(status_code=403, detail="permission denied")

    path = _latest_tplus_export_file("current_stock")
    if path is None:
        raise HTTPException(status_code=404, detail="现存量数据尚未同步，请先在 T+ 同步任务跑一轮全量。")

    import pandas as pd

    df = pd.read_excel(path, dtype=str)
    for column in _STOCK_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = df[list(_STOCK_COLUMNS)].fillna("")

    codes = df["WarehouseCode"].str.strip()
    if scope == "raw":
        df = df[codes.isin(_RAW_STOCK_WAREHOUSES)]
    else:
        df = df[~codes.isin(_FINISHED_EXCLUDED_WAREHOUSES)]

    warehouses = (
        df[["WarehouseCode", "WarehouseName"]]
        .drop_duplicates()
        .sort_values("WarehouseCode")
        .to_dict("records")
    )

    requested_warehouse = warehouse.strip()
    if requested_warehouse:
        if requested_warehouse not in {str(item["WarehouseCode"]).strip() for item in warehouses}:
            raise HTTPException(status_code=400, detail="warehouse not in scope")
        df = df[df["WarehouseCode"].str.strip() == requested_warehouse]
    keyword = q.strip()
    if keyword:
        lowered = keyword.lower()
        mask = (
            df["InventoryName"].str.lower().str.contains(lowered, na=False)
            | df["InventoryCode"].str.lower().str.contains(lowered, na=False)
            | df["Specification"].str.lower().str.contains(lowered, na=False)
        )
        df = df[mask]

    for column in ("ExistingQuantity", "AvailableQuantity"):
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)

    stat = path.stat()
    return {
        "items": df.to_dict("records"),
        "total": int(len(df)),
        "warehouses": warehouses,
        "source_file": path.name,
        "synced_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
    }


@app.get("/v1/exports/tplus/{file_name}")
def exports_tplus_download(file_name: str, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    if Path(file_name).name != file_name or not file_name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="invalid file name")
    path = _tplus_export_dir() / file_name
    if not path.is_file():
        raise HTTPException(status_code=404, detail="export file not found")
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=file_name)


def _routing_projects(channel: str) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT peer_id, display_name, project_url, project_name
                FROM managed_contacts
                WHERE channel = %s
                  AND enabled = true
                  AND COALESCE(project_url, '') <> ''
                ORDER BY peer_id
                """,
                (channel,),
            )
            rows = cur.fetchall()
    lanes: dict[str, dict[str, str]] = {}
    for peer_id, display_name, project_url, project_name in rows:
        if not peer_id or not project_url:
            continue
        lanes[str(peer_id)] = {
            "name": str(display_name or project_name or peer_id),
            "project_url": str(project_url),
        }
    return {"lanes": lanes}


@app.get("/v1/routing/wechat-projects.json")
def routing_wechat_projects() -> dict[str, Any]:
    return _routing_projects("wechat")


@app.get("/v1/routing/feishu-projects.json")
def routing_feishu_projects() -> dict[str, Any]:
    return _routing_projects("feishu")


@app.get("/v1/exports/external/{source_id}")
def exports_external_download(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT provider, env_profile, document_name, sheet_name FROM external_sources WHERE id = %s",
                (source_id,),
            )
            source = cur.fetchone()
            if not source:
                raise HTTPException(status_code=404, detail="external source not found")
            cur.execute(
                """
                SELECT external_record_id, normalized_json, external_updated_at, synced_at
                FROM external_records
                WHERE source_id = %s
                ORDER BY id
                """,
                (source_id,),
            )
            records = cur.fetchall()

    provider, env_profile, document_name, sheet_name = source
    columns: list[str] = []
    for _record_id, normalized, _ext_updated, _synced in records:
        if isinstance(normalized, dict):
            for column in normalized:
                if column not in columns:
                    columns.append(column)

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(sheet_name or "data")[:31] or "data"
    header = ["external_record_id", *columns, "external_updated_at", "synced_at"]
    sheet.append(header)
    for record_id, normalized, ext_updated, synced in records:
        data = normalized if isinstance(normalized, dict) else {}
        sheet.append(
            [record_id, *[data.get(column, "") for column in columns], ext_updated or "", str(synced or "")]
        )

    export_dir = Path(os.getenv("EXTERNAL_EXPORT_DIR", "/tmp/aliecs-external-exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"{provider}_{(env_profile or 'default').lower()}_{source_id}_{timestamp}.xlsx"
    workbook.save(path)
    label = f"{document_name or provider}-{sheet_name or source_id}"
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=f"{label}_{timestamp}.xlsx")


def _append_records_worksheet(workbook: Any, title: str, records: list[tuple]) -> None:
    columns: list[str] = []
    for _record_id, normalized, _ext_updated, _synced in records:
        if isinstance(normalized, dict):
            for column in normalized:
                if column not in columns:
                    columns.append(column)
    base = (str(title or "data").strip() or "data")[:31]
    name = base
    suffix = 2
    while name in workbook.sheetnames:
        name = f"{base[:28]}_{suffix}"
        suffix += 1
    sheet = workbook.create_sheet(title=name)
    sheet.append(["external_record_id", *columns, "external_updated_at", "synced_at"])
    for record_id, normalized, ext_updated, synced in records:
        data = normalized if isinstance(normalized, dict) else {}
        sheet.append([record_id, *[data.get(column, "") for column in columns], ext_updated or "", str(synced or "")])


@app.get("/v1/exports/external-doc/{source_id}")
def exports_external_doc_download(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    """按所属工作簿导出：同文档的全部 sheet 各占一个工作表。source_id 为该文档任一 sheet 级源。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT provider, env_profile, external_doc_id, document_name FROM external_sources WHERE id = %s",
                (source_id,),
            )
            anchor = cur.fetchone()
            if not anchor:
                raise HTTPException(status_code=404, detail="external source not found")
            provider, env_profile, external_doc_id, document_name = anchor
            cur.execute(
                """
                SELECT id, sheet_name
                FROM external_sources
                WHERE provider = %s AND env_profile = %s AND external_doc_id = %s
                  AND external_sheet_id <> '' AND status = 'active'
                ORDER BY id
                """,
                (provider, env_profile, external_doc_id),
            )
            sheet_sources = cur.fetchall()
            if not sheet_sources:
                raise HTTPException(status_code=404, detail="document has no active sheets")

            from openpyxl import Workbook

            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_source_id, sheet_name in sheet_sources:
                cur.execute(
                    """
                    SELECT external_record_id, normalized_json, external_updated_at, synced_at
                    FROM external_records
                    WHERE source_id = %s
                    ORDER BY id
                    """,
                    (sheet_source_id,),
                )
                _append_records_worksheet(workbook, str(sheet_name or sheet_source_id), cur.fetchall())

    export_dir = Path(os.getenv("EXTERNAL_EXPORT_DIR", "/tmp/aliecs-external-exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"{provider}_{(env_profile or 'default').lower()}_doc_{source_id}_{timestamp}.xlsx"
    workbook.save(path)
    label = str(document_name or provider)
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=f"{label}_{timestamp}.xlsx")


def _external_doc_anchor(cur: Any, source_id: int) -> tuple[str, str, str, str]:
    cur.execute(
        "SELECT provider, env_profile, external_doc_id, document_name FROM external_sources WHERE id = %s",
        (source_id,),
    )
    anchor = cur.fetchone()
    if not anchor:
        raise HTTPException(status_code=404, detail="external source not found")
    return anchor


def _ensure_doc_row(cur: Any, provider: str, env_profile: str, external_doc_id: str, document_name: str) -> int:
    """确保 doc 级登记行存在并返回其 id（worker 对 doc 级请求整簿重扫，含新 sheet 发现）。"""
    cur.execute(
        """
        INSERT INTO external_sources(
            provider, env_profile, source_name, source_type,
            external_doc_id, external_sheet_id, source_url,
            document_name, sheet_name, status, updated_at
        )
        VALUES (%s, %s, %s, 'smartsheet_doc', %s, '', '', %s, '', 'active', NOW())
        ON CONFLICT(provider, env_profile, external_doc_id, external_sheet_id)
        DO UPDATE SET status = 'active', updated_at = NOW()
        RETURNING id
        """,
        (provider, env_profile, document_name, external_doc_id, document_name),
    )
    return int(cur.fetchone()[0])


def _create_doc_sync_request(cur: Any, doc_row_id: int, provider: str, env_profile: str, requested_by: str) -> None:
    cur.execute(
        """
        INSERT INTO sync_requests(source_id, provider, env_profile, mode, status, requested_by)
        VALUES (%s, %s, %s, 'manual', 'pending', %s)
        """,
        (doc_row_id, provider, env_profile, requested_by),
    )


@app.post("/v1/exports/external-doc/{source_id}/sync-requests")
def exports_external_doc_sync(source_id: int, user: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """为该工作簿创建 doc 级同步请求（worker 整簿重扫，含新 sheet 发现），约 30 秒内开始。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            provider, env_profile, external_doc_id, document_name = _external_doc_anchor(cur, source_id)
            doc_row_id = _ensure_doc_row(cur, str(provider), str(env_profile), str(external_doc_id), str(document_name or ""))
            _create_doc_sync_request(cur, doc_row_id, str(provider), str(env_profile), str(user.get("sub") or ""))
        conn.commit()
    return {
        "document_name": document_name,
        "requests_created": 1,
        "message": f"已为「{document_name}」创建整簿同步请求，约 30 秒内开始同步。",
    }


@app.post("/v1/exports/sync-all")
def exports_sync_all(user: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """同步数据列表：为全部已登记文档各建一条 doc 级同步请求（发现新文档/新表/改名/新记录）。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, provider, env_profile FROM external_sources
                WHERE external_sheet_id = '' AND status = 'active' AND external_doc_id <> ''
                ORDER BY id
                """
            )
            doc_rows = cur.fetchall()
            for doc_row_id, provider, env_profile in doc_rows:
                _create_doc_sync_request(cur, int(doc_row_id), str(provider), str(env_profile), str(user.get("sub") or ""))
        conn.commit()
    return {
        "requests_created": len(doc_rows),
        "message": f"已为 {len(doc_rows)} 个文档创建同步请求，列表将在 1-2 分钟内陆续刷新。",
    }


@app.post("/v1/exports/external-doc/{source_id}/copy")
def exports_external_doc_copy(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """在企业微信中创建该智能表格的完整副本（全部工作表结构 + 全部记录）。"""
    from app.integrations.wecom_docs import WeComDocError, copy_smartsheet_doc

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            provider, env_profile, external_doc_id, document_name = _external_doc_anchor(cur, source_id)
    if provider != "wecom":
        raise HTTPException(status_code=400, detail="仅支持企业微信智能表格创建副本")

    new_name = f"{document_name or '智能表格'}-副本{datetime.now(tz=timezone.utc).strftime('%y%m%d_%H%M')}"
    try:
        result = copy_smartsheet_doc(env_profile=str(env_profile), source_docid=str(external_doc_id), new_doc_name=new_name)
    except WeComDocError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    # 新副本自动登记为同步源并触发首次整簿同步，30 秒内出现在数据导出列表。
    new_docid = str(result.get("new_docid") or "")
    if new_docid:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                doc_row_id = _ensure_doc_row(cur, str(provider), str(env_profile), new_docid, new_name)
                _create_doc_sync_request(cur, doc_row_id, str(provider), str(env_profile), "copy-auto")
            conn.commit()
        result["registered"] = True
    return {"document_name": document_name, "new_doc_name": new_name, **result}


@app.get("/v1/admin/rbac-overview")
def admin_rbac_overview(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name, description FROM roles ORDER BY id")
            roles = cur.fetchall()
            cur.execute(
                """
                SELECT r.id, p.code
                FROM roles r
                LEFT JOIN role_permissions rp ON rp.role_id = r.id
                LEFT JOIN permissions p ON p.id = rp.permission_id
                ORDER BY r.id, p.code
                """
            )
            rows = cur.fetchall()

    mapping: dict[int, list[str]] = {}
    for role_id, perm_code in rows:
        mapping.setdefault(role_id, [])
        if perm_code:
            mapping[role_id].append(perm_code)

    return {
        "roles": [
            {
                "id": r[0],
                "code": r[1],
                "name": r[2],
                "description": r[3],
                "permissions": mapping.get(r[0], []),
            }
            for r in roles
        ]
    }


@app.get("/couple/access")
def couple_access(authorization: str | None = Header(default=None)) -> dict[str, Any] | None:
    if not authorization:
        raise HTTPException(status_code=404, detail="not found")

    try:
        user = get_current_user(authorization)
    except HTTPException:
        raise HTTPException(status_code=404, detail="not found")

    if not _has_couple_access(user):
        raise HTTPException(status_code=404, detail="not found")

    return {"allowed": True, "route": _couple_route()}


@app.get("/v1/immich/status")
def immich_status(user: dict[str, Any] = Depends(require_login)) -> dict[str, object]:
    require_permission("couple_memory_access", user)
    from app.immich_client import ImmichClient

    return ImmichClient().status()


def _public_immich_thumbnail_url(asset_id: str) -> str:
    public_base = os.getenv("APP_BASE_URL", "").rstrip("/")
    relative_url = f"/api/v1/immich/assets/{urllib.parse.quote(asset_id, safe='')}/thumbnail"
    return f"{public_base}{relative_url}" if public_base else relative_url


@app.get("/v1/immich/assets")
def immich_assets(
    query: str | None = Query(default=None),
    taken_after: str | None = Query(default=None),
    taken_before: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    require_permission("couple_memory_access", user)
    from app.immich_client import ImmichClient, load_immich_config

    config = load_immich_config()
    if not config.enabled:
        return {"enabled": False, "items": []}
    try:
        assets = ImmichClient(config).search_assets(
            query=query,
            taken_after=taken_after,
            taken_before=taken_before,
            page=page,
        )
    except Exception as exc:
        return {"enabled": True, "items": [], "detail": str(exc)}
    return {
        "enabled": True,
        "items": [
            {
                "asset_id": asset.asset_id,
                "original_filename": asset.original_filename,
                "taken_at": asset.taken_at,
                "latitude": asset.latitude,
                "longitude": asset.longitude,
                "thumbnail_url": _public_immich_thumbnail_url(asset.asset_id),
            }
            for asset in assets
            if asset.asset_id
        ],
    }


@app.get("/v1/immich/assets/{asset_id}/thumbnail")
def immich_asset_thumbnail(asset_id: str, user: dict[str, Any] = Depends(require_login)) -> Response:
    require_permission("couple_memory_access", user)
    from app.immich_client import ImmichClient

    try:
        content, content_type = ImmichClient().get_thumbnail(asset_id)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return Response(content=content, media_type=content_type, headers={"Cache-Control": "private, max-age=3600"})


def _user_id_by_username(username: str) -> int | None:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE username = %s", (username,))
            row = cur.fetchone()
    return row[0] if row else None


def _resolve_couple_space_id(user_id: int, requested_space_id: int | None) -> int:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            if requested_space_id:
                cur.execute(
                    """
                    SELECT couple_space_id
                    FROM couple_members
                    WHERE user_id = %s AND couple_space_id = %s
                    """,
                    (user_id, requested_space_id),
                )
                row = cur.fetchone()
                if row:
                    return row[0]
                raise HTTPException(status_code=403, detail="permission denied")

            cur.execute(
                """
                SELECT couple_space_id
                FROM couple_members
                WHERE user_id = %s
                ORDER BY joined_at ASC
                LIMIT 1
                """,
                (user_id,),
            )
            row = cur.fetchone()
            if row:
                return row[0]
    raise HTTPException(status_code=404, detail="not found")


def _require_couple_user(user: dict[str, Any]) -> int:
    if not _has_couple_access(user):
        raise HTTPException(status_code=403, detail="permission denied")
    user_id = _user_id_by_username(str(user.get("sub", "")))
    if not user_id:
        raise HTTPException(status_code=401, detail="invalid user")
    return user_id


def _is_space_owner(user: dict[str, Any], user_id: int, space_id: int) -> bool:
    roles = [str(r).lower() for r in user.get("roles", [])]
    permissions = [str(p).lower() for p in user.get("permissions", [])]
    if "admin" in roles or "admin.access" in permissions:
        return True
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT role
                FROM couple_members
                WHERE couple_space_id = %s AND user_id = %s
                LIMIT 1
                """,
                (space_id, user_id),
            )
            row = cur.fetchone()
    return bool(row and row[0] == "owner")


def _parse_date_or_none(value: str | date | None) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="invalid date") from exc


def _normalize_tags(tags: list[str]) -> list[str]:
    seen: set[str] = set()
    cleaned: list[str] = []
    for raw in tags:
        tag = str(raw).strip()
        if not tag or tag in seen:
            continue
        seen.add(tag)
        cleaned.append(tag[:40])
    return cleaned


def _safe_date(year: int, month: int, day: int) -> date:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def _next_anniversary_occurrence(base_date: date, repeat_type: str, today: date | None = None) -> date | None:
    today = today or date.today()
    if repeat_type == "none":
        return base_date if base_date >= today else None
    if repeat_type == "yearly":
        current = _safe_date(today.year, base_date.month, base_date.day)
        if current < today:
            current = _safe_date(today.year + 1, base_date.month, base_date.day)
        return current
    if repeat_type == "monthly":
        current = _safe_date(today.year, today.month, base_date.day)
        if current < today:
            year = today.year + (1 if today.month == 12 else 0)
            month = 1 if today.month == 12 else today.month + 1
            current = _safe_date(year, month, base_date.day)
        return current
    raise HTTPException(status_code=400, detail="invalid repeat_type")


def _anniversary_payload(row: tuple[Any, ...], today: date | None = None) -> dict[str, Any] | None:
    occurrence = _next_anniversary_occurrence(row[2], row[3], today)
    if occurrence is None:
        return None
    today = today or date.today()
    return {
        "id": row[0],
        "title": row[1],
        "date": str(row[2]),
        "repeat_type": row[3],
        "description": row[4],
        "next_occurrence": str(occurrence),
        "days_remaining": (occurrence - today).days,
    }


def _share_base_url() -> str:
    return os.getenv("SHARE_BASE_URL", os.getenv("APP_BASE_URL", "")).rstrip("/")


_share_hits: dict[str, list[float]] = {}


def _check_share_rate(token: str) -> None:
    now = time.monotonic()
    hits = [ts for ts in _share_hits.get(token, []) if now - ts < 60]
    hits.append(now)
    _share_hits[token] = hits
    if len(hits) > 120:
        raise HTTPException(status_code=429, detail="too many requests")


@app.get("/v1/memories")
def list_memories(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=200),
    couple_space_id: int | None = Query(default=None),
    archived: str = Query(default="false"),
    tag: str | None = Query(default=None),
    from_date: str | None = Query(default=None, alias="from"),
    to_date: str | None = Query(default=None, alias="to"),
    q: str | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    offset = (page - 1) * page_size
    where = ["m.couple_space_id = %s"]
    params: list[Any] = [space_id]
    archived_value = archived.strip().lower()
    if archived_value not in {"all", "true", "1", "yes"}:
        where.append("m.archived = false")
    elif archived_value in {"true", "1", "yes"}:
        where.append("m.archived = true")
    if tag:
        where.append("EXISTS (SELECT 1 FROM memory_tags mt WHERE mt.memory_id = m.id AND mt.tag = %s)")
        params.append(tag.strip())
    start = _parse_date_or_none(from_date)
    end = _parse_date_or_none(to_date)
    if start:
        where.append("m.memory_date >= %s")
        params.append(start)
    if end:
        where.append("m.memory_date <= %s")
        params.append(end)
    if q and q.strip():
        where.append("(m.title ILIKE %s OR COALESCE(m.content, '') ILIKE %s)")
        pattern = f"%{q.strip()}%"
        params.extend([pattern, pattern])
    where_sql = " AND ".join(where)

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM memories m WHERE {where_sql}", params)
            total = cur.fetchone()[0]
            cur.execute(
                f"""
                SELECT id, couple_space_id, title, content, memory_date, place_name, latitude, longitude,
                       cover_photo_url, visibility, created_by, created_at, updated_at, archived,
                       COALESCE((SELECT COUNT(*) FROM photos p WHERE p.memory_id = m.id), 0) AS photo_count
                FROM memories m
                WHERE {where_sql}
                ORDER BY memory_date DESC NULLS LAST, id DESC
                LIMIT %s OFFSET %s
                """,
                [*params, page_size, offset],
            )
            rows = cur.fetchall()

            memory_ids = [row[0] for row in rows]
            tags_map: dict[int, list[str]] = {mid: [] for mid in memory_ids}
            if memory_ids:
                cur.execute(
                    "SELECT memory_id, tag FROM memory_tags WHERE memory_id = ANY(%s::bigint[]) ORDER BY id",
                    (memory_ids,),
                )
                for mid, tag in cur.fetchall():
                    tags_map.setdefault(mid, []).append(tag)

    return {
        "items": [
            {
                "id": row[0],
                "couple_space_id": row[1],
                "title": row[2],
                "content": row[3],
                "memory_date": str(row[4]) if row[4] else None,
                "place_name": row[5],
                "latitude": row[6],
                "longitude": row[7],
                "cover_photo_url": row[8],
                "visibility": row[9],
                "created_by": row[10],
                "created_at": str(row[11]),
                "updated_at": str(row[12]),
                "archived": row[13],
                "photo_count": row[14],
                "tags": tags_map.get(row[0], []),
            }
            for row in rows
        ],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": (total + page_size - 1) // page_size,
    }


@app.post("/v1/memories")
def create_memory(body: MemoryUpsertRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, body.couple_space_id)
    if body.visibility not in {"private", "shareable"}:
        raise HTTPException(status_code=400, detail="invalid visibility")
    memory_date = _parse_date_or_none(body.memory_date)

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO memories(
                    couple_space_id, title, content, memory_date, place_name, latitude, longitude,
                    cover_photo_url, visibility, created_by
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    space_id,
                    body.title,
                    body.content,
                    memory_date,
                    body.place_name,
                    body.latitude,
                    body.longitude,
                    body.cover_photo_url,
                    body.visibility,
                    user_id,
                ),
            )
            memory_id = cur.fetchone()[0]
            for tag_clean in _normalize_tags(body.tags):
                cur.execute(
                    "INSERT INTO memory_tags(memory_id, tag) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    (memory_id, tag_clean),
                )
        conn.commit()
    return {"id": memory_id}


@app.get("/v1/memories/{memory_id}")
def get_memory(memory_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT m.id, m.couple_space_id, m.title, m.content, m.memory_date, m.place_name, m.latitude, m.longitude,
                       m.cover_photo_url, m.visibility, m.created_by, m.created_at, m.updated_at, m.archived,
                       COALESCE((SELECT COUNT(*) FROM photos p WHERE p.memory_id = m.id), 0) AS photo_count
                FROM memories m
                JOIN couple_members cm ON cm.couple_space_id = m.couple_space_id
                WHERE m.id = %s AND cm.user_id = %s
                LIMIT 1
                """,
                (memory_id, user_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
            cur.execute("SELECT tag FROM memory_tags WHERE memory_id = %s ORDER BY id", (memory_id,))
            tags = [r[0] for r in cur.fetchall()]
    return {
        "id": row[0],
        "couple_space_id": row[1],
        "title": row[2],
        "content": row[3],
        "memory_date": str(row[4]) if row[4] else None,
        "place_name": row[5],
        "latitude": row[6],
        "longitude": row[7],
        "cover_photo_url": row[8],
        "visibility": row[9],
        "created_by": row[10],
        "created_at": str(row[11]),
        "updated_at": str(row[12]),
        "archived": row[13],
        "photo_count": row[14],
        "tags": tags,
    }


def _memory_space_for_user(cur: Any, memory_id: int, user_id: int) -> int:
    cur.execute(
        """
        SELECT m.couple_space_id
        FROM memories m
        JOIN couple_members cm ON cm.couple_space_id = m.couple_space_id
        WHERE m.id = %s AND cm.user_id = %s
        LIMIT 1
        """,
        (memory_id, user_id),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="not found")
    return int(row[0])


def _immich_asset_payload(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "id": row[0],
        "couple_space_id": row[1],
        "memory_id": row[2],
        "provider": row[3],
        "immich_asset_id": row[4],
        "immich_album_id": row[5],
        "original_filename": row[6],
        "taken_at": str(row[7]) if row[7] else None,
        "latitude": row[8],
        "longitude": row[9],
        "thumbnail_cache_key": row[10],
        "sort_order": row[11],
        "selected_by": row[12],
        "created_at": str(row[13]),
        "updated_at": str(row[14]),
    }


@app.post("/v1/memories/{memory_id}/immich-assets")
def bind_memory_immich_asset(
    memory_id: int,
    body: ImmichAssetBindRequest,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    asset_ids = [asset_id.strip() for asset_id in body.asset_ids if asset_id and asset_id.strip()]
    if asset_ids:
        items = [
            _bind_memory_immich_asset(
                memory_id,
                body.model_copy(update={"asset_ids": [], "immich_asset_id": asset_id}),
                user,
            )
            for asset_id in dict.fromkeys(asset_ids)
        ]
        return {"items": items}
    return _bind_memory_immich_asset(memory_id, body, user)


def _bind_memory_immich_asset(
    memory_id: int,
    body: ImmichAssetBindRequest,
    user: dict[str, Any],
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    if not body.immich_asset_id and not body.immich_album_id:
        raise HTTPException(status_code=400, detail="immich_asset_id or immich_album_id is required")

    from app.immich_client import ImmichClient, load_immich_config

    config = load_immich_config()
    original_filename = body.original_filename
    taken_at = body.taken_at
    latitude = body.latitude
    longitude = body.longitude

    if config.enabled and body.immich_asset_id:
        asset = ImmichClient(config).get_asset(body.immich_asset_id)
        original_filename = original_filename or asset.original_filename
        taken_at = taken_at or asset.taken_at
        latitude = latitude if latitude is not None else asset.latitude
        longitude = longitude if longitude is not None else asset.longitude
    elif not original_filename:
        raise HTTPException(status_code=503, detail="Immich integration disabled")

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            space_id = _memory_space_for_user(cur, memory_id, user_id)
            cur.execute(
                """
                INSERT INTO couple_memory_assets(
                    couple_space_id, memory_id, provider, immich_asset_id, immich_album_id,
                    original_filename, taken_at, latitude, longitude, thumbnail_cache_key,
                    sort_order, selected_by
                )
                VALUES (%s, %s, 'immich', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (memory_id, provider, immich_asset_id) WHERE immich_asset_id IS NOT NULL
                DO UPDATE SET
                    original_filename = COALESCE(EXCLUDED.original_filename, couple_memory_assets.original_filename),
                    taken_at = COALESCE(EXCLUDED.taken_at, couple_memory_assets.taken_at),
                    latitude = COALESCE(EXCLUDED.latitude, couple_memory_assets.latitude),
                    longitude = COALESCE(EXCLUDED.longitude, couple_memory_assets.longitude),
                    thumbnail_cache_key = COALESCE(EXCLUDED.thumbnail_cache_key, couple_memory_assets.thumbnail_cache_key),
                    sort_order = EXCLUDED.sort_order,
                    selected_by = EXCLUDED.selected_by,
                    updated_at = NOW()
                RETURNING id, couple_space_id, memory_id, provider, immich_asset_id, immich_album_id,
                          original_filename, taken_at, latitude, longitude, thumbnail_cache_key,
                          sort_order, selected_by, created_at, updated_at
                """,
                (
                    space_id,
                    memory_id,
                    body.immich_asset_id,
                    body.immich_album_id,
                    original_filename,
                    taken_at,
                    latitude,
                    longitude,
                    body.thumbnail_cache_key,
                    body.sort_order,
                    user_id,
                ),
            )
            row = cur.fetchone()
        conn.commit()
    return _immich_asset_payload(row)


@app.get("/v1/memories/{memory_id}/immich-assets")
def list_memory_immich_assets(
    memory_id: int,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            _memory_space_for_user(cur, memory_id, user_id)
            cur.execute(
                """
                SELECT id, couple_space_id, memory_id, provider, immich_asset_id, immich_album_id,
                       original_filename, taken_at, latitude, longitude, thumbnail_cache_key,
                       sort_order, selected_by, created_at, updated_at
                FROM couple_memory_assets
                WHERE memory_id = %s
                ORDER BY sort_order ASC, id ASC
                """,
                (memory_id,),
            )
            rows = cur.fetchall()
    return {"items": [_immich_asset_payload(row) for row in rows]}


@app.delete("/v1/memories/{memory_id}/immich-assets/{binding_id}")
def delete_memory_immich_asset(
    memory_id: int,
    binding_id: int,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            _memory_space_for_user(cur, memory_id, user_id)
            cur.execute(
                """
                DELETE FROM couple_memory_assets
                WHERE id = %s AND memory_id = %s
                RETURNING id
                """,
                (binding_id, memory_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.put("/v1/memories/{memory_id}")
def update_memory(memory_id: int, body: MemoryUpsertRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    user_id = _require_couple_user(user)
    if body.visibility not in {"private", "shareable"}:
        raise HTTPException(status_code=400, detail="invalid visibility")
    memory_date = _parse_date_or_none(body.memory_date)

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT m.couple_space_id
                FROM memories m
                JOIN couple_members cm ON cm.couple_space_id = m.couple_space_id
                WHERE m.id = %s AND cm.user_id = %s
                LIMIT 1
                """,
                (memory_id, user_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
            cur.execute(
                """
                UPDATE memories
                SET title=%s, content=%s, memory_date=%s, place_name=%s, latitude=%s, longitude=%s,
                    cover_photo_url=%s, visibility=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (
                    body.title,
                    body.content,
                    memory_date,
                    body.place_name,
                    body.latitude,
                    body.longitude,
                    body.cover_photo_url,
                    body.visibility,
                    memory_id,
                ),
            )
            cur.execute("DELETE FROM memory_tags WHERE memory_id = %s", (memory_id,))
            for tag_clean in _normalize_tags(body.tags):
                cur.execute(
                    "INSERT INTO memory_tags(memory_id, tag) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    (memory_id, tag_clean),
                )
        conn.commit()
    return {"status": "ok"}


@app.post("/v1/memories/{memory_id}/archive")
def archive_memory(memory_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    return _set_memory_archived(memory_id, True, user)


@app.post("/v1/memories/{memory_id}/unarchive")
def unarchive_memory(memory_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    return _set_memory_archived(memory_id, False, user)


def _set_memory_archived(memory_id: int, archived: bool, user: dict[str, Any]) -> dict[str, str]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE memories m
                SET archived = %s, updated_at = NOW()
                FROM couple_members cm
                WHERE m.id = %s AND cm.user_id = %s AND cm.couple_space_id = m.couple_space_id
                """,
                (archived, memory_id, user_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.delete("/v1/memories/{memory_id}")
def delete_memory(memory_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM memories m
                USING couple_members cm
                WHERE m.id = %s AND cm.user_id = %s AND cm.couple_space_id = m.couple_space_id
                """,
                (memory_id, user_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.get("/v1/map/memories")
def map_memories(
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, title, memory_date, place_name, latitude, longitude, cover_photo_url
                FROM memories
                WHERE couple_space_id = %s AND archived = false AND latitude IS NOT NULL AND longitude IS NOT NULL
                ORDER BY memory_date DESC NULLS LAST, id DESC
                """,
                (space_id,),
            )
            rows = cur.fetchall()
            memory_ids = [row[0] for row in rows]
            tags_map: dict[int, list[str]] = {mid: [] for mid in memory_ids}
            if memory_ids:
                cur.execute(
                    "SELECT memory_id, tag FROM memory_tags WHERE memory_id = ANY(%s::bigint[]) ORDER BY id",
                    (memory_ids,),
                )
                for mid, tag_value in cur.fetchall():
                    tags_map.setdefault(mid, []).append(tag_value)
    return {
        "items": [
            {
                "id": row[0],
                "title": row[1],
                "memory_date": str(row[2]) if row[2] else None,
                "place_name": row[3],
                "latitude": row[4],
                "longitude": row[5],
                "cover_photo_url": row[6],
                "tags": tags_map.get(row[0], []),
            }
            for row in rows
        ]
    }


@app.post("/v1/photos/upload")
async def upload_photo(
    file: UploadFile = File(...),
    memory_id: int | None = Query(default=None),
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    if memory_id is not None:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 1
                    FROM memories
                    WHERE id = %s AND couple_space_id = %s
                    LIMIT 1
                    """,
                    (memory_id, space_id),
                )
                if not cur.fetchone():
                    raise HTTPException(status_code=400, detail="memory_id not in current couple space")
    saved = await photo_storage().save(file)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO photos(
                    couple_space_id, memory_id, original_filename, original_storage_url,
                    thumbnail_url, display_url, storage_driver
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id, created_at
                """,
                (
                    space_id,
                    memory_id,
                    file.filename,
                    saved["original_storage_url"],
                    saved["thumbnail_url"],
                    saved["display_url"],
                    saved["storage_driver"],
                ),
            )
            row = cur.fetchone()
        conn.commit()
    return {
        "id": row[0],
        "created_at": str(row[1]),
        "memory_id": memory_id,
        "display_url": saved["display_url"],
        "thumbnail_url": saved["thumbnail_url"],
        "storage_driver": saved["storage_driver"],
    }


@app.get("/v1/photos")
def list_photos(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    memory_id: int | None = Query(default=None),
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    offset = (page - 1) * page_size
    where = ["p.couple_space_id = %s"]
    params: list[Any] = [space_id]
    if memory_id is not None:
        where.append("p.memory_id = %s")
        params.append(memory_id)
    where_sql = " AND ".join(where)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM photos p WHERE {where_sql}", params)
            total = cur.fetchone()[0]
            cur.execute(
                f"""
                SELECT p.id, p.memory_id, p.original_filename, p.display_url, p.thumbnail_url, p.taken_at, p.created_at,
                       m.title AS memory_title
                FROM photos p
                LEFT JOIN memories m ON m.id = p.memory_id
                WHERE {where_sql}
                ORDER BY COALESCE(p.taken_at, p.created_at) DESC, p.id DESC
                LIMIT %s OFFSET %s
                """,
                [*params, page_size, offset],
            )
            rows = cur.fetchall()
    return {
        "items": [
            {
                "id": row[0],
                "memory_id": row[1],
                "original_filename": row[2],
                "display_url": row[3],
                "thumbnail_url": row[4],
                "taken_at": str(row[5]) if row[5] else None,
                "created_at": str(row[6]),
                "memory_title": row[7],
            }
            for row in rows
        ],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": (total + page_size - 1) // page_size,
    }


@app.get("/v1/photos/content/{key}")
def get_photo_content(key: str) -> Response:
    content = _webdock_photo_request("GET", key)
    mime = _detect_image_mime(content) or "application/octet-stream"
    return Response(
        content=content,
        media_type=mime,
        headers={"Cache-Control": "private, max-age=86400"},
    )


@app.get("/v1/photos/{photo_id}")
def get_photo(photo_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT p.id, p.memory_id, p.original_filename, p.original_storage_url, p.display_url, p.thumbnail_url,
                       p.storage_driver, p.created_at, m.title AS memory_title
                FROM photos p
                LEFT JOIN memories m ON m.id = p.memory_id
                JOIN couple_members cm ON cm.couple_space_id = p.couple_space_id
                WHERE p.id = %s AND cm.user_id = %s
                LIMIT 1
                """,
                (photo_id, user_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
    return {
        "id": row[0],
        "memory_id": row[1],
        "original_filename": row[2],
        "original_storage_url": row[3],
        "display_url": row[4],
        "thumbnail_url": row[5],
        "storage_driver": row[6],
        "created_at": str(row[7]),
        "memory_title": row[8],
    }


@app.delete("/v1/photos/{photo_id}")
def delete_photo(photo_id: int, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM photos p
                USING couple_members cm
                WHERE p.id = %s AND cm.user_id = %s AND cm.couple_space_id = p.couple_space_id
                RETURNING p.original_storage_url, p.storage_driver
                """,
                (photo_id, user_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    try:
        if row[1] == "local":
            LocalPhotoStorage().delete(row[0])
        elif row[1] == "webdock":
            WebDockPhotoStorage().delete(row[0])
    except Exception:
        pass
    return {"status": "ok"}


@app.get("/v1/couple/space")
def get_couple_space(
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    today = date.today()
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, name, start_date, theme, cover_image_url
                FROM couple_spaces
                WHERE id = %s
                """,
                (space_id,),
            )
            space = cur.fetchone()
            if not space:
                raise HTTPException(status_code=404, detail="not found")
            cur.execute(
                """
                SELECT u.id, u.display_name, u.username, cm.role
                FROM couple_members cm
                JOIN users u ON u.id = cm.user_id
                WHERE cm.couple_space_id = %s
                ORDER BY cm.role DESC, cm.joined_at ASC
                """,
                (space_id,),
            )
            members = cur.fetchall()
            counts: dict[str, int] = {}
            for key, table in [
                ("memories", "memories"),
                ("photos", "photos"),
                ("anniversaries", "anniversaries"),
                ("bucket_items", "bucket_items"),
            ]:
                cur.execute(f"SELECT COUNT(*) FROM {table} WHERE couple_space_id = %s", (space_id,))
                counts[key] = cur.fetchone()[0]
            cur.execute(
                """
                SELECT id, title, date, repeat_type, description
                FROM anniversaries
                WHERE couple_space_id = %s
                """,
                (space_id,),
            )
            anniversary_rows = cur.fetchall()

    next_items = [item for row in anniversary_rows if (item := _anniversary_payload(row, today))]
    next_items.sort(key=lambda item: (item["days_remaining"], item["id"]))
    start_date = space[2]
    return {
        "id": space[0],
        "name": space[1],
        "start_date": str(start_date) if start_date else None,
        "theme": space[3],
        "cover_image_url": space[4],
        "days_together": (today - start_date).days if start_date else None,
        "members": [
            {
                "user_id": row[0],
                "display_name": row[1] or row[2],
                "username": row[2],
                "role": row[3],
            }
            for row in members
        ],
        "next_anniversary": next_items[0] if next_items else None,
        "counts": counts,
    }


@app.patch("/v1/couple/space")
def patch_couple_space(
    body: PatchCoupleSpaceRequest,
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    if not _is_space_owner(user, user_id, space_id):
        raise HTTPException(status_code=403, detail="permission denied")
    fields: list[str] = []
    params: list[Any] = []
    if body.name is not None:
        name = body.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="name is required")
        fields.append("name = %s")
        params.append(name)
    if body.start_date is not None:
        fields.append("start_date = %s")
        params.append(_parse_date_or_none(body.start_date))
    if body.theme is not None:
        fields.append("theme = %s")
        params.append(body.theme.strip() or None)
    if body.cover_image_url is not None:
        fields.append("cover_image_url = %s")
        params.append(body.cover_image_url.strip() or None)
    if not fields:
        return {"status": "ok"}
    params.append(space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE couple_spaces SET {', '.join(fields)}, updated_at = NOW() WHERE id = %s",
                params,
            )
        conn.commit()
    return {"status": "ok"}


@app.get("/v1/tags")
def list_tags(
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT mt.tag, COUNT(*) AS count
                FROM memory_tags mt
                JOIN memories m ON m.id = mt.memory_id
                WHERE m.couple_space_id = %s
                GROUP BY mt.tag
                ORDER BY count DESC, mt.tag ASC
                """,
                (space_id,),
            )
            rows = cur.fetchall()
    return {"items": [{"tag": row[0], "count": row[1]} for row in rows]}


@app.get("/v1/anniversaries")
def list_anniversaries(
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    today = date.today()
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, title, date, repeat_type, description
                FROM anniversaries
                WHERE couple_space_id = %s
                ORDER BY date ASC, id ASC
                """,
                (space_id,),
            )
            rows = cur.fetchall()
    items = [item for row in rows if (item := _anniversary_payload(row, today))]
    items.sort(key=lambda item: (item["days_remaining"], item["id"]))
    return {"items": items}


@app.get("/v1/anniversaries/next")
def next_anniversary(
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any] | None:
    items = list_anniversaries(couple_space_id, user)["items"]
    return items[0] if items else None


@app.post("/v1/anniversaries")
def create_anniversary(body: AnniversaryUpsertRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, body.couple_space_id)
    if body.repeat_type not in {"none", "yearly", "monthly"}:
        raise HTTPException(status_code=400, detail="invalid repeat_type")
    title = body.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO anniversaries(couple_space_id, title, date, repeat_type, description)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
                """,
                (space_id, title, _parse_date_or_none(body.date), body.repeat_type, body.description),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id}


@app.put("/v1/anniversaries/{anniversary_id}")
def update_anniversary(
    anniversary_id: int,
    body: AnniversaryUpsertRequest,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, body.couple_space_id)
    if body.repeat_type not in {"none", "yearly", "monthly"}:
        raise HTTPException(status_code=400, detail="invalid repeat_type")
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE anniversaries
                SET title = %s, date = %s, repeat_type = %s, description = %s, updated_at = NOW()
                WHERE id = %s AND couple_space_id = %s
                """,
                (body.title.strip(), _parse_date_or_none(body.date), body.repeat_type, body.description, anniversary_id, space_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.delete("/v1/anniversaries/{anniversary_id}")
def delete_anniversary(
    anniversary_id: int,
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM anniversaries WHERE id = %s AND couple_space_id = %s", (anniversary_id, space_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.get("/v1/bucket-items")
def list_bucket_items(
    status: str | None = Query(default=None),
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    where = ["b.couple_space_id = %s"]
    params: list[Any] = [space_id]
    if status:
        if status not in {"want", "planned", "done"}:
            raise HTTPException(status_code=400, detail="invalid status")
        where.append("b.status = %s")
        params.append(status)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT b.id, b.title, b.description, b.status, b.target_date, b.completed_memory_id,
                       m.title AS completed_memory_title
                FROM bucket_items b
                LEFT JOIN memories m ON m.id = b.completed_memory_id
                WHERE {' AND '.join(where)}
                ORDER BY CASE b.status WHEN 'planned' THEN 1 WHEN 'want' THEN 2 ELSE 3 END,
                         b.target_date ASC NULLS LAST, b.id DESC
                """,
                params,
            )
            rows = cur.fetchall()
    return {
        "items": [
            {
                "id": row[0],
                "title": row[1],
                "description": row[2],
                "status": row[3],
                "target_date": str(row[4]) if row[4] else None,
                "completed_memory_id": row[5],
                "completed_memory_title": row[6],
            }
            for row in rows
        ]
    }


def _validate_bucket_body(body: BucketItemUpsertRequest, space_id: int) -> None:
    if body.status not in {"want", "planned", "done"}:
        raise HTTPException(status_code=400, detail="invalid status")
    if body.completed_memory_id is not None:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT 1 FROM memories WHERE id = %s AND couple_space_id = %s LIMIT 1",
                    (body.completed_memory_id, space_id),
                )
                if not cur.fetchone():
                    raise HTTPException(status_code=400, detail="completed_memory_id not in current couple space")


@app.post("/v1/bucket-items")
def create_bucket_item(body: BucketItemUpsertRequest, user: dict[str, Any] = Depends(require_login)) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, body.couple_space_id)
    _validate_bucket_body(body, space_id)
    title = body.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO bucket_items(couple_space_id, title, description, status, target_date, completed_memory_id)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    space_id,
                    title,
                    body.description,
                    body.status,
                    _parse_date_or_none(body.target_date),
                    body.completed_memory_id,
                ),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id}


@app.put("/v1/bucket-items/{item_id}")
def update_bucket_item(
    item_id: int,
    body: BucketItemUpsertRequest,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, body.couple_space_id)
    _validate_bucket_body(body, space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE bucket_items
                SET title = %s, description = %s, status = %s, target_date = %s,
                    completed_memory_id = %s, updated_at = NOW()
                WHERE id = %s AND couple_space_id = %s
                """,
                (
                    body.title.strip(),
                    body.description,
                    body.status,
                    _parse_date_or_none(body.target_date),
                    body.completed_memory_id,
                    item_id,
                    space_id,
                ),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.delete("/v1/bucket-items/{item_id}")
def delete_bucket_item(
    item_id: int,
    couple_space_id: int | None = Query(default=None),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, str]:
    user_id = _require_couple_user(user)
    space_id = _resolve_couple_space_id(user_id, couple_space_id)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM bucket_items WHERE id = %s AND couple_space_id = %s", (item_id, space_id))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.post("/v1/memories/{memory_id}/share")
def create_memory_share(
    memory_id: int,
    body: CreateShareLinkRequest,
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT m.couple_space_id
                FROM memories m
                JOIN couple_members cm ON cm.couple_space_id = m.couple_space_id
                WHERE m.id = %s AND cm.user_id = %s
                LIMIT 1
                """,
                (memory_id, user_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="not found")
            token = secrets.token_urlsafe(24)
            expires_at = None
            if body.expires_in_days:
                expires_at = datetime.now(timezone.utc) + timedelta(days=body.expires_in_days)
            cur.execute(
                """
                INSERT INTO share_links(couple_space_id, memory_id, token, expires_at)
                VALUES (%s, %s, %s, %s)
                """,
                (row[0], memory_id, token, expires_at),
            )
            cur.execute("UPDATE memories SET visibility = 'shareable', updated_at = NOW() WHERE id = %s", (memory_id,))
        conn.commit()
    url = f"{_share_base_url()}/s/{token}"
    return {"token": token, "url": url, "expires_at": expires_at.isoformat() if expires_at else None}


@app.delete("/v1/share/{token}")
def revoke_share_link(token: str, user: dict[str, Any] = Depends(require_login)) -> dict[str, str]:
    user_id = _require_couple_user(user)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE share_links sl
                SET revoked_at = NOW()
                FROM couple_members cm
                WHERE sl.token = %s AND cm.user_id = %s AND cm.couple_space_id = sl.couple_space_id
                """,
                (token, user_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="not found")
        conn.commit()
    return {"status": "ok"}


@app.get("/v1/share/{token}")
def get_shared_memory(token: str) -> dict[str, Any]:
    _check_share_rate(token)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT m.id, m.title, m.content, m.memory_date, m.place_name, m.latitude, m.longitude,
                       m.cover_photo_url, m.visibility, sl.expires_at, sl.revoked_at
                FROM share_links sl
                JOIN memories m ON m.id = sl.memory_id
                WHERE sl.token = %s
                LIMIT 1
                """,
                (token,),
            )
            row = cur.fetchone()
            if not row or row[8] != "shareable" or row[10] is not None:
                raise HTTPException(status_code=404, detail="not found")
            if row[9] is not None:
                now = datetime.now(row[9].tzinfo) if row[9].tzinfo else datetime.now()
                if row[9] <= now:
                    raise HTTPException(status_code=404, detail="not found")
            cur.execute(
                """
                SELECT id, display_url, thumbnail_url, original_filename, taken_at
                FROM photos
                WHERE memory_id = %s
                ORDER BY COALESCE(taken_at, created_at) ASC, id ASC
                """,
                (row[0],),
            )
            photos = cur.fetchall()
            cur.execute(
                """
                SELECT id, couple_space_id, memory_id, provider, immich_asset_id, immich_album_id,
                       original_filename, taken_at, latitude, longitude, thumbnail_cache_key,
                       sort_order, selected_by, created_at, updated_at
                FROM couple_memory_assets
                WHERE memory_id = %s
                ORDER BY sort_order ASC, id ASC
                """,
                (row[0],),
            )
            immich_assets = cur.fetchall()
            cur.execute("SELECT tag FROM memory_tags WHERE memory_id = %s ORDER BY id", (row[0],))
            tags = [r[0] for r in cur.fetchall()]
    return {
        "memory": {
            "id": row[0],
            "title": row[1],
            "content": row[2],
            "memory_date": str(row[3]) if row[3] else None,
            "place_name": row[4],
            "latitude": row[5],
            "longitude": row[6],
            "cover_photo_url": row[7],
            "tags": tags,
        },
        "photos": [
            {
                "id": p[0],
                "display_url": p[1],
                "thumbnail_url": p[2],
                "original_filename": p[3],
                "taken_at": str(p[4]) if p[4] else None,
            }
            for p in photos
        ],
        "immich_assets": [_immich_asset_payload(r) for r in immich_assets],
    }

@app.get("/v1/admin/users")
def admin_users(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, username, display_name, status, is_admin, created_at, last_login_at
                FROM users
                ORDER BY id
                """
            )
            users = cur.fetchall()

            cur.execute(
                """
                SELECT ur.user_id, r.id, r.code, r.name
                FROM user_roles ur
                JOIN roles r ON r.id = ur.role_id
                ORDER BY r.id
                """
            )
            role_rows = cur.fetchall()

    roles_by_user: dict[int, list[dict[str, Any]]] = {}
    for user_id, role_id, code, name in role_rows:
        roles_by_user.setdefault(user_id, []).append({"id": role_id, "code": code, "name": name})

    return {
        "items": [
            {
                "id": row[0],
                "username": row[1],
                "display_name": row[2],
                "status": row[3],
                "is_admin": row[4],
                "created_at": str(row[5]),
                "last_login_at": str(row[6]) if row[6] else None,
                "roles": roles_by_user.get(row[0], []),
            }
            for row in users
        ]
    }


@app.get("/v1/admin/contacts")
def admin_contacts(channel: str | None = Query(default=None), _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    where = ""
    params: tuple[Any, ...] = ()
    if channel:
        where = "WHERE channel = %s"
        params = (channel,)
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, channel, peer_id, display_name, remark, enabled, project_url,
                       project_name, tags, daily_quota, notes, source_sheet, updated_at
                FROM managed_contacts
                {where}
                ORDER BY channel, peer_id
                """,
                params,
            )
            rows = cur.fetchall()
    return {
        "items": [
            {
                "id": row[0],
                "channel": row[1],
                "peer_id": row[2],
                "display_name": row[3],
                "remark": row[4],
                "enabled": row[5],
                "project_url": row[6],
                "project_name": row[7],
                "tags": row[8],
                "daily_quota": row[9],
                "notes": row[10],
                "source_sheet": row[11],
                "updated_at": str(row[12]),
            }
            for row in rows
        ]
    }


@app.post("/v1/admin/contacts")
def admin_upsert_contact(body: ManagedContactUpsertRequest, actor: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    if body.channel not in {"wechat", "feishu"}:
        raise HTTPException(status_code=400, detail="invalid channel")
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO managed_contacts(
                    channel, peer_id, display_name, remark, enabled, project_url,
                    project_name, tags, daily_quota, notes, source_sheet, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT(channel, peer_id)
                DO UPDATE SET
                    display_name = EXCLUDED.display_name,
                    remark = EXCLUDED.remark,
                    enabled = EXCLUDED.enabled,
                    project_url = EXCLUDED.project_url,
                    project_name = EXCLUDED.project_name,
                    tags = EXCLUDED.tags,
                    daily_quota = EXCLUDED.daily_quota,
                    notes = EXCLUDED.notes,
                    source_sheet = EXCLUDED.source_sheet,
                    updated_at = NOW()
                RETURNING id
                """,
                (
                    body.channel,
                    body.peer_id,
                    body.display_name,
                    body.remark,
                    body.enabled,
                    body.project_url,
                    body.project_name,
                    body.tags,
                    body.daily_quota,
                    body.notes,
                    body.source_sheet,
                ),
            )
            contact_id = cur.fetchone()[0]
        conn.commit()
    _audit(actor.get("sub"), "admin.contacts.upsert", "managed_contacts", str(contact_id), body.model_dump())
    return {"id": contact_id}


@app.post("/v1/admin/users")
def admin_create_user(body: CreateUserRequest, actor: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO users(username, display_name, password_hash, is_admin, status)
                VALUES (%s, %s, %s, %s, 'active')
                RETURNING id
                """,
                (body.username, body.display_name, pwd_ctx.hash(body.password), body.is_admin),
            )
            user_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.users.create", "users", str(user_id))
    return {"id": user_id}


@app.patch("/v1/admin/users/{user_id}")
def admin_patch_user(
    user_id: int,
    body: PatchUserRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    fields = body.model_dump(exclude_none=True)
    if not fields:
        return {"status": "ok"}

    allowed = {"display_name", "status", "is_admin"}
    fields = {k: v for k, v in fields.items() if k in allowed}
    sets = ", ".join(f"{key} = %s" for key in fields)
    values = list(fields.values()) + [user_id]

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE users SET {sets}, updated_at = NOW() WHERE id = %s", values)
        conn.commit()

    _audit(actor.get("sub"), "admin.users.patch", "users", str(user_id), fields)
    return {"status": "ok"}


@app.post("/v1/admin/users/{user_id}/reset-password")
def admin_reset_password(
    user_id: int,
    body: ResetPasswordRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET password_hash = %s, updated_at = NOW() WHERE id = %s",
                (pwd_ctx.hash(body.new_password), user_id),
            )
        conn.commit()

    _audit(actor.get("sub"), "admin.users.reset_password", "users", str(user_id))
    return {"status": "ok"}


@app.post("/v1/admin/users/{user_id}/revoke-sessions")
def admin_revoke_user_sessions(user_id: int, user: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET token_version = token_version + 1, updated_at = NOW() WHERE id = %s RETURNING token_version",
                (user_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="user not found")
        conn.commit()

    _audit(user.get("sub"), "admin.users.revoke_sessions", "users", str(user_id))
    return {"user_id": user_id, "token_version": int(row[0])}


@app.post("/v1/admin/users/{user_id}/disable")
def admin_disable_user(user_id: int, actor: dict[str, Any] = Depends(require_admin)) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE users SET status = 'disabled', updated_at = NOW() WHERE id = %s", (user_id,))
        conn.commit()

    _audit(actor.get("sub"), "admin.users.disable", "users", str(user_id))
    return {"status": "ok"}


@app.post("/v1/admin/users/{user_id}/enable")
def admin_enable_user(user_id: int, actor: dict[str, Any] = Depends(require_admin)) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE users SET status = 'active', updated_at = NOW() WHERE id = %s", (user_id,))
        conn.commit()

    _audit(actor.get("sub"), "admin.users.enable", "users", str(user_id))
    return {"status": "ok"}


@app.get("/v1/admin/roles")
def admin_roles(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name, description FROM roles ORDER BY id")
            rows = cur.fetchall()

    return {
        "items": [
            {"id": row[0], "code": row[1], "name": row[2], "description": row[3]}
            for row in rows
        ]
    }


@app.post("/v1/admin/roles")
def admin_create_role(body: CreateRoleRequest, actor: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO roles(code, name, description) VALUES (%s, %s, %s) RETURNING id",
                (body.code, body.name, body.description),
            )
            role_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.roles.create", "roles", str(role_id))
    return {"id": role_id}


@app.patch("/v1/admin/roles/{role_id}")
def admin_patch_role(
    role_id: int,
    body: PatchRoleRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    fields = body.model_dump(exclude_none=True)
    if not fields:
        return {"status": "ok"}

    sets = ", ".join(f"{key} = %s" for key in fields)
    values = list(fields.values()) + [role_id]

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE roles SET {sets} WHERE id = %s", values)
        conn.commit()

    _audit(actor.get("sub"), "admin.roles.patch", "roles", str(role_id), fields)
    return {"status": "ok"}


@app.get("/v1/admin/permissions")
def admin_permissions(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name, description FROM permissions ORDER BY id")
            rows = cur.fetchall()

    return {
        "items": [
            {"id": row[0], "code": row[1], "name": row[2], "description": row[3]}
            for row in rows
        ]
    }


@app.put("/v1/admin/users/{user_id}/roles")
def admin_set_user_roles(
    user_id: int,
    body: PutRoleIdsRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM user_roles WHERE user_id = %s", (user_id,))
            for role_id in body.role_ids:
                cur.execute(
                    """
                    INSERT INTO user_roles(user_id, role_id)
                    VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    (user_id, role_id),
                )
        conn.commit()

    _audit(actor.get("sub"), "admin.users.set_roles", "users", str(user_id), {"role_ids": body.role_ids})
    return {"status": "ok"}


@app.put("/v1/admin/roles/{role_id}/permissions")
def admin_set_role_permissions(
    role_id: int,
    body: PutPermissionIdsRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM role_permissions WHERE role_id = %s", (role_id,))
            for permission_id in body.permission_ids:
                cur.execute(
                    """
                    INSERT INTO role_permissions(role_id, permission_id)
                    VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    (role_id, permission_id),
                )
        conn.commit()

    _audit(
        actor.get("sub"),
        "admin.roles.set_permissions",
        "roles",
        str(role_id),
        {"permission_ids": body.permission_ids},
    )
    return {"status": "ok"}


@app.get("/v1/admin/features")
def admin_features(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, code, title, description, url, category, required_permission, status, sort_order
                FROM features
                ORDER BY sort_order, id
                """
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "code": row[1],
                "title": row[2],
                "description": row[3],
                "url": row[4],
                "category": row[5],
                "required_permission": row[6],
                "status": row[7],
                "sort_order": row[8],
            }
            for row in rows
        ]
    }


@app.post("/v1/admin/features")
def admin_create_feature(
    body: CreateFeatureRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO features(
                    code, title, description, url, category, required_permission, status, sort_order, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                RETURNING id
                """,
                (
                    body.code,
                    body.title,
                    body.description,
                    body.url,
                    body.category,
                    body.required_permission,
                    body.status,
                    body.sort_order,
                ),
            )
            feature_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.features.create", "features", str(feature_id))
    return {"id": feature_id}


@app.patch("/v1/admin/features/{feature_id}")
def admin_patch_feature(
    feature_id: int,
    body: PatchFeatureRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    fields = body.model_dump(exclude_none=True)
    if not fields:
        return {"status": "ok"}

    allowed = {"title", "description", "url", "category", "required_permission", "status", "sort_order"}
    fields = {k: v for k, v in fields.items() if k in allowed}

    sets = ", ".join(f"{key} = %s" for key in fields)
    values = list(fields.values()) + [feature_id]

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE features SET {sets}, updated_at = NOW() WHERE id = %s", values)
        conn.commit()

    _audit(actor.get("sub"), "admin.features.patch", "features", str(feature_id), fields)
    return {"status": "ok"}


@app.get("/v1/admin/audit-logs")
def admin_audit_logs(
    page: int = Query(default=1),
    page_size: int = Query(default=50),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    offset = (page - 1) * page_size
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT count(*) FROM audit_logs")
            total = int(cur.fetchone()[0])
            cur.execute(
                """
                SELECT id, actor_username, action, target_type, target_id, detail, created_at
                FROM audit_logs
                ORDER BY id DESC
                LIMIT %s OFFSET %s
                """,
                (page_size, offset),
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "actor_username": row[1],
                "action": row[2],
                "target_type": row[3],
                "target_id": row[4],
                "detail": row[5],
                "created_at": str(row[6]),
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/v1/admin/doc-sync/sources")
def admin_doc_sync_sources(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    id, provider, env_profile, source_name, source_type,
                    external_doc_id, external_sheet_id, source_url, status,
                    document_name, sheet_name,
                    last_sync_at, created_at, updated_at,
                    (
                        SELECT COUNT(*)
                        FROM external_records er
                        WHERE er.source_id = external_sources.id
                    ) AS record_count,
                    (
                        SELECT COALESCE(array_agg(ef.field_title ORDER BY ef.id), ARRAY[]::TEXT[])
                        FROM external_fields ef
                        WHERE ef.source_id = external_sources.id
                    ) AS field_titles
                FROM external_sources
                ORDER BY updated_at DESC, id DESC
                LIMIT 500
                """
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "provider": row[1],
                "env_profile": row[2],
                "source_name": row[3],
                "source_type": row[4],
                "external_doc_id": row[5],
                "external_sheet_id": row[6],
                "source_url": row[7],
                "status": row[8],
                "document_name": row[9] or row[3],
                "sheet_name": row[10] or "",
                "last_sync_at": str(row[11]) if row[11] else None,
                "created_at": str(row[12]),
                "updated_at": str(row[13]),
                "record_count": row[14],
                "field_titles": row[15] or [],
                "open_url": row[7] or (
                    f"https://doc.weixin.qq.com/smartsheet/{row[5]}?sheet_id={row[6]}"
                    if row[5] and row[6]
                    else (f"https://doc.weixin.qq.com/smartsheet/{row[5]}" if row[5] else "")
                ),
            }
            for row in rows
        ]
    }


@app.get("/v1/admin/doc-sync/runs")
def admin_doc_sync_runs(
    limit: int = Query(default=100, ge=1, le=500),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    id, provider, env_profile, mode, status, started_at, finished_at,
                    source_count, sheet_count, record_count, created_count, updated_count,
                    error_count, error_json
                FROM sync_runs
                ORDER BY started_at DESC, id DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "provider": row[1],
                "env_profile": row[2],
                "mode": row[3],
                "status": row[4],
                "started_at": str(row[5]),
                "finished_at": str(row[6]) if row[6] else None,
                "source_count": row[7],
                "sheet_count": row[8],
                "record_count": row[9],
                "created_count": row[10],
                "updated_count": row[11],
                "error_count": row[12],
                "error_json": row[13],
            }
            for row in rows
        ]
    }


def _doc_sync_records(source_id: int | None, limit: int, offset: int) -> dict[str, Any]:
    where = "WHERE er.source_id = %s" if source_id is not None else ""
    params: list[Any] = []
    if source_id is not None:
        params.append(source_id)
    params.extend([limit, offset])

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    er.id, er.source_id, es.source_name, es.env_profile,
                    er.external_record_id, er.record_hash, er.normalized_json,
                    er.external_created_at, er.external_updated_at, er.synced_at
                FROM external_records er
                JOIN external_sources es ON es.id = er.source_id
                {where}
                ORDER BY er.synced_at DESC, er.id DESC
                LIMIT %s OFFSET %s
                """,
                params,
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "source_id": row[1],
                "source_name": row[2],
                "env_profile": row[3],
                "external_record_id": row[4],
                "record_hash": row[5],
                "normalized_json": row[6],
                "external_created_at": row[7],
                "external_updated_at": row[8],
                "synced_at": str(row[9]),
            }
            for row in rows
        ]
    }


@app.get("/v1/admin/doc-sync/records")
def admin_doc_sync_records(
    source_id: int | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    return _doc_sync_records(source_id=source_id, limit=limit, offset=offset)


@app.get("/v1/admin/doc-sync/sources/{source_id}/records")
def admin_doc_sync_source_records(
    source_id: int,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    return _doc_sync_records(source_id=source_id, limit=limit, offset=offset)


@app.get("/v1/admin/doc-sync/requests")
def admin_doc_sync_requests(
    limit: int = Query(default=100, ge=1, le=500),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    sr.id, sr.source_id, es.source_name, sr.provider, sr.env_profile,
                    sr.mode, sr.status, sr.requested_by, sr.requested_at,
                    sr.started_at, sr.finished_at, sr.sync_run_id, sr.error_json
                FROM sync_requests sr
                JOIN external_sources es ON es.id = sr.source_id
                ORDER BY sr.requested_at DESC, sr.id DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "source_id": row[1],
                "source_name": row[2],
                "provider": row[3],
                "env_profile": row[4],
                "mode": row[5],
                "status": row[6],
                "requested_by": row[7],
                "requested_at": str(row[8]),
                "started_at": str(row[9]) if row[9] else None,
                "finished_at": str(row[10]) if row[10] else None,
                "sync_run_id": row[11],
                "error_json": row[12],
            }
            for row in rows
        ]
    }


@app.post("/v1/admin/doc-sync/sources/{source_id}/sync-requests")
def admin_create_doc_sync_request(
    source_id: int,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT provider, env_profile
                FROM external_sources
                WHERE id = %s AND status = 'active'
                """,
                (source_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="doc sync source not found")
            cur.execute(
                """
                INSERT INTO sync_requests(source_id, provider, env_profile, mode, status, requested_by)
                VALUES (%s, %s, %s, 'manual', 'pending', %s)
                RETURNING id
                """,
                (source_id, row[0], row[1], actor.get("sub")),
            )
            request_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.doc_sync.request", "external_sources", str(source_id), {"request_id": request_id})
    return {"id": request_id, "status": "pending"}


@app.get("/v1/admin/couple-spaces")
def admin_couple_spaces(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name, start_date, theme, created_at FROM couple_spaces ORDER BY id DESC")
            spaces = cur.fetchall()
            cur.execute(
                """
                SELECT cm.couple_space_id, u.id, u.username, u.display_name, cm.role, cm.joined_at
                FROM couple_members cm
                JOIN users u ON u.id = cm.user_id
                ORDER BY cm.couple_space_id DESC, cm.id ASC
                """
            )
            member_rows = cur.fetchall()

    members_map: dict[int, list[dict[str, Any]]] = {}
    for row in member_rows:
        members_map.setdefault(row[0], []).append(
            {
                "user_id": row[1],
                "username": row[2],
                "display_name": row[3],
                "role": row[4],
                "joined_at": str(row[5]),
            }
        )

    return {
        "items": [
            {
                "id": row[0],
                "name": row[1],
                "start_date": str(row[2]) if row[2] else None,
                "theme": row[3],
                "created_at": str(row[4]),
                "members": members_map.get(row[0], []),
            }
            for row in spaces
        ]
    }


@app.post("/v1/admin/couple-spaces")
def admin_create_couple_space(
    body: CreateCoupleSpaceRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO couple_spaces(name, start_date, theme) VALUES (%s, %s, %s) RETURNING id",
                (name, body.start_date, body.theme),
            )
            new_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.couple_spaces.create", "couple_spaces", str(new_id), {"name": name})
    return {"id": new_id}


@app.post("/v1/admin/couple-spaces/{space_id}/members")
def admin_add_couple_member(
    space_id: int,
    body: AddCoupleMemberRequest,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    username = body.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="username is required")
    role = body.role.strip() or "member"
    if role not in {"member", "owner"}:
        raise HTTPException(status_code=400, detail="invalid role")

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM couple_spaces WHERE id = %s", (space_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="couple space not found")
            cur.execute("SELECT id FROM users WHERE username = %s", (username,))
            user_row = cur.fetchone()
            if not user_row:
                raise HTTPException(status_code=404, detail="user not found")
            cur.execute(
                """
                INSERT INTO couple_members(couple_space_id, user_id, role)
                VALUES (%s, %s, %s)
                ON CONFLICT(couple_space_id, user_id)
                DO UPDATE SET role = EXCLUDED.role
                """,
                (space_id, user_row[0], role),
            )
        conn.commit()

    _audit(actor.get("sub"), "admin.couple_spaces.add_member", "couple_spaces", str(space_id), {"username": username, "role": role})
    return {"status": "ok"}


@app.delete("/v1/admin/couple-spaces/{space_id}/members/{user_id}")
def admin_remove_couple_member(
    space_id: int,
    user_id: int,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM couple_members WHERE couple_space_id = %s AND user_id = %s",
                (space_id, user_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="member not found")
        conn.commit()

    _audit(actor.get("sub"), "admin.couple_spaces.remove_member", "couple_spaces", str(space_id), {"user_id": user_id})
    return {"status": "ok"}


@app.post("/v1/admin/users/{user_id}/grant-couple-access")
def admin_grant_couple_access(
    user_id: int,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, str]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE id = %s", (user_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="user not found")
            cur.execute(
                """
                INSERT INTO permissions(code, name, description)
                VALUES ('couple_memory_access', 'Couple Memory 访问', '访问双人私密回忆空间')
                ON CONFLICT(code) DO NOTHING
                """
            )
            cur.execute(
                """
                INSERT INTO roles(code, name, description)
                VALUES ('couple_memory', 'Couple Memory 成员', '可访问 Couple Memory 私密空间')
                ON CONFLICT(code) DO NOTHING
                """
            )
            cur.execute(
                """
                INSERT INTO role_permissions(role_id, permission_id)
                SELECT r.id, p.id
                FROM roles r, permissions p
                WHERE r.code = 'couple_memory' AND p.code = 'couple_memory_access'
                ON CONFLICT DO NOTHING
                """
            )
            cur.execute(
                """
                INSERT INTO user_roles(user_id, role_id)
                SELECT %s, id FROM roles WHERE code = 'couple_memory'
                ON CONFLICT DO NOTHING
                """,
                (user_id,),
            )
        conn.commit()

    _audit(actor.get("sub"), "admin.users.grant_couple_access", "users", str(user_id))
    return {"status": "ok"}
