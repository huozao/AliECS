"""数据导出与同步域：T+导出下载、库存查询、微信/飞书路由配置下发、企微文档同步管理与外部数据源导出。"""

from __future__ import annotations

import os

from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse

from app.core import _audit, _conn, require_admin, require_login


router = APIRouter()

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


_TPLUS_EXPORT_DESCRIPTIONS = {
    "bom": "BOM 父件和子件用量；看成品由哪些原材料组成，不含价格。",
    "inventory": "存货基础档案；含分类、采购/销售/材料标记、税率、RetailPriceNew/InvUnitPriceDTOs。销售价格先看这里；采购价未接入。",
    "current_stock": "仓库×存货现存量/可用量；看原材库库存数量，不含价格。",
    "partner": "往来单位档案；客户/供应商及价格等级，不是商品价格。",
    "warehouse": "仓库档案；用于识别原材库、成品库等仓库编码。",
    "unit_group": "计量单位组档案；辅助理解单位换算关系。",
    "unit": "计量单位档案；含主单位、换算率等单位信息。",
    "project": "项目档案；当前可能为空，主要用于项目辅助核算。",
    "project_class": "项目分类档案；用于识别项目类别。",
    "brand": "品牌档案；当前可能为空。",
    "district": "地区档案；当前可能为空。",
    "sale_order_list": "销售订单列表，仅单据ID/日期/单号；不含明细售价金额。",
    "sale_delivery_list": "销货单列表，仅单据ID/日期/单号；不含明细售价金额。",
    "purchase_order_list": "采购订单列表，仅单据ID/日期/单号；不含明细单价金额，不能核对原材料采购价。",
    "purchase_arrival_list": "采购到货单列表，仅单据ID/日期/单号；不含明细单价金额。",
    "purchase_receive_list": "采购入库单列表，仅单据ID/日期/单号；不含明细单价金额。",
    "material_dispatch_list": "材料出库/领料单列表，仅单据ID/日期/单号；不含价格。",
    "purchase_price": "采购价格表（采购到货明细，来自 T+ 报表）；原材料采购单价/含税单价优先看这里，成本核算系统单价取此最新价。",
    "sales_price": "销售价格表（销货单明细，来自 T+ 报表）；商品销售单价/含税单价优先看这里。",
}


def _system_config_record(sheet_name: str) -> dict[str, Any]:
    """Read one singleton row from the doc-sync mirror of Feishu「系统配置」.

    This is intentionally best-effort: backend must keep serving with code
    defaults when doc-sync, Postgres, or the mirror row is temporarily absent.
    """
    try:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT er.raw_json, er.normalized_json, er.source_id
                    FROM external_sources es
                    JOIN external_records er ON er.source_id = es.id
                    WHERE es.provider = 'feishu'
                      AND es.document_name = '系统配置'
                      AND es.sheet_name = %s
                      AND es.status = 'active'
                    ORDER BY er.id
                    """,
                    (sheet_name,),
                )
                rows = cur.fetchall()
                source_ids = sorted({int(row[2]) for row in rows if len(row) >= 3 and row[2] is not None})
                field_titles: dict[tuple[int, str], str] = {}
                if source_ids:
                    cur.execute(
                        """
                        SELECT source_id, external_field_id, field_title
                        FROM external_fields
                        WHERE source_id = ANY(%s)
                        """,
                        (source_ids,),
                    )
                    field_titles = {
                        (int(source_id), str(field_id)): str(title or field_id)
                        for source_id, field_id, title in cur.fetchall()
                    }
    except Exception:  # noqa: BLE001 - 配置镜像不可用时回退代码默认
        return {}
    for raw_json, normalized_json, source_id in rows:
        source_id = int(source_id)
        fields: dict[str, Any] = {}
        if isinstance(normalized_json, dict):
            fields.update(normalized_json)
        raw = raw_json if isinstance(raw_json, dict) else {}
        raw_fields = raw.get("fields") if isinstance(raw.get("fields"), dict) else raw
        for key, value in raw_fields.items():
            fields[field_titles.get((source_id, str(key)), str(key))] = value
        if _config_text(fields.get("配置编号")) == "global-default":
            return fields
    return {}


def _config_cell_text(value: Any) -> str:
    if isinstance(value, dict):
        for key in ("text", "name", "value"):
            if key in value:
                return str(value.get(key) or "").strip()
        return ""
    return str(value or "").strip()


def _config_text(value: Any) -> str:
    if isinstance(value, list):
        return "".join(_config_cell_text(item) for item in value).strip()
    return _config_cell_text(value)


def _config_codes(value: Any) -> list[str]:
    values = (
        value
        if isinstance(value, list)
        else str(value or "").replace("；", ";").replace("，", ";").replace(",", ";").split(";")
    )
    result: list[str] = []
    for item in values:
        code = _config_cell_text(item)
        if code and code not in result:
            result.append(code)
    return result


def _tplus_export_description(module: str) -> str:
    configured = _config_text(_system_config_record("T+导出说明").get(module))
    if configured:
        return configured
    return _TPLUS_EXPORT_DESCRIPTIONS.get(module, "暂未配置说明；请按表头人工判断内容。")


def _tplus_export_dir() -> Path:
    return Path(os.getenv("TPLUS_EXPORT_DIR", "/app/tplus-output/excel"))


def _tplus_module_of(file_name: str) -> str:
    stem = file_name[:-5] if file_name.endswith(".xlsx") else file_name
    parts = stem.rsplit("_", 2)
    if len(parts) == 3 and parts[1].isdigit() and parts[2].isdigit():
        return parts[0]
    return stem


def _parse_export_timestamp(file_name: str) -> datetime | None:
    stem = file_name[:-5] if file_name.endswith(".xlsx") else file_name
    parts = stem.rsplit("_", 2)
    if len(parts) == 3 and len(parts[1]) == 8 and len(parts[2]) == 6 and parts[1].isdigit() and parts[2].isdigit():
        try:
            return datetime.strptime(parts[1] + parts[2], "%Y%m%d%H%M%S")
        except ValueError:
            return None
    return None


def _match_export_files_to_runs(runs: list[tuple[Any, Any]], files: list[str]) -> dict[Any, list[str]]:
    """把每个文件归给 finished_at >= 文件时间戳 的最早一次 run。
    runs: [(run_id, finished_at_iso_or_dt)]，可乱序。返回 {run_id: [file,...]}。"""
    parsed_runs = []
    for run_id, finished in runs:
        if finished is None:
            continue
        dt = finished if isinstance(finished, datetime) else datetime.fromisoformat(str(finished).replace("Z", "")[:19])
        # psycopg 返回 timestamptz 为 tz-aware；文件名时间戳解析为 naive。统一为 naive-UTC 再比较，避免 TypeError。
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        parsed_runs.append((dt, run_id))
    parsed_runs.sort()
    mapping: dict[Any, list[str]] = {}
    for name in files:
        t = _parse_export_timestamp(name)
        if t is None:
            continue
        chosen = next((rid for dt, rid in parsed_runs if dt >= t), None)
        if chosen is not None:
            mapping.setdefault(chosen, []).append(name)
    return mapping


def _latest_tplus_exports() -> list[dict[str, Any]]:
    directory = _tplus_export_dir()
    latest: dict[str, Path] = {}
    if directory.is_dir():
        for item in directory.glob("*.xlsx"):
            module = _tplus_module_of(item.name)
            current = latest.get(module)
            if current is None or item.name > current.name:
                latest[module] = item
    items: list[dict[str, Any]] = []
    for module in sorted(latest):
        path = latest[module]
        stat = path.stat()
        items.append(
            {
                "name": module,
                "file_name": path.name,
                "description": _tplus_export_description(module),
                "size_bytes": stat.st_size,
                "updated_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "download_url": f"/v1/exports/tplus/{path.name}",
            }
        )
    return items


def _external_source_tab_key(provider: str, env_profile: str) -> str:
    if provider == "feishu":
        return "feishu"
    return f"{provider}_{env_profile.lower()}"


@router.get("/v1/exports/catalog")
def exports_catalog(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    tabs: dict[str, dict[str, Any]] = {
        "tplus": {"key": "tplus", "title": "T+ ERP", "items": _latest_tplus_exports()},
        "wecom_company_a": {"key": "wecom_company_a", "title": "企微A", "items": []},
        "wecom_company_b": {"key": "wecom_company_b", "title": "企微B", "items": []},
        "feishu": {"key": "feishu", "title": "飞书", "items": []},
    }
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            # 按工作簿（文档）聚合：每个智能表格文档一条，下载时整簿多 sheet 导出。
            cur.execute(
                """
                SELECT s.provider, s.env_profile, s.external_doc_id,
                       COALESCE(
                           MAX(s.document_name) FILTER (WHERE s.external_sheet_id = ''),
                           MAX(NULLIF(s.document_name, '')),
                           MAX(NULLIF(s.source_name, ''))
                       ) AS document_name,
                       COALESCE(MIN(s.id) FILTER (WHERE s.external_sheet_id = ''), MIN(s.id)) AS first_source_id,
                       COUNT(DISTINCT s.id) FILTER (WHERE s.external_sheet_id <> '') AS sheet_count,
                       COUNT(r.id) FILTER (WHERE s.external_sheet_id <> '') AS row_count,
                       MAX(s.last_sync_at) FILTER (WHERE s.external_sheet_id <> '') AS last_sync_at
                FROM external_sources s
                LEFT JOIN external_records r ON r.source_id = s.id AND s.external_sheet_id <> ''
                WHERE s.status = 'active' AND s.external_doc_id <> ''
                GROUP BY s.provider, s.env_profile, s.external_doc_id
                ORDER BY MIN(s.id)
                """
            )
            rows = cur.fetchall()
    for provider, env_profile, _doc_id, document_name, first_source_id, sheet_count, row_count, last_sync_at in rows:
        key = _external_source_tab_key(str(provider or ""), str(env_profile or ""))
        tab = tabs.setdefault(key, {"key": key, "title": key, "items": []})
        tab["items"].append(
            {
                "name": document_name or f"{provider} 文档",
                "source_id": first_source_id,
                "sheets": int(sheet_count or 0),
                "rows": int(row_count or 0),
                "updated_at": str(last_sync_at) if last_sync_at else None,
                "download_url": f"/v1/exports/external-doc/{first_source_id}" if int(sheet_count or 0) > 0 else None,
            }
        )
    return {"tabs": list(tabs.values())}


_STOCK_COLUMNS = {
    "WarehouseCode": "仓库编码",
    "WarehouseName": "仓库",
    "InventoryCode": "存货编码",
    "InventoryName": "存货名称",
    "InventoryClassName": "存货分类",
    "Specification": "规格型号",
    "UnitName": "单位",
    "ExistingQuantity": "现存量",
    "AvailableQuantity": "可用量",
}


def _latest_tplus_export_file(module: str) -> Path | None:
    directory = _tplus_export_dir()
    if not directory.is_dir():
        return None
    candidates = [item for item in directory.glob(f"{module}_*.xlsx") if _tplus_module_of(item.name) == module]
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.name)


# 库存页仓库范围：原材料=原材库(001)+L-代加工库(012)；成品=除原材库外全部。
_RAW_STOCK_WAREHOUSES = {"001", "012"}
_FINISHED_EXCLUDED_WAREHOUSES = {"001"}


def _latest_tplus_warehouse_codes() -> set[str]:
    path = _latest_tplus_export_file("warehouse")
    if path is None:
        return set()
    try:
        import pandas as pd

        df = pd.read_excel(path, dtype=str).fillna("")
    except Exception:  # noqa: BLE001 - 仓库档案不可读时不阻断库存页
        return set()
    for column in ("WarehouseCode", "仓库编码", "Code"):
        if column in df.columns:
            return {str(code).strip() for code in df[column].tolist() if str(code).strip()}
    return set()


def _validated_config_codes(configured: list[str], defaults: set[str], valid_codes: set[str]) -> set[str]:
    if not configured:
        return set(defaults)
    if not valid_codes:
        return set(defaults)
    selected = {code for code in configured if code in valid_codes}
    return selected or set(defaults)


def _inventory_scope_config() -> tuple[set[str], set[str]]:
    record = _system_config_record("库存仓库范围")
    valid_codes = _latest_tplus_warehouse_codes()
    raw_warehouses = _validated_config_codes(
        _config_codes(record.get("库存原料仓库")),
        _RAW_STOCK_WAREHOUSES,
        valid_codes,
    )
    finished_excluded = _validated_config_codes(
        _config_codes(record.get("成品排除仓库")),
        _FINISHED_EXCLUDED_WAREHOUSES,
        valid_codes,
    )
    return raw_warehouses, finished_excluded


@router.get("/v1/inventory/current-stock")
def inventory_current_stock(
    q: str = Query(default=""),
    warehouse: str = Query(default=""),
    scope: str = Query(default="raw"),
    user: dict[str, Any] = Depends(require_login),
) -> dict[str, Any]:
    if scope not in ("raw", "finished"):
        raise HTTPException(status_code=400, detail="scope must be raw or finished")
    roles = user.get("roles", [])
    permissions = user.get("permissions", [])
    scope_permission = "inventory.raw.read" if scope == "raw" else "inventory.finished.read"
    allowed = "admin" in roles or "admin.access" in permissions or scope_permission in permissions
    if not allowed:
        raise HTTPException(status_code=403, detail="permission denied")

    path = _latest_tplus_export_file("current_stock")
    if path is None:
        raise HTTPException(status_code=404, detail="现存量数据尚未同步，请先在 T+ 同步任务跑一轮全量。")

    import pandas as pd

    df = pd.read_excel(path, dtype=str)
    for column in _STOCK_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = df[list(_STOCK_COLUMNS)].fillna("")

    raw_warehouses, finished_excluded = _inventory_scope_config()
    codes = df["WarehouseCode"].str.strip()
    if scope == "raw":
        df = df[codes.isin(raw_warehouses)]
    else:
        df = df[~codes.isin(finished_excluded)]

    warehouses = (
        df[["WarehouseCode", "WarehouseName"]]
        .drop_duplicates()
        .sort_values("WarehouseCode")
        .to_dict("records")
    )

    requested_warehouse = warehouse.strip()
    if requested_warehouse:
        if requested_warehouse not in {str(item["WarehouseCode"]).strip() for item in warehouses}:
            raise HTTPException(status_code=400, detail="warehouse not in scope")
        df = df[df["WarehouseCode"].str.strip() == requested_warehouse]
    keyword = q.strip()
    if keyword:
        lowered = keyword.lower()
        mask = (
            df["InventoryName"].str.lower().str.contains(lowered, na=False)
            | df["InventoryCode"].str.lower().str.contains(lowered, na=False)
            | df["Specification"].str.lower().str.contains(lowered, na=False)
        )
        df = df[mask]

    for column in ("ExistingQuantity", "AvailableQuantity"):
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)

    stat = path.stat()
    return {
        "items": df.to_dict("records"),
        "total": int(len(df)),
        "warehouses": warehouses,
        "source_file": path.name,
        "synced_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
    }


@router.get("/v1/exports/tplus/{file_name}")
def exports_tplus_download(file_name: str, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    if Path(file_name).name != file_name or not file_name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="invalid file name")
    path = _tplus_export_dir() / file_name
    if not path.is_file():
        raise HTTPException(status_code=404, detail="export file not found")
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=file_name)


def _routing_projects(channel: str) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT peer_id, display_name, project_url, project_name
                FROM managed_contacts
                WHERE channel = %s
                  AND enabled = true
                  AND COALESCE(project_url, '') <> ''
                ORDER BY peer_id
                """,
                (channel,),
            )
            rows = cur.fetchall()
    lanes: dict[str, dict[str, str]] = {}
    for peer_id, display_name, project_url, project_name in rows:
        if not peer_id or not project_url:
            continue
        lanes[str(peer_id)] = {
            "name": str(display_name or project_name or peer_id),
            "project_url": str(project_url),
        }
    return {"lanes": lanes}


@router.get("/v1/routing/wechat-projects.json")
def routing_wechat_projects() -> dict[str, Any]:
    return _routing_projects("wechat")


@router.get("/v1/routing/feishu-projects.json")
def routing_feishu_projects() -> dict[str, Any]:
    return _routing_projects("feishu")


@router.get("/v1/exports/external/{source_id}")
def exports_external_download(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT provider, env_profile, document_name, sheet_name FROM external_sources WHERE id = %s",
                (source_id,),
            )
            source = cur.fetchone()
            if not source:
                raise HTTPException(status_code=404, detail="external source not found")
            cur.execute(
                """
                SELECT external_record_id, normalized_json, external_updated_at, synced_at
                FROM external_records
                WHERE source_id = %s
                ORDER BY id
                """,
                (source_id,),
            )
            records = cur.fetchall()

    provider, env_profile, document_name, sheet_name = source
    columns: list[str] = []
    for _record_id, normalized, _ext_updated, _synced in records:
        if isinstance(normalized, dict):
            for column in normalized:
                if column not in columns:
                    columns.append(column)

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(sheet_name or "data")[:31] or "data"
    header = ["external_record_id", *columns, "external_updated_at", "synced_at"]
    sheet.append(header)
    for record_id, normalized, ext_updated, synced in records:
        data = normalized if isinstance(normalized, dict) else {}
        sheet.append(
            [record_id, *[data.get(column, "") for column in columns], ext_updated or "", str(synced or "")]
        )

    export_dir = Path(os.getenv("EXTERNAL_EXPORT_DIR", "/tmp/aliecs-external-exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"{provider}_{(env_profile or 'default').lower()}_{source_id}_{timestamp}.xlsx"
    workbook.save(path)
    label = f"{document_name or provider}-{sheet_name or source_id}"
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=f"{label}_{timestamp}.xlsx")


def _append_records_worksheet(workbook: Any, title: str, records: list[tuple]) -> None:
    columns: list[str] = []
    for _record_id, normalized, _ext_updated, _synced in records:
        if isinstance(normalized, dict):
            for column in normalized:
                if column not in columns:
                    columns.append(column)
    base = (str(title or "data").strip() or "data")[:31]
    name = base
    suffix = 2
    while name in workbook.sheetnames:
        name = f"{base[:28]}_{suffix}"
        suffix += 1
    sheet = workbook.create_sheet(title=name)
    sheet.append(["external_record_id", *columns, "external_updated_at", "synced_at"])
    for record_id, normalized, ext_updated, synced in records:
        data = normalized if isinstance(normalized, dict) else {}
        sheet.append([record_id, *[data.get(column, "") for column in columns], ext_updated or "", str(synced or "")])


@router.get("/v1/exports/external-doc/{source_id}")
def exports_external_doc_download(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> FileResponse:
    """按所属工作簿导出：同文档的全部 sheet 各占一个工作表。source_id 为该文档任一 sheet 级源。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT provider, env_profile, external_doc_id, document_name FROM external_sources WHERE id = %s",
                (source_id,),
            )
            anchor = cur.fetchone()
            if not anchor:
                raise HTTPException(status_code=404, detail="external source not found")
            provider, env_profile, external_doc_id, document_name = anchor
            cur.execute(
                """
                SELECT id, sheet_name
                FROM external_sources
                WHERE provider = %s AND env_profile = %s AND external_doc_id = %s
                  AND external_sheet_id <> '' AND status = 'active'
                ORDER BY id
                """,
                (provider, env_profile, external_doc_id),
            )
            sheet_sources = cur.fetchall()
            if not sheet_sources:
                raise HTTPException(status_code=404, detail="document has no active sheets")

            from openpyxl import Workbook

            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_source_id, sheet_name in sheet_sources:
                cur.execute(
                    """
                    SELECT external_record_id, normalized_json, external_updated_at, synced_at
                    FROM external_records
                    WHERE source_id = %s
                    ORDER BY id
                    """,
                    (sheet_source_id,),
                )
                _append_records_worksheet(workbook, str(sheet_name or sheet_source_id), cur.fetchall())

    export_dir = Path(os.getenv("EXTERNAL_EXPORT_DIR", "/tmp/aliecs-external-exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"{provider}_{(env_profile or 'default').lower()}_doc_{source_id}_{timestamp}.xlsx"
    workbook.save(path)
    label = str(document_name or provider)
    return FileResponse(path, media_type=_XLSX_MEDIA_TYPE, filename=f"{label}_{timestamp}.xlsx")


def _external_doc_anchor(cur: Any, source_id: int) -> tuple[str, str, str, str]:
    cur.execute(
        "SELECT provider, env_profile, external_doc_id, document_name FROM external_sources WHERE id = %s",
        (source_id,),
    )
    anchor = cur.fetchone()
    if not anchor:
        raise HTTPException(status_code=404, detail="external source not found")
    return anchor


def _ensure_doc_row(cur: Any, provider: str, env_profile: str, external_doc_id: str, document_name: str) -> int:
    """确保 doc 级登记行存在并返回其 id（worker 对 doc 级请求整簿重扫，含新 sheet 发现）。"""
    cur.execute(
        """
        INSERT INTO external_sources(
            provider, env_profile, source_name, source_type,
            external_doc_id, external_sheet_id, source_url,
            document_name, sheet_name, status, updated_at
        )
        VALUES (%s, %s, %s, 'smartsheet_doc', %s, '', '', %s, '', 'active', NOW())
        ON CONFLICT(provider, env_profile, external_doc_id, external_sheet_id)
        DO UPDATE SET status = 'active', updated_at = NOW()
        RETURNING id
        """,
        (provider, env_profile, document_name, external_doc_id, document_name),
    )
    return int(cur.fetchone()[0])


def _create_doc_sync_request(cur: Any, doc_row_id: int, provider: str, env_profile: str, requested_by: str) -> None:
    cur.execute(
        """
        INSERT INTO sync_requests(source_id, provider, env_profile, mode, status, requested_by)
        VALUES (%s, %s, %s, 'manual', 'pending', %s)
        """,
        (doc_row_id, provider, env_profile, requested_by),
    )


@router.post("/v1/exports/external-doc/{source_id}/sync-requests")
def exports_external_doc_sync(source_id: int, user: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """为该工作簿创建 doc 级同步请求（worker 整簿重扫，含新 sheet 发现），约 30 秒内开始。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            provider, env_profile, external_doc_id, document_name = _external_doc_anchor(cur, source_id)
            doc_row_id = _ensure_doc_row(cur, str(provider), str(env_profile), str(external_doc_id), str(document_name or ""))
            _create_doc_sync_request(cur, doc_row_id, str(provider), str(env_profile), str(user.get("sub") or ""))
        conn.commit()
    return {
        "document_name": document_name,
        "requests_created": 1,
        "message": f"已为「{document_name}」创建整簿同步请求，约 30 秒内开始同步。",
    }


@router.post("/v1/exports/sync-all")
def exports_sync_all(user: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """同步数据列表：为全部已登记文档各建一条 doc 级同步请求（发现新文档/新表/改名/新记录）。"""
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, provider, env_profile FROM external_sources
                WHERE external_sheet_id = '' AND status = 'active' AND external_doc_id <> ''
                ORDER BY id
                """
            )
            doc_rows = cur.fetchall()
            for doc_row_id, provider, env_profile in doc_rows:
                _create_doc_sync_request(cur, int(doc_row_id), str(provider), str(env_profile), str(user.get("sub") or ""))
        conn.commit()
    return {
        "requests_created": len(doc_rows),
        "message": f"已为 {len(doc_rows)} 个文档创建同步请求，列表将在 1-2 分钟内陆续刷新。",
    }


@router.post("/v1/exports/external-doc/{source_id}/copy")
def exports_external_doc_copy(source_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    """在企业微信中创建该智能表格的完整副本（全部工作表结构 + 全部记录）。"""
    from app.integrations.wecom_docs import WeComDocError, copy_smartsheet_doc

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            provider, env_profile, external_doc_id, document_name = _external_doc_anchor(cur, source_id)
    if provider != "wecom":
        raise HTTPException(status_code=400, detail="仅支持企业微信智能表格创建副本")

    new_name = f"{document_name or '智能表格'}-副本{datetime.now(tz=timezone.utc).strftime('%y%m%d_%H%M')}"
    try:
        result = copy_smartsheet_doc(env_profile=str(env_profile), source_docid=str(external_doc_id), new_doc_name=new_name)
    except WeComDocError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    # 新副本自动登记为同步源并触发首次整簿同步，30 秒内出现在数据导出列表。
    new_docid = str(result.get("new_docid") or "")
    if new_docid:
        with closing(_conn()) as conn:
            with conn.cursor() as cur:
                doc_row_id = _ensure_doc_row(cur, str(provider), str(env_profile), new_docid, new_name)
                _create_doc_sync_request(cur, doc_row_id, str(provider), str(env_profile), "copy-auto")
            conn.commit()
        result["registered"] = True
    return {"document_name": document_name, "new_doc_name": new_name, **result}


@router.get("/v1/admin/doc-sync/sources")
def admin_doc_sync_sources(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    id, provider, env_profile, source_name, source_type,
                    external_doc_id, external_sheet_id, source_url, status,
                    document_name, sheet_name,
                    last_sync_at, created_at, updated_at,
                    (
                        SELECT COUNT(*)
                        FROM external_records er
                        WHERE er.source_id = external_sources.id
                    ) AS record_count,
                    (
                        SELECT COALESCE(array_agg(ef.field_title ORDER BY ef.id), ARRAY[]::TEXT[])
                        FROM external_fields ef
                        WHERE ef.source_id = external_sources.id
                    ) AS field_titles
                FROM external_sources
                ORDER BY updated_at DESC, id DESC
                LIMIT 500
                """
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "provider": row[1],
                "env_profile": row[2],
                "source_name": row[3],
                "source_type": row[4],
                "external_doc_id": row[5],
                "external_sheet_id": row[6],
                "source_url": row[7],
                "status": row[8],
                "document_name": row[9] or row[3],
                "sheet_name": row[10] or "",
                "last_sync_at": str(row[11]) if row[11] else None,
                "created_at": str(row[12]),
                "updated_at": str(row[13]),
                "record_count": row[14],
                "field_titles": row[15] or [],
                "open_url": row[7] or (
                    f"https://doc.weixin.qq.com/smartsheet/{row[5]}?sheet_id={row[6]}"
                    if row[5] and row[6]
                    else (f"https://doc.weixin.qq.com/smartsheet/{row[5]}" if row[5] else "")
                ),
            }
            for row in rows
        ]
    }


@router.get("/v1/admin/doc-sync/runs")
def admin_doc_sync_runs(
    limit: int = Query(default=100, ge=1, le=500),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    id, provider, env_profile, mode, status, started_at, finished_at,
                    source_count, sheet_count, record_count, created_count, updated_count,
                    error_count, error_json
                FROM sync_runs
                ORDER BY started_at DESC, id DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "provider": row[1],
                "env_profile": row[2],
                "mode": row[3],
                "status": row[4],
                "started_at": str(row[5]),
                "finished_at": str(row[6]) if row[6] else None,
                "source_count": row[7],
                "sheet_count": row[8],
                "record_count": row[9],
                "created_count": row[10],
                "updated_count": row[11],
                "error_count": row[12],
                "error_json": row[13],
            }
            for row in rows
        ]
    }


def _doc_sync_records(source_id: int | None, limit: int, offset: int) -> dict[str, Any]:
    where = "WHERE er.source_id = %s" if source_id is not None else ""
    params: list[Any] = []
    if source_id is not None:
        params.append(source_id)
    params.extend([limit, offset])

    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    er.id, er.source_id, es.source_name, es.env_profile,
                    er.external_record_id, er.record_hash, er.normalized_json,
                    er.external_created_at, er.external_updated_at, er.synced_at
                FROM external_records er
                JOIN external_sources es ON es.id = er.source_id
                {where}
                ORDER BY er.synced_at DESC, er.id DESC
                LIMIT %s OFFSET %s
                """,
                params,
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "source_id": row[1],
                "source_name": row[2],
                "env_profile": row[3],
                "external_record_id": row[4],
                "record_hash": row[5],
                "normalized_json": row[6],
                "external_created_at": row[7],
                "external_updated_at": row[8],
                "synced_at": str(row[9]),
            }
            for row in rows
        ]
    }


@router.get("/v1/admin/doc-sync/records")
def admin_doc_sync_records(
    source_id: int | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    return _doc_sync_records(source_id=source_id, limit=limit, offset=offset)


@router.get("/v1/admin/doc-sync/sources/{source_id}/records")
def admin_doc_sync_source_records(
    source_id: int,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    return _doc_sync_records(source_id=source_id, limit=limit, offset=offset)


@router.get("/v1/admin/doc-sync/requests")
def admin_doc_sync_requests(
    limit: int = Query(default=100, ge=1, le=500),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    sr.id, sr.source_id, es.source_name, sr.provider, sr.env_profile,
                    sr.mode, sr.status, sr.requested_by, sr.requested_at,
                    sr.started_at, sr.finished_at, sr.sync_run_id, sr.error_json
                FROM sync_requests sr
                JOIN external_sources es ON es.id = sr.source_id
                ORDER BY sr.requested_at DESC, sr.id DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()

    return {
        "items": [
            {
                "id": row[0],
                "source_id": row[1],
                "source_name": row[2],
                "provider": row[3],
                "env_profile": row[4],
                "mode": row[5],
                "status": row[6],
                "requested_by": row[7],
                "requested_at": str(row[8]),
                "started_at": str(row[9]) if row[9] else None,
                "finished_at": str(row[10]) if row[10] else None,
                "sync_run_id": row[11],
                "error_json": row[12],
            }
            for row in rows
        ]
    }


@router.post("/v1/admin/doc-sync/sources/{source_id}/sync-requests")
def admin_create_doc_sync_request(
    source_id: int,
    actor: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with closing(_conn()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT provider, env_profile
                FROM external_sources
                WHERE id = %s AND status = 'active'
                """,
                (source_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="doc sync source not found")
            cur.execute(
                """
                INSERT INTO sync_requests(source_id, provider, env_profile, mode, status, requested_by)
                VALUES (%s, %s, %s, 'manual', 'pending', %s)
                RETURNING id
                """,
                (source_id, row[0], row[1], actor.get("sub")),
            )
            request_id = cur.fetchone()[0]
        conn.commit()

    _audit(actor.get("sub"), "admin.doc_sync.request", "external_sources", str(source_id), {"request_id": request_id})
    return {"id": request_id, "status": "pending"}
