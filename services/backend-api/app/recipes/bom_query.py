from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_MATERIAL = "物料清单"
SHEET_COMPONENT = "子件明细"
SHEET_DETAIL = "父件子件明细"
DETAIL_SHEET = "父件子件明细_提取"
SUMMARY_SHEET = "版本父件分组合计"
HUMAN_SHEET = "配方表_人眼版"
MATRIX_SHEET = "横向对比_矩阵"

OUTPUT_COLUMNS = [
    "版本号_子件",
    "父件编码",
    "子件编码",
    "父件名称",
    "子件名称",
    "规格型号_子件",
    "计量单位_子件",
    "需用数量",
    "比例",
    "系统单价",
    "默认BOM",
    "停用",
]

PRICE_COLUMNS = [
    "系统单价",
    "材料单价",
    "存货单价",
    "最新成本",
    "参考成本",
    "成本单价",
    "含税单价",
    "单价",
]

CUSTOM_COLUMNS_ORDER = [
    "版本号_子件",
    "父件编码",
    "子件编码",
    "父件名称",
    "子件名称",
    "规格型号_子件",
    "计量单位_子件",
    "需用数量",
    "系统单价",
    "生产数量_子件",
    "比例",
    "规格型号_父件",
    "创建时间",
    "标准用量",
    "版本号_父件",
    "存货图片",
    "备注",
    "子件BOM",
    "子件默认BOM",
    "损耗率%",
    "预出仓库编码",
    "预出仓库",
    "倒冲料",
    "材料倒冲方式",
    "计量单位_父件",
    "生产数量_父件",
    "生产车间编码",
    "生产车间",
    "预入仓库编码",
    "预入仓库",
    "默认BOM",
    "成品率%",
    "停用",
]

SKIP_CHILD_CODES = {"30008", "30004", "90011", "30009", "90024", "03000012", "3000011", "0300001"}
# 包装类子件按单位兜底排除：新增包装袋编码不在 SKIP_CHILD_CODES 时（如 HYD-0601 的新包装袋），
# 凡单位为"条"/"个"的子件一律不参与比例分母。
SKIP_RATIO_UNITS = {"条", "个"}
GRAM_UNITS = {"克", "g", "G", "gram", "grams"}
TEXT_COLUMNS = {
    "版本号",
    "版本号_子件",
    "版本号_父件",
    "父件编码",
    "父件编码_子件",
    "父件编码_父件",
    "子件编码",
    "父件名称",
    "子件名称",
    "规格型号",
    "规格型号_子件",
    "规格型号_父件",
    "计量单位",
    "计量单位_子件",
    "计量单位_父件",
    "存货图片",
    "备注",
    "子件BOM",
    "子件默认BOM",
    "预出仓库编码",
    "预出仓库",
    "材料倒冲方式",
    "生产车间编码",
    "生产车间",
    "预入仓库编码",
    "预入仓库",
}


@dataclass(frozen=True)
class RecipeQueryResult:
    source_path: Path
    query_text: str
    codes: list[str]
    detail: pd.DataFrame
    summary: pd.DataFrame
    default_bom: str
    include_disabled: bool

    @property
    def match_count(self) -> int:
        return len(self.detail)

    @property
    def recipe_count(self) -> int:
        if self.detail.empty:
            return 0
        return len(self.detail[["版本号_子件", "父件编码"]].drop_duplicates())

    def preview_rows(self, limit: int = 20) -> list[dict[str, object]]:
        preview = self.detail.head(max(limit, 1)).copy()
        return preview.where(pd.notna(preview), "").to_dict("records")


def normalize_cell(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_codes_text(text: str | None) -> list[str]:
    if not text:
        return []
    return [item for item in re.split(r"[\s,，;；、]+", text.strip()) if item]


def flatten_codes(values: Iterable[str] | None) -> list[str]:
    result: list[str] = []
    for value in values or []:
        result.extend(parse_codes_text(value))
    return list(dict.fromkeys(result))


def locate_recipe_source(*, include_active: bool = True) -> Path:
    configured_path = os.getenv("RECIPE_BOM_INPUT_PATH", "").strip()
    if configured_path:
        path = Path(configured_path)
        if path.is_file():
            return path
        raise FileNotFoundError(f"RECIPE_BOM_INPUT_PATH 不存在：{path}")

    if include_active:
        active_dir = Path(os.getenv("RECIPE_ACTIVE_BOM_DIR", "/app/recipe-active-bom"))
        active_files = _matching_recipe_files(active_dir, ["bom_*.xlsx", "*bom*.xlsx", "*物料清单*.xlsx"])
        if active_files:
            return max(active_files, key=lambda path: path.stat().st_mtime)

    input_dir = Path(os.getenv("RECIPE_BOM_INPUT_DIR", "/app/tplus-output/excel"))
    patterns = flatten_codes([os.getenv("RECIPE_BOM_INPUT_GLOB", "*物料清单合并*.xlsx;*bom*.xlsx;*物料清单*.xlsx")])
    files = _matching_recipe_files(input_dir, patterns)
    if not files:
        raise FileNotFoundError(f"未找到 BOM 输入文件：{input_dir} / {patterns}")
    return max(files, key=lambda path: path.stat().st_mtime)


def _matching_recipe_files(input_dir: Path, patterns: Iterable[str]) -> list[Path]:
    if not input_dir.exists():
        return []
    files: list[Path] = []
    for pattern in patterns:
        files.extend(path for path in input_dir.glob(pattern) if path.is_file() and not path.name.startswith("~$"))
    return files


def recipe_export_dir() -> Path:
    path = Path(os.getenv("RECIPE_EXPORT_DIR", "/tmp/aliecs-recipe-exports"))
    path.mkdir(parents=True, exist_ok=True)
    return path


def new_export_path() -> tuple[str, Path]:
    file_id = uuid.uuid4().hex
    return file_id, recipe_export_dir() / f"{file_id}.xlsx"


def export_path_for_id(file_id: str) -> Path:
    if not re.fullmatch(r"[a-f0-9]{32}", file_id):
        raise ValueError("invalid file id")
    return recipe_export_dir() / f"{file_id}.xlsx"


def _normalize_key(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip()


def _read_recipe_sheet(workbook: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(
        workbook,
        sheet_name=sheet_name,
        dtype={column: str for column in TEXT_COLUMNS},
    )


def merge_frames(df_component: pd.DataFrame, df_material: pd.DataFrame) -> pd.DataFrame:
    component = df_component.copy()
    material = df_material.copy()

    if "版本号" in component.columns:
        component = component.rename(columns={"版本号": "版本号_子件"})
    if "版本号" in material.columns:
        material = material.rename(columns={"版本号": "版本号_父件"})

    component["_merge_parent_code"] = _normalize_key(component["父件编码"])
    material["_merge_parent_code"] = _normalize_key(material["父件编码"])

    if "版本号_子件" in component.columns and "版本号_父件" in material.columns:
        component["_merge_version"] = _normalize_key(component["版本号_子件"])
        material["_merge_version"] = _normalize_key(material["版本号_父件"])
        merged = pd.merge(
            component,
            material,
            on=["_merge_parent_code", "_merge_version"],
            how="left",
            suffixes=("_子件", "_父件"),
        )
    else:
        merged = pd.merge(component, material, on="_merge_parent_code", how="left", suffixes=("_子件", "_父件"))

    merged = merged.drop(columns=[c for c in ["_merge_parent_code", "_merge_version"] if c in merged.columns])
    if "父件编码_子件" in merged.columns:
        merged["父件编码"] = merged["父件编码_子件"]
        merged = merged.drop(columns=[c for c in ["父件编码_子件", "父件编码_父件"] if c in merged.columns])
    return merged


def compute_ratio_series(df: pd.DataFrame) -> pd.Series:
    qty = pd.to_numeric(df.get("需用数量"), errors="coerce")
    unit = df.get("计量单位_子件").astype(str).str.strip() if "计量单位_子件" in df.columns else ""
    qty_kg = qty.where(~unit.isin(GRAM_UNITS), qty / 1000)
    child = df.get("子件编码").astype(str).str.strip() if "子件编码" in df.columns else ""
    mask = qty_kg.notna() & (~child.isin(SKIP_CHILD_CODES))
    if "计量单位_子件" in df.columns:
        mask &= ~unit.isin(SKIP_RATIO_UNITS)
    group_key = (
        df.get("版本号_子件").astype(str).str.strip().fillna("")
        .str.cat(df.get("父件编码").astype(str).str.strip().fillna(""), sep="||")
    )
    denom = qty_kg.where(mask, 0).groupby(group_key).transform("sum")
    return (qty_kg / denom).where(mask & (denom > 0))


def _dedup_items(df: pd.DataFrame) -> pd.DataFrame:
    keys = ["版本号_子件", "父件编码", "子件编码"]
    if any(key not in df.columns for key in keys):
        return df
    tmp = df[keys].astype(str).apply(lambda series: series.str.strip())
    return df.loc[~tmp.duplicated(subset=keys, keep="first")].copy()


def _detail_from_merged(merged: pd.DataFrame) -> pd.DataFrame:
    columns = [column for column in CUSTOM_COLUMNS_ORDER if column in merged.columns]
    columns.extend(column for column in PRICE_COLUMNS if column in merged.columns and column not in columns)
    aligned = merged.reindex(columns=columns).copy()
    if "比例" not in aligned.columns:
        aligned["比例"] = compute_ratio_series(aligned)
    aligned = _dedup_items(aligned)
    for column in ["版本号_子件", "父件编码", "子件编码", "父件名称", "子件名称", "规格型号_子件", "计量单位_子件", "默认BOM", "停用"]:
        if column not in aligned.columns:
            aligned[column] = ""
        aligned[column] = aligned[column].map(normalize_cell)
    aligned["需用数量"] = pd.to_numeric(aligned.get("需用数量"), errors="coerce")
    aligned["比例"] = pd.to_numeric(aligned.get("比例"), errors="coerce")
    aligned["系统单价"] = _first_numeric_column(aligned, PRICE_COLUMNS)
    return aligned.reindex(columns=OUTPUT_COLUMNS)


def _first_numeric_column(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    values = pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    for column in columns:
        if column not in df.columns:
            continue
        numeric = pd.to_numeric(df[column], errors="coerce")
        values = values.where(values.notna(), numeric)
    return pd.to_numeric(values, errors="coerce")


def load_detail_from_workbook(input_path: Path) -> pd.DataFrame:
    with pd.ExcelFile(input_path) as workbook:
        sheet_names = set(workbook.sheet_names)
        if SHEET_DETAIL in sheet_names:
            return _detail_from_merged(_read_recipe_sheet(workbook, SHEET_DETAIL))
        if {SHEET_MATERIAL, SHEET_COMPONENT}.issubset(sheet_names):
            material = _read_recipe_sheet(workbook, SHEET_MATERIAL)
            component = _read_recipe_sheet(workbook, SHEET_COMPONENT)
            return _detail_from_merged(merge_frames(component, material))
    raise ValueError(f"工作簿缺少 {SHEET_DETAIL}，或缺少 {SHEET_MATERIAL}/{SHEET_COMPONENT}。")


def _is_all_filter(value: str | None) -> bool:
    return normalize_cell(value).lower() in {"", "all", "全部", "none", "不筛选"}


def _extract_recipe_identifier(parent_name: str, parent_code: str) -> str:
    for text in [normalize_cell(parent_name), normalize_cell(parent_code)]:
        for pattern in [r"0*(\d{3,4})", r"[A-Za-z\-]*0*(\d{3,4})"]:
            match = re.search(pattern, text)
            if match:
                return match.group(1).lstrip("0")
    return ""


def _matches_query(row: pd.Series, keywords: list[str]) -> bool:
    parent_code = normalize_cell(row.get("父件编码"))
    parent_name = normalize_cell(row.get("父件名称"))
    identifier = _extract_recipe_identifier(parent_name, parent_code)
    for keyword in keywords:
        text = normalize_cell(keyword)
        if not text:
            continue
        normalized_number = text.lstrip("0")
        if text in parent_code or text in parent_name:
            return True
        if normalized_number and identifier == normalized_number:
            return True
    return False


def build_group_summary(detail: pd.DataFrame) -> pd.DataFrame:
    columns = ["版本号_子件", "父件编码", "父件名称", "明细行数", "需用数量合计", "比例合计", "停用"]
    if detail.empty:
        return pd.DataFrame(columns=columns)
    return (
        detail.groupby(["版本号_子件", "父件编码"], as_index=False, dropna=False)
        .agg(
            父件名称=("父件名称", "first"),
            明细行数=("子件编码", "size"),
            需用数量合计=("需用数量", "sum"),
            比例合计=("比例", "sum"),
            停用=("停用", "first"),
        )
        .sort_values(["停用", "版本号_子件", "父件编码"], kind="stable")
        .reindex(columns=columns)
    )


def query_recipe_workbook(
    input_path: Path,
    query_text: str,
    default_bom: str | None = "all",
    include_disabled: bool = True,
) -> RecipeQueryResult:
    detail = load_detail_from_workbook(input_path)
    if not _is_all_filter(default_bom):
        wanted = normalize_cell(default_bom)
        detail = detail[detail["默认BOM"].map(normalize_cell) == wanted]
    if not include_disabled and "停用" in detail.columns:
        detail = detail[detail["停用"].map(normalize_cell).isin({"", "0", "否", "False", "false"})]

    codes = parse_codes_text(query_text)
    keywords = codes or [query_text]
    detail = detail[detail.apply(lambda row: _matches_query(row, keywords), axis=1)].copy()
    summary = build_group_summary(detail)
    return RecipeQueryResult(
        source_path=input_path,
        query_text=query_text,
        codes=codes,
        detail=detail,
        summary=summary,
        default_bom=default_bom or "all",
        include_disabled=include_disabled,
    )


def _cost_quantity(quantity: object, unit: object) -> float:
    value = pd.to_numeric(pd.Series([quantity]), errors="coerce").iloc[0]
    if pd.isna(value):
        return 0.0
    unit_text = normalize_cell(unit)
    if unit_text in GRAM_UNITS:
        return float(value) / 1000
    return float(value)


def _float_or_zero(value: object) -> float:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return 0.0
    return float(numeric)


def _normalized_float_mapping(values: dict[str, float] | None) -> dict[str, float]:
    result: dict[str, float] = {}
    for key, value in (values or {}).items():
        numeric = float(value)
        if numeric < 0:
            raise ValueError("simulated quantities must be non-negative")
        result[normalize_cell(key)] = numeric
    return result


def _ratio_excluded(child_code: str, unit: str) -> bool:
    return child_code in SKIP_CHILD_CODES or unit in SKIP_RATIO_UNITS


def calculate_recipe_costs(
    result: RecipeQueryResult,
    manual_prices: dict[str, float] | None = None,
    simulated_quantities: dict[str, float] | None = None,
) -> list[dict[str, object]]:
    manual_prices = {normalize_cell(key): float(value) for key, value in (manual_prices or {}).items()}
    simulated_quantities = _normalized_float_mapping(simulated_quantities)
    recipes: list[dict[str, object]] = []
    if result.detail.empty:
        return recipes

    for (version, parent_code), group in result.detail.groupby(["版本号_子件", "父件编码"], sort=False, dropna=False):
        lines: list[dict[str, object]] = []
        system_total = 0.0
        current_total = 0.0
        simulated_total = 0.0
        group = group.reset_index(drop=True)
        simulated_inputs: list[tuple[str, str, float, float]] = []
        for _, row in group.iterrows():
            child_code = normalize_cell(row.get("子件编码"))
            unit = normalize_cell(row.get("计量单位_子件"))
            quantity = _float_or_zero(row.get("需用数量"))
            simulated_quantity = simulated_quantities.get(child_code, quantity)
            simulated_inputs.append((child_code, unit, simulated_quantity, _cost_quantity(simulated_quantity, unit)))
        simulated_denom = sum(cost_qty for child_code, unit, _, cost_qty in simulated_inputs if not _ratio_excluded(child_code, unit))

        for _, row in group.iterrows():
            child_code = normalize_cell(row.get("子件编码"))
            unit = normalize_cell(row.get("计量单位_子件"))
            quantity = _float_or_zero(row.get("需用数量"))
            cost_quantity = _cost_quantity(row.get("需用数量"), unit)
            system_price = _float_or_zero(row.get("系统单价"))
            manual_price = manual_prices.get(child_code)
            current_price = system_price if manual_price is None else float(manual_price)
            # 分价按「比例 × 单价」计（不是数量 × 单价）。
            ratio_value = _float_or_zero(row.get("比例"))
            system_amount = ratio_value * system_price
            current_amount = ratio_value * current_price
            simulated_quantity = simulated_quantities.get(child_code, quantity)
            simulated_ratio = 0.0
            if simulated_denom > 0 and not _ratio_excluded(child_code, unit):
                simulated_ratio = _cost_quantity(simulated_quantity, unit) / simulated_denom
            simulated_amount = simulated_ratio * current_price
            system_total += system_amount
            current_total += current_amount
            simulated_total += simulated_amount
            lines.append(
                {
                    "child_code": child_code,
                    "child_name": normalize_cell(row.get("子件名称")),
                    "spec": normalize_cell(row.get("规格型号_子件")),
                    "unit": unit,
                    "quantity": quantity,
                    "cost_quantity": cost_quantity,
                    "ratio": _float_or_zero(row.get("比例")),
                    "system_price": system_price,
                    "current_price": current_price,
                    "system_amount": system_amount,
                    "current_amount": current_amount,
                    "simulated_quantity": simulated_quantity,
                    "simulated_ratio": simulated_ratio,
                    "simulated_amount": simulated_amount,
                    "manual_price_applied": manual_price is not None,
                    "disabled": normalize_cell(row.get("停用")),
                }
            )

        first = group.iloc[0]
        recipes.append(
            {
                "key": f"{normalize_cell(parent_code)}::{normalize_cell(version)}",
                "parent_code": normalize_cell(parent_code),
                "parent_name": normalize_cell(first.get("父件名称")),
                "version": normalize_cell(version),
                "disabled": normalize_cell(first.get("停用")),
                "system_total": system_total,
                "current_total": current_total,
                "simulated_total": simulated_total,
                "lines": lines,
            }
        )
    return recipes


def _safe_sheet_title(raw: object, used: set[str]) -> str:
    title = re.sub(r"[\[\]:*?/\\]", " ", normalize_cell(raw)).strip() or "配方"
    title = re.sub(r"\s+", " ", title)[:31] or "配方"
    candidate = title
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def save_recipe_cost_workbook(output_path: Path, recipes: list[dict[str, object]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = [
        "子件编码",
        "子件名称",
        "规格型号",
        "单位",
        "数量",
        "模拟数量",
        "比例",
        "模拟比例",
        "系统单价",
        "当下价格",
        "系统分价",
        "当下分价",
        "模拟分价",
    ]
    used_titles: set[str] = set()
    for recipe in recipes:
        title = _safe_sheet_title(f"{recipe.get('parent_name') or recipe.get('parent_code')} {recipe.get('version')}", used_titles)
        ws = wb.create_sheet(title)
        ws.append(headers)
        _style_header(ws[1])
        for line in recipe.get("lines", []):
            ws.append(
                [
                    line.get("child_code"),
                    line.get("child_name"),
                    line.get("spec"),
                    line.get("unit"),
                    line.get("quantity"),
                    line.get("simulated_quantity"),
                    line.get("ratio"),
                    line.get("simulated_ratio"),
                    line.get("system_price"),
                    line.get("current_price"),
                    line.get("system_amount"),
                    line.get("current_amount"),
                    line.get("simulated_amount"),
                ]
            )
        ws.append(
            [
                "汇总",
                "",
                "",
                "",
                sum(float(line.get("quantity") or 0) for line in recipe.get("lines", [])),
                sum(float(line.get("simulated_quantity") or 0) for line in recipe.get("lines", [])),
                sum(float(line.get("ratio") or 0) for line in recipe.get("lines", [])),
                sum(float(line.get("simulated_ratio") or 0) for line in recipe.get("lines", [])),
                "",
                "",
                recipe.get("system_total"),
                recipe.get("current_total"),
                recipe.get("simulated_total"),
            ]
        )
        _style_header(ws[ws.max_row], fill="FF6B7280")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 32)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 7).number_format = "0.00%"
            ws.cell(row, 8).number_format = "0.00%"
        _apply_range_border(ws)
        ws.sheet_view.showGridLines = False

    if not recipes:
        ws = wb.create_sheet("核算结果")
        ws.append(headers)
        _style_header(ws[1])
    wb.save(output_path)


def _style_title(cell, fill: str = "FF1F2937") -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color="FFFFFFFF", size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_header(row_cells, fill: str = "FF374151") -> None:
    for cell in row_cells:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    side = Side(style="thin", color="FFD1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_range_border(ws) -> None:
    border = _thin_border()
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def _recipe_label(row: pd.Series) -> str:
    return f"{row['父件名称']} | 父件编码：{row['父件编码']} | 版本：{row['版本号_子件']}"


def _write_human_review_sheet(wb, detail: pd.DataFrame, codes: list[str]) -> None:
    ws = wb.create_sheet(HUMAN_SHEET, 0)
    ws.merge_cells("A1:I1")
    ws["A1"] = "物料清单配方表"
    _style_title(ws["A1"])
    ws.merge_cells("A2:I2")
    ws["A2"] = f"筛选编号：{'、'.join(codes) if codes else '全部'} | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].fill = PatternFill("solid", fgColor="FFE5E7EB")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    current_row = 4
    if detail.empty:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(current_row, 1).value = "没有筛选到配方"
        ws.cell(current_row, 1).alignment = Alignment(horizontal="center")
    else:
        for _, group in detail.groupby(["版本号_子件", "父件编码"], sort=False, dropna=False):
            group = group.reset_index(drop=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            ws.cell(current_row, 1).value = _recipe_label(group.iloc[0])
            ws.cell(current_row, 1).fill = PatternFill("solid", fgColor="FFE5E7EB")
            ws.cell(current_row, 1).font = Font(bold=True)
            current_row += 1
            headers = ["序号", "子件编码", "子件名称", "规格型号", "单位", "需用数量", "比例", "默认BOM", "停用"]
            for col, value in enumerate(headers, 1):
                ws.cell(current_row, col).value = value
            _style_header(ws[current_row])
            first_data_row = current_row + 1
            current_row += 1
            for idx, row in group.iterrows():
                values = [
                    idx + 1,
                    row["子件编码"],
                    row["子件名称"],
                    row["规格型号_子件"],
                    row["计量单位_子件"],
                    row["需用数量"],
                    row["比例"],
                    row["默认BOM"],
                    row["停用"],
                ]
                for col, value in enumerate(values, 1):
                    ws.cell(current_row, col).value = value
                current_row += 1
            last_data_row = current_row - 1
            ws.cell(current_row, 1).value = "合计"
            ws.cell(current_row, 6).value = f"=SUM(F{first_data_row}:F{last_data_row})"
            ws.cell(current_row, 7).value = f"=SUM(G{first_data_row}:G{last_data_row})"
            for col in range(1, 10):
                ws.cell(current_row, col).fill = PatternFill("solid", fgColor="FFFFF7ED")
                ws.cell(current_row, col).font = Font(bold=True)
            current_row += 2

    for col, width in {"A": 8, "B": 18, "C": 28, "D": 20, "E": 10, "F": 13, "G": 13, "H": 10, "I": 10}.items():
        ws.column_dimensions[col].width = width
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 6).number_format = "0.0000"
        ws.cell(row, 7).number_format = "0.00%"
    _apply_range_border(ws)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def _write_matrix_sheet(wb, detail: pd.DataFrame) -> None:
    ws = wb.create_sheet(MATRIX_SHEET, 1)
    ws.merge_cells("A1:H1")
    ws["A1"] = "配方横向对比矩阵"
    _style_title(ws["A1"])
    if detail.empty:
        ws.append(["无数据"])
        return

    data = detail.copy()
    data["配方"] = data["父件名称"].astype(str) + "\n" + data["父件编码"].astype(str) + "\n" + data["版本号_子件"].astype(str)
    recipes = list(dict.fromkeys(data["配方"].tolist()))
    headers = ["子件编码", "子件名称"]
    for recipe in recipes:
        headers.extend([f"{recipe}(需用数量)", f"{recipe}(比例)"])
    headers.append("出现配方数")
    ws.append(headers)
    _style_header(ws[2])

    for (child_code, child_name), group in data.groupby(["子件编码", "子件名称"], sort=False, dropna=False):
        row = [child_code, child_name]
        present = 0
        by_recipe = {recipe: item for recipe, item in group.groupby("配方", sort=False)}
        for recipe in recipes:
            if recipe in by_recipe:
                item = by_recipe[recipe]
                row.extend([item["需用数量"].sum(), item["比例"].sum()])
                present += 1
            else:
                row.extend(["-", "-"])
        row.append(present)
        ws.append(row)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    _apply_range_border(ws)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(3, ws.max_row + 1):
        for col in range(4, ws.max_column + 1, 2):
            ws.cell(row, col).number_format = "0.00%"
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False


def save_recipe_workbook(output_path: Path, result: RecipeQueryResult) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result.detail.to_excel(writer, sheet_name=DETAIL_SHEET, index=False)
        result.summary.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)

    wb = load_workbook(output_path)
    _write_human_review_sheet(wb, result.detail, result.codes)
    _write_matrix_sheet(wb, result.detail)
    for ws in wb.worksheets:
        if ws.title in {HUMAN_SHEET, MATRIX_SHEET}:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _style_header(ws[1])
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 35)
        _apply_range_border(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.sheet_view.showGridLines = False
    wb.save(output_path)
