"""对比表 Excel（方案B 紧凑式）真 .xlsx 渲染。

前端把对比矩阵的呈现态（版本列、行状态、每格 比例/数量/±基准）原样传来，
本模块只负责渲染，不重新解析 BOM——对比逻辑的唯一事实源仍在前端。
改用 openpyxl 真 xlsx 的动机：HTML 伪 .xls 在 Excel/WPS 只认内联样式、
在移动端查看器连内联背景都丢（2026-07-05 真机取证）；真 xlsx 全端一致，
且找回了伪 xls 做不到的冻结窗格。
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATUS_TEXT = {"replace": "替换", "add": "新增", "del": "删除", "change": "变更", "same": "一致", "history": "历史"}
STATUS_ORDER = {"replace": 0, "add": 1, "del": 2, "change": 3, "same": 4, "history": 5}
# (填充, 字色)，与网页/旧导出同一套色板
STATUS_COLOR = {
    "replace": ("FFDBEAFE", "FF1D4ED8"),
    "add": ("FFEDE9FE", "FF6D28D9"),
    "del": ("FFFEE2E2", "FFB91C1C"),
    "change": ("FFFFEDD5", "FFC2410C"),
    "same": ("FFF1F5F9", "FF64748B"),
    "history": ("FFF5EFE6", "FF7B6852"),
}

HEADER_FILL = PatternFill("solid", fgColor="FF4A3C28")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
BASE_FILL = PatternFill("solid", fgColor="FFF6EFE2")
MISS_FILL = PatternFill("solid", fgColor="FFF2F0EC")
META_FILL = PatternFill("solid", fgColor="FFFAF6EE")
THIN_BORDER = Border(*[Side(style="thin", color="FFD8CCB8")] * 4)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

UP_COLOR, DOWN_COLOR, NEW_COLOR = "FFC0392B", "FF1E8449", "FF6D28D9"  # 升红降绿，与网页一致


def _heat_color(ratio: float, max_ratio: float) -> str:
    """占比越高底色越深（近白→暖棕），非零给 12% 可见下限。"""
    p = min(1.0, max(0.0, ratio / max_ratio)) if max_ratio > 0 else 0.0
    q = 0.12 + 0.88 * p if p > 0 else 0.0
    mix = lambda a, b: round(a + (b - a) * q)
    return f"FF{mix(0xFF, 0xE3):02X}{mix(0xFB, 0xC9):02X}{mix(0xF3, 0x9E):02X}"


def _fmt_ratio(value: float) -> str:
    return f"{value * 100:.3f}%"


def _fmt_delta(value: float) -> str:
    return f"{value * 100:+.3f}%"


def _fmt_qty(value: float) -> str:
    text = f"{value:,.6f}".rstrip("0").rstrip(".")
    return text or "0"


def compare_export_filename(query_text: str | None, *, now: datetime | None = None) -> str:
    stamp = (now or datetime.now()).strftime("%Y%m%d-%H%M%S")
    label = re.sub(r'[\\/:*?"<>|\r\n\t]+', "-", str(query_text or "")).strip()
    label = re.sub(r"\s+", "", label)[:36] or "配方"
    return f"配方比例对比表_{label}_{stamp}.xlsx"


def save_compare_workbook(output_path: Path, payload: dict) -> None:
    """payload 结构见 recipes.py 的 CompareExportRequest；行序由前端给定（已按状态分组）。"""
    versions = payload["versions"]
    rows = payload["rows"]
    view = payload.get("view", {})
    show_code = bool(view.get("code", True))
    show_spec = bool(view.get("spec", True))
    show_qty = bool(view.get("qty", True))
    show_pct = bool(view.get("pct", True))
    show_arrow = bool(view.get("arrow", True))
    show_delta = bool(view.get("delta", True))
    show_new = bool(view.get("newTag", True))

    info_heads = [("状态", 8)] + ([("子件编码", 13)] if show_code else []) + [("子件名称", 26)] + ([("规格型号", 14)] if show_spec else []) + [("单位", 7)]
    n_info = len(info_heads)
    # 信息列的列号由开关算出，不能写死：少一列会让后面所有列号（含左对齐判据）整体错位。
    col_code = 2 if show_code else 0
    col_name = 3 if show_code else 2
    col_spec = col_name + 1 if show_spec else 0
    col_unit = col_name + (1 if show_spec else 0) + 1
    left_cols = {col_name} | ({col_spec} if show_spec else set())
    n_cols = n_info + len(versions)

    # 每版本列内部最大占比 → 热力分母
    max_ratio = [0.0] * len(versions)
    for row in rows:
        for idx, cell in enumerate(row["cells"]):
            if cell and cell.get("ratio") is not None:
                max_ratio[idx] = max(max_ratio[idx], float(cell["ratio"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "比例对比"
    ws.sheet_view.showGridLines = False

    def merged_row(row_idx: int, text, *, font: Font, fill: PatternFill | None = None, height: float = 18) -> None:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = font
        cell.alignment = LEFT_WRAP
        if fill:
            for col in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col).fill = fill
        ws.row_dimensions[row_idx].height = height

    merged_row(1, "配方比例对比表", font=Font(bold=True, size=16, color="FF2D2418"), height=26)

    base = next((v for v in versions if v.get("is_base")), None)
    targets = [v for v in versions if v.get("is_target")]
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta = (
        f"基准：{base['label'] if base else '-'}　|　目标：{'、'.join(t['label'] for t in targets) or '-'}"
        f"　|　导出：{stamp}　|　筛选：{payload.get('filter_label') or '全部'}"
    )
    merged_row(2, meta, font=Font(size=9, color="FF6B6258"), fill=META_FILL)

    stats: dict[str, int] = {}
    for row in rows:
        label = STATUS_TEXT.get(row["status"], row["status"])
        stats[label] = stats.get(label, 0) + 1
    chips = " · ".join(f"{k} {v}" for k, v in stats.items())
    merged_row(3, f"差异统计：{chips}　（行按状态分组排序）", font=Font(size=10, bold=True, color="FF6B6258"), fill=META_FILL)

    header_row = 4
    for col, (label, width) in enumerate(info_heads, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER_WRAP, THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    for idx, version in enumerate(versions):
        col = n_info + 1 + idx
        tag = "\n★ 基准" if version.get("is_base") else ""
        cell = ws.cell(row=header_row, column=col, value=f"{version.get('code', '')}\n{version.get('version', '') or '-'}{tag}")
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER_WRAP, THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.row_dimensions[header_row].height = 42

    small = lambda color="FF777777", bold=False: InlineFont(sz=9, color=color, b=bold)

    for r_idx, row in enumerate(rows, start=header_row + 1):
        status = row["status"]
        st_fill, st_font = STATUS_COLOR.get(status, ("FFFFFFFF", "FF333333"))
        cell = ws.cell(row=r_idx, column=1, value=STATUS_TEXT.get(status, status))
        cell.fill = PatternFill("solid", fgColor=st_fill)
        cell.font = Font(bold=True, color=st_font, size=10)

        warn = bool(row.get("code_warn"))
        if show_code:
            code_cell = ws.cell(row=r_idx, column=col_code, value=row["item_code"])
            code_cell.font = Font(name="Consolas", size=10, color="FFA83D2F" if warn else "FF1F2328", bold=warn)
        name_cell = ws.cell(row=r_idx, column=col_name, value=row["item_name"])
        # 编码列隐藏时，编码异常的红色标记挪到名称列，告警不随列一起消失（与页面一致）。
        if warn and not show_code:
            name_cell.font = Font(size=11, color="FFA83D2F", bold=True)
        if show_spec:
            spec_cell = ws.cell(row=r_idx, column=col_spec, value=row.get("spec", ""))
            spec_cell.font = Font(size=10, color="FF777777")
            spec_cell.alignment = LEFT_WRAP
        ws.cell(row=r_idx, column=col_unit, value=row.get("unit", ""))

        lines_max = 1
        for idx, version in enumerate(versions):
            col = n_info + 1 + idx
            xcell = ws.cell(row=r_idx, column=col)
            data = row["cells"][idx]
            if not data or data.get("ratio") is None:
                xcell.value = "—"
                xcell.fill = MISS_FILL
                xcell.font = Font(color="FFB8B2A8", size=10)
                continue
            ratio_v = float(data["ratio"])
            parts: list[TextBlock] = []
            if show_pct or not show_qty:
                parts.append(TextBlock(InlineFont(b=True, sz=11), _fmt_ratio(ratio_v)))
            if show_qty:
                if parts:
                    parts.append(TextBlock(small(), "\n"))
                parts.append(TextBlock(small(), _fmt_qty(float(data.get("qty") or 0))))
            if data.get("is_new") and show_new and not version.get("is_base"):
                parts.append(TextBlock(small(NEW_COLOR, bold=True), "\n新增"))
            elif data.get("delta") is not None and (show_arrow or show_delta) and not version.get("is_base"):
                delta_v = float(data["delta"])
                if abs(delta_v) > 0.000001:
                    color = UP_COLOR if delta_v > 0 else DOWN_COLOR
                    glyph = ("↑" if delta_v > 0 else "↓") if show_arrow else ""
                    num = f" {_fmt_delta(delta_v)}" if show_delta else ""
                    parts.append(TextBlock(small(color, bold=True), f"\n{glyph}{num.strip() if not glyph else num}"))
            xcell.value = CellRichText(*parts)
            lines_max = max(lines_max, 1 + sum(1 for p in parts if str(p.text).startswith("\n") or "\n" in str(p.text)))
            xcell.fill = BASE_FILL if version.get("is_base") else PatternFill("solid", fgColor=_heat_color(ratio_v, max_ratio[idx]))

        for c in range(1, n_cols + 1):
            target = ws.cell(row=r_idx, column=c)
            target.border = THIN_BORDER
            if target.alignment is None or not target.alignment.wrap_text:
                target.alignment = LEFT_WRAP if c in left_cols else CENTER_WRAP
        ws.row_dimensions[r_idx].height = 14 * max(2, lines_max) + 6

    legend_row = header_row + len(rows) + 1
    merged_row(
        legend_row,
        "说明：每格=比例/数量/相对基准变化（数量单位见「单位」列）；红↑=比基准高、绿↓=低；「—」=该配方不含此物料；"
        "底色深浅=该配方内占比高低；★基准列米色底；"
        + ("红色编码=格式与多数物料不一致。" if show_code else "红色名称=该物料编码格式与多数物料不一致。"),
        font=Font(size=8, color="FF8A8172"),
        height=24,
    )

    # 用坐标字符串而非 Cell 对象：当 rows 为空时 legend_row 与 header_row+1 重合，
    # 该单元格会落在 merged_row() 合并区内变成 MergedCell，Cell 对象赋值会在
    # freeze_panes setter 里因 isinstance 判断落空而崩溃；坐标字符串不受此影响。
    ws.freeze_panes = f"{get_column_letter(n_info + 1)}{header_row + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
